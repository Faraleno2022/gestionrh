"""
Exports génériques Excel / PDF des états « livres » SYSCOHADA
==============================================================
Chaque état fournit un titre, des en-têtes et des lignes (listes de valeurs) ;
les deux moteurs produisent un fichier téléchargeable homogène.
Contient aussi la conversion des montants en toutes lettres (chèques).
"""
import io
from decimal import Decimal

from django.http import HttpResponse


# ═══════════════════════════════════════════════════════════════════════════
# MONTANT EN TOUTES LETTRES (français — francs guinéens)
# ═══════════════════════════════════════════════════════════════════════════

_UNITES = ['', 'un', 'deux', 'trois', 'quatre', 'cinq', 'six', 'sept', 'huit', 'neuf',
           'dix', 'onze', 'douze', 'treize', 'quatorze', 'quinze', 'seize',
           'dix-sept', 'dix-huit', 'dix-neuf']
_DIZAINES = ['', 'dix', 'vingt', 'trente', 'quarante', 'cinquante',
             'soixante', 'soixante', 'quatre-vingt', 'quatre-vingt']


def _moins_de_cent(n):
    if n < 20:
        return _UNITES[n]
    d, u = divmod(n, 10)
    if d in (7, 9):          # 70-79 et 90-99 : base 60/80 + 10-19
        u += 10
        base = _DIZAINES[d]
    else:
        base = _DIZAINES[d]
    if u == 0:
        if d == 8:
            return 'quatre-vingts'
        return base
    if u == 1 and d in (2, 3, 4, 5, 6):
        return f"{base} et un"
    if u == 11 and d == 7:
        return "soixante et onze"
    return f"{base}-{_UNITES[u]}"


def _moins_de_mille(n):
    c, r = divmod(n, 100)
    if c == 0:
        return _moins_de_cent(r)
    prefixe = 'cent' if c == 1 else f"{_UNITES[c]} cent"
    if r == 0:
        return prefixe + ('s' if c > 1 else '')
    return f"{prefixe} {_moins_de_cent(r)}"


def nombre_en_lettres(n):
    """Convertit un entier positif en toutes lettres (français)."""
    n = int(n)
    if n == 0:
        return 'zéro'
    parties = []
    milliards, reste = divmod(n, 1_000_000_000)
    millions, reste = divmod(reste, 1_000_000)
    milliers, unites = divmod(reste, 1_000)
    if milliards:
        parties.append(f"{_moins_de_mille(milliards)} milliard{'s' if milliards > 1 else ''}")
    if millions:
        parties.append(f"{_moins_de_mille(millions)} million{'s' if millions > 1 else ''}")
    if milliers:
        if milliers == 1:
            parties.append('mille')
        else:
            # « cent » et « vingt » sont invariables devant « mille »
            prefixe = _moins_de_mille(milliers)
            if prefixe.endswith(('cents', 'vingts')):
                prefixe = prefixe[:-1]
            parties.append(f"{prefixe} mille")
    if unites:
        parties.append(_moins_de_mille(unites))
    return ' '.join(parties)


def montant_en_lettres(montant, devise='francs guinéens'):
    """Montant en toutes lettres pour chèques et documents officiels."""
    entier = int(Decimal(montant))
    return f"{nombre_en_lettres(entier)} {devise}".strip().capitalize()


# ═══════════════════════════════════════════════════════════════════════════
# EXPORT EXCEL (openpyxl)
# ═══════════════════════════════════════════════════════════════════════════

def reponse_excel(titre, sous_titre, entetes, lignes, nom_fichier, totaux=None):
    """Génère un fichier .xlsx : titre, sous-titre, tableau, ligne de totaux."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = (titre[:28] or 'Etat')

    gras = Font(bold=True)
    entete_fill = PatternFill(start_color='1A5276', end_color='1A5276', fill_type='solid')
    entete_font = Font(bold=True, color='FFFFFF')
    bordure = Border(*[Side(style='thin')] * 4)

    ws.cell(row=1, column=1, value=titre).font = Font(bold=True, size=14)
    ws.cell(row=2, column=1, value=sous_titre or '')
    ligne_debut = 4

    for col, entete in enumerate(entetes, start=1):
        c = ws.cell(row=ligne_debut, column=col, value=entete)
        c.font = entete_font
        c.fill = entete_fill
        c.border = bordure
        c.alignment = Alignment(horizontal='center')

    r = ligne_debut + 1
    for ligne in lignes:
        for col, valeur in enumerate(ligne, start=1):
            if isinstance(valeur, Decimal):
                valeur = float(valeur)
            c = ws.cell(row=r, column=col, value=valeur)
            c.border = bordure
            if isinstance(valeur, (int, float)):
                c.number_format = '#,##0'
        r += 1

    if totaux:
        for col, valeur in enumerate(totaux, start=1):
            if isinstance(valeur, Decimal):
                valeur = float(valeur)
            c = ws.cell(row=r, column=col, value=valeur)
            c.font = gras
            c.border = bordure
            if isinstance(valeur, (int, float)):
                c.number_format = '#,##0'

    # Largeurs de colonnes approximatives
    for col in range(1, len(entetes) + 1):
        longueur = len(str(entetes[col - 1]))
        for ligne in lignes[:50]:
            if col <= len(ligne):
                longueur = max(longueur, len(str(ligne[col - 1])))
        ws.column_dimensions[ws.cell(row=ligne_debut, column=col).column_letter].width = min(longueur + 4, 45)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{nom_fichier}.xlsx"'
    return response


# ═══════════════════════════════════════════════════════════════════════════
# EXPORT PDF (reportlab)
# ═══════════════════════════════════════════════════════════════════════════

def reponse_pdf(titre, sous_titre, entetes, lignes, nom_fichier, entreprise=None, totaux=None):
    """Génère un PDF paysage : en-tête entreprise, titre, tableau, totaux."""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4),
                            leftMargin=12 * mm, rightMargin=12 * mm,
                            topMargin=12 * mm, bottomMargin=12 * mm)
    styles = getSampleStyleSheet()
    story = []

    if entreprise is not None:
        story.append(Paragraph(str(entreprise.nom_entreprise),
                               ParagraphStyle('ent', parent=styles['Normal'], fontSize=10)))
    story.append(Paragraph(titre, ParagraphStyle('titre', parent=styles['Title'],
                                                 fontSize=15, spaceAfter=2)))
    if sous_titre:
        story.append(Paragraph(sous_titre, ParagraphStyle('st', parent=styles['Normal'],
                                                          fontSize=9, textColor=colors.grey)))
    story.append(Spacer(1, 6))

    def _fmt(v):
        if isinstance(v, Decimal):
            v = float(v)
        if isinstance(v, float) and v.is_integer():
            v = int(v)
        if isinstance(v, (int, float)):
            return f"{v:,.0f}".replace(',', ' ')
        return str(v) if v is not None else ''

    donnees = [list(entetes)] + [[_fmt(v) for v in ligne] for ligne in lignes]
    if totaux:
        donnees.append([_fmt(v) for v in totaux])

    table = Table(donnees, repeatRows=1)
    style = [
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1A5276')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 7.5),
        ('GRID', (0, 0), (-1, -1), 0.4, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F4F6F7')]),
        ('ALIGN', (1, 1), (-1, -1), 'RIGHT'),
        ('ALIGN', (0, 0), (0, -1), 'LEFT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]
    if totaux:
        style += [
            ('FONTNAME', (0, len(donnees) - 1), (-1, len(donnees) - 1), 'Helvetica-Bold'),
            ('BACKGROUND', (0, len(donnees) - 1), (-1, len(donnees) - 1), colors.HexColor('#D6EAF8')),
        ]
    table.setStyle(TableStyle(style))
    story.append(table)

    doc.build(story)
    buffer.seek(0)
    response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{nom_fichier}.pdf"'
    return response


def exporter_etat(request, titre, sous_titre, entetes, lignes, nom_fichier,
                  entreprise=None, totaux=None):
    """Retourne la réponse Excel ou PDF si ?export=excel|pdf, sinon None."""
    fmt = request.GET.get('export', '')
    if fmt == 'excel':
        return reponse_excel(titre, sous_titre, entetes, lignes, nom_fichier, totaux=totaux)
    if fmt == 'pdf':
        return reponse_pdf(titre, sous_titre, entetes, lignes, nom_fichier,
                           entreprise=entreprise, totaux=totaux)
    return None
