"""
Vues « livres et documents » — Comptabilité SYSCOHADA
======================================================
Journaux par type, balance auxiliaire, caisse (pièces, livre, situation),
reçus/quittances, bordereaux de versement/chèques, emprunts, TFT,
notes annexes, registre des immobilisations.
"""
from collections import OrderedDict
from decimal import Decimal
from datetime import date, timedelta

from django import forms
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db import transaction
from django.db.models import Sum, Q
from django.shortcuts import render, redirect, get_object_or_404
from django.utils import timezone

from core.decorators import reauth_required
from .exports_livres import exporter_etat, montant_en_lettres
from .models import (
    Journal, EcritureComptable, LigneEcriture, PlanComptable, Tiers, Facture,
    Reglement, CompteBancaire, Immobilisation, Amortissement, CessionImmobilisation,
    ExerciceComptable, RapprochementBancaire,
    PieceCaisse, BordereauRemise, LigneBordereau, Emprunt, ArreteCaisse,
    ChequeEmis, DeclarationPatente,
)


def compta_required(view_func):
    """Vérifie l'accès au module comptabilité (copie locale : le package
    comptabilite/views/ masque views.py, qui n'est pas importable ici)."""
    def wrapper(request, *args, **kwargs):
        if not request.user.is_authenticated:
            return redirect('core:login')
        if not request.user.entreprise:
            messages.error(request, "Vous devez être associé à une entreprise.")
            return redirect('core:index')
        if not request.user.entreprise.has_compta:
            messages.error(request, "Votre entreprise n'a pas accès au module comptabilité.")
            return redirect('dashboard:index')
        return view_func(request, *args, **kwargs)
    return wrapper


ZERO = Decimal('0')


def _periode(request):
    """Extrait la période (date_debut, date_fin) des paramètres GET."""
    return request.GET.get('date_debut', ''), request.GET.get('date_fin', '')


# ═══════════════════════════════════════════════════════════════════════════
# 1. JOURNAUX PAR TYPE (achats, ventes, caisse, banque, OD, salaires)
# ═══════════════════════════════════════════════════════════════════════════

TYPES_JOURNAUX_LIBELLES = {
    'AC': 'Journal des Achats',
    'VT': 'Journal des Ventes',
    'CA': 'Journal de Caisse',
    'BQ': 'Journal de Banque',
    'OD': 'Journal des Opérations Diverses',
    'SA': 'Journal des Salaires',
}


@reauth_required
@login_required
@compta_required
def journal_par_type(request, type_journal):
    """Journal comptable filtré par type SYSCOHADA (AC/VT/CA/BQ/OD/SA)."""
    type_journal = type_journal.upper()
    if type_journal not in TYPES_JOURNAUX_LIBELLES:
        messages.error(request, "Type de journal inconnu.")
        return redirect('comptabilite:journal_list')

    entreprise = request.user.entreprise
    date_debut, date_fin = _periode(request)

    ecritures = (EcritureComptable.objects
                 .filter(entreprise=entreprise, journal__type_journal=type_journal, est_validee=True)
                 .select_related('journal')
                 .prefetch_related('lignes__compte')
                 .order_by('date_ecriture', 'numero_ecriture'))
    if date_debut:
        ecritures = ecritures.filter(date_ecriture__gte=date_debut)
    if date_fin:
        ecritures = ecritures.filter(date_ecriture__lte=date_fin)

    total_debit = total_credit = ZERO
    for e in ecritures:
        for l in e.lignes.all():
            total_debit += l.montant_debit or ZERO
            total_credit += l.montant_credit or ZERO

    export = exporter_etat(
        request, TYPES_JOURNAUX_LIBELLES[type_journal],
        f"Période : {date_debut or '…'} au {date_fin or '…'}",
        ['Date', 'N° Écriture', 'Journal', 'Compte', 'Libellé', 'Débit', 'Crédit'],
        [[e.date_ecriture.strftime('%d/%m/%Y'), e.numero_ecriture, e.journal.code,
          f"{l.compte.numero_compte} {l.compte.intitule}", l.libelle or e.libelle,
          l.montant_debit or ZERO, l.montant_credit or ZERO]
         for e in ecritures for l in e.lignes.all()],
        f'journal_{type_journal}', entreprise=entreprise,
        totaux=['TOTAUX', '', '', '', '', total_debit, total_credit])
    if export:
        return export

    return render(request, 'comptabilite/livres/journal_type.html', {
        'titre': TYPES_JOURNAUX_LIBELLES[type_journal],
        'type_journal': type_journal,
        'ecritures': ecritures,
        'total_debit': total_debit,
        'total_credit': total_credit,
        'date_debut': date_debut,
        'date_fin': date_fin,
        'types_journaux': TYPES_JOURNAUX_LIBELLES,
    })


# ═══════════════════════════════════════════════════════════════════════════
# 2. BALANCE AUXILIAIRE (clients / fournisseurs)
# ═══════════════════════════════════════════════════════════════════════════

@reauth_required
@login_required
@compta_required
def balance_auxiliaire(request, categorie):
    """Balance auxiliaire clients ou fournisseurs : par tiers, facturé /
    réglé / solde, à partir des factures validées."""
    if categorie not in ('clients', 'fournisseurs'):
        return redirect('comptabilite:dashboard')

    entreprise = request.user.entreprise
    date_debut, date_fin = _periode(request)
    type_facture = 'vente' if categorie == 'clients' else 'achat'
    types_tiers = ['client', 'mixte'] if categorie == 'clients' else ['fournisseur', 'mixte']

    factures = Facture.objects.filter(
        entreprise=entreprise, type_facture=type_facture
    ).exclude(statut__in=['brouillon', 'annulee'])
    if date_debut:
        factures = factures.filter(date_facture__gte=date_debut)
    if date_fin:
        factures = factures.filter(date_facture__lte=date_fin)

    lignes = []
    total_facture = total_regle = total_solde = ZERO
    tiers_qs = Tiers.objects.filter(entreprise=entreprise, type_tiers__in=types_tiers, est_actif=True)
    for tiers in tiers_qs.order_by('raison_sociale'):
        aggr = factures.filter(tiers=tiers).aggregate(
            facture=Sum('montant_ttc'), regle=Sum('montant_paye'))
        facture_t = aggr['facture'] or ZERO
        regle_t = aggr['regle'] or ZERO
        if facture_t == 0 and regle_t == 0:
            continue
        solde = facture_t - regle_t
        lignes.append({'tiers': tiers, 'facture': facture_t, 'regle': regle_t, 'solde': solde})
        total_facture += facture_t
        total_regle += regle_t
        total_solde += solde

    export = exporter_etat(
        request, f"Balance auxiliaire {'Clients' if categorie == 'clients' else 'Fournisseurs'}",
        f"Période : {date_debut or '…'} au {date_fin or '…'}",
        ['Code', 'Tiers', 'Total facturé (TTC)', 'Total réglé', 'Solde'],
        [[l['tiers'].code, l['tiers'].raison_sociale, l['facture'], l['regle'], l['solde']]
         for l in lignes],
        f'balance_auxiliaire_{categorie}', entreprise=entreprise,
        totaux=['', 'TOTAUX', total_facture, total_regle, total_solde])
    if export:
        return export

    return render(request, 'comptabilite/livres/balance_auxiliaire.html', {
        'categorie': categorie,
        'titre': f"Balance auxiliaire {'Clients' if categorie == 'clients' else 'Fournisseurs'}",
        'lignes': lignes,
        'total_facture': total_facture,
        'total_regle': total_regle,
        'total_solde': total_solde,
        'date_debut': date_debut,
        'date_fin': date_fin,
    })


# ═══════════════════════════════════════════════════════════════════════════
# 3. CAISSE : pièces, livre, situation
# ═══════════════════════════════════════════════════════════════════════════

class PieceCaisseForm(forms.ModelForm):
    class Meta:
        model = PieceCaisse
        fields = ['type_piece', 'date_operation', 'libelle', 'tiers', 'beneficiaire',
                  'montant', 'reference', 'observation']
        widgets = {
            'type_piece': forms.Select(attrs={'class': 'form-select'}),
            'date_operation': forms.DateInput(attrs={'type': 'date', 'class': 'form-control'}),
            'libelle': forms.TextInput(attrs={'class': 'form-control'}),
            'tiers': forms.Select(attrs={'class': 'form-select'}),
            'beneficiaire': forms.TextInput(attrs={'class': 'form-control'}),
            'montant': forms.NumberInput(attrs={'class': 'form-control', 'min': '0', 'step': '1'}),
            'reference': forms.TextInput(attrs={'class': 'form-control'}),
            'observation': forms.Textarea(attrs={'class': 'form-control', 'rows': 2}),
        }

    def __init__(self, *args, entreprise=None, **kwargs):
        super().__init__(*args, **kwargs)
        if entreprise:
            self.fields['tiers'].queryset = Tiers.objects.filter(entreprise=entreprise, est_actif=True)
        self.fields['tiers'].required = False

    def clean_montant(self):
        montant = self.cleaned_data['montant']
        if montant is None or montant <= 0:
            raise forms.ValidationError("Le montant doit être supérieur à zéro.")
        return montant


@reauth_required
@login_required
@compta_required
def piece_caisse_list(request):
    """Liste des pièces de caisse (entrées/sorties)."""
    entreprise = request.user.entreprise
    date_debut, date_fin = _periode(request)
    type_piece = request.GET.get('type', '')

    pieces = PieceCaisse.objects.filter(entreprise=entreprise).select_related('tiers')
    if type_piece in ('entree', 'sortie'):
        pieces = pieces.filter(type_piece=type_piece)
    if date_debut:
        pieces = pieces.filter(date_operation__gte=date_debut)
    if date_fin:
        pieces = pieces.filter(date_operation__lte=date_fin)

    total_entrees = pieces.filter(type_piece='entree').aggregate(t=Sum('montant'))['t'] or ZERO
    total_sorties = pieces.filter(type_piece='sortie').aggregate(t=Sum('montant'))['t'] or ZERO

    export = exporter_etat(
        request, 'Pièces de caisse',
        f"Période : {date_debut or '…'} au {date_fin or '…'}",
        ['N° Pièce', 'Date', 'Type', 'Motif', 'Tiers/Bénéficiaire', 'Montant'],
        [[p.numero, p.date_operation.strftime('%d/%m/%Y'), p.get_type_piece_display(),
          p.libelle, (p.tiers.raison_sociale if p.tiers else p.beneficiaire) or '', p.montant]
         for p in pieces],
        'pieces_caisse', entreprise=entreprise,
        totaux=['TOTAUX', '', '', '',
                f'Entrées {total_entrees:,.0f} / Sorties {total_sorties:,.0f}',
                total_entrees - total_sorties])
    if export:
        return export

    return render(request, 'comptabilite/livres/piece_caisse_list.html', {
        'pieces': pieces,
        'total_entrees': total_entrees,
        'total_sorties': total_sorties,
        'solde_periode': total_entrees - total_sorties,
        'type_piece': type_piece,
        'date_debut': date_debut,
        'date_fin': date_fin,
    })


@reauth_required
@login_required
@compta_required
def piece_caisse_create(request):
    """Créer une pièce de caisse (entrée ou sortie)."""
    entreprise = request.user.entreprise
    if request.method == 'POST':
        form = PieceCaisseForm(request.POST, entreprise=entreprise)
        if form.is_valid():
            piece = form.save(commit=False)
            piece.entreprise = entreprise
            piece.cree_par = request.user
            piece.numero = PieceCaisse.prochain_numero(entreprise, piece.type_piece)
            piece.save()
            # Moteur comptable : écriture de caisse automatique
            try:
                from .moteur_comptable import comptabiliser_piece_caisse
                ecriture = comptabiliser_piece_caisse(piece, request.user)
                messages.success(request,
                                 f"Pièce de caisse {piece.numero} enregistrée. "
                                 f"Écriture {ecriture.numero_ecriture} générée automatiquement.")
            except Exception as exc:
                messages.warning(request,
                                 f"Pièce {piece.numero} enregistrée, mais l'écriture n'a pas pu "
                                 f"être générée : {exc}")
            return redirect('comptabilite:piece_caisse_print', pk=piece.pk)
    else:
        form = PieceCaisseForm(entreprise=entreprise,
                               initial={'type_piece': request.GET.get('type', 'entree')})
    return render(request, 'comptabilite/livres/piece_caisse_form.html', {'form': form})


@reauth_required
@login_required
@compta_required
def piece_caisse_print(request, pk):
    """Pièce de caisse imprimable."""
    piece = get_object_or_404(PieceCaisse, pk=pk, entreprise=request.user.entreprise)
    return render(request, 'comptabilite/livres/piece_caisse_print.html', {
        'piece': piece, 'entreprise': request.user.entreprise,
    })


@reauth_required
@login_required
@compta_required
def livre_caisse(request):
    """Livre de caisse chronologique avec solde progressif."""
    entreprise = request.user.entreprise
    date_debut, date_fin = _periode(request)

    pieces = PieceCaisse.objects.filter(entreprise=entreprise).select_related('tiers')
    # Solde initial = cumul des pièces antérieures à la période
    solde_initial = ZERO
    if date_debut:
        anterieures = PieceCaisse.objects.filter(entreprise=entreprise, date_operation__lt=date_debut)
        entrees = anterieures.filter(type_piece='entree').aggregate(t=Sum('montant'))['t'] or ZERO
        sorties = anterieures.filter(type_piece='sortie').aggregate(t=Sum('montant'))['t'] or ZERO
        solde_initial = entrees - sorties
        pieces = pieces.filter(date_operation__gte=date_debut)
    if date_fin:
        pieces = pieces.filter(date_operation__lte=date_fin)

    pieces = pieces.order_by('date_operation', 'id')
    lignes = []
    solde = solde_initial
    total_entrees = total_sorties = ZERO
    for p in pieces:
        entree = p.montant if p.type_piece == 'entree' else None
        sortie = p.montant if p.type_piece == 'sortie' else None
        solde += p.montant_signe
        total_entrees += entree or ZERO
        total_sorties += sortie or ZERO
        lignes.append({'piece': p, 'entree': entree, 'sortie': sortie, 'solde': solde})

    export = exporter_etat(
        request, 'Livre de caisse',
        f"Période : {date_debut or '…'} au {date_fin or '…'} — solde initial : {solde_initial:,.0f} GNF",
        ['Date', 'N° Pièce', 'Libellé', 'Tiers/Bénéficiaire', 'Entrées', 'Sorties', 'Solde'],
        [[l['piece'].date_operation.strftime('%d/%m/%Y'), l['piece'].numero, l['piece'].libelle,
          (l['piece'].tiers.raison_sociale if l['piece'].tiers else l['piece'].beneficiaire) or '',
          l['entree'] or '', l['sortie'] or '', l['solde']]
         for l in lignes],
        'livre_caisse', entreprise=entreprise,
        totaux=['', '', '', 'TOTAUX', total_entrees, total_sorties, solde])
    if export:
        return export

    return render(request, 'comptabilite/livres/livre_caisse.html', {
        'lignes': lignes,
        'solde_initial': solde_initial,
        'solde_final': solde,
        'total_entrees': total_entrees,
        'total_sorties': total_sorties,
        'date_debut': date_debut,
        'date_fin': date_fin,
    })


@reauth_required
@login_required
@compta_required
def situation_caisse(request):
    """Situation de caisse à une date + récapitulatif mensuel de l'année."""
    entreprise = request.user.entreprise
    a_date = request.GET.get('a_date', '') or timezone.now().date().isoformat()

    pieces = PieceCaisse.objects.filter(entreprise=entreprise, date_operation__lte=a_date)
    entrees = pieces.filter(type_piece='entree').aggregate(t=Sum('montant'))['t'] or ZERO
    sorties = pieces.filter(type_piece='sortie').aggregate(t=Sum('montant'))['t'] or ZERO
    solde = entrees - sorties

    # Récapitulatif mensuel de l'année en cours
    annee = int(a_date[:4])
    recap_mensuel = []
    solde_cumule = ZERO
    pieces_avant = PieceCaisse.objects.filter(entreprise=entreprise, date_operation__lt=date(annee, 1, 1))
    e0 = pieces_avant.filter(type_piece='entree').aggregate(t=Sum('montant'))['t'] or ZERO
    s0 = pieces_avant.filter(type_piece='sortie').aggregate(t=Sum('montant'))['t'] or ZERO
    solde_cumule = e0 - s0
    for mois in range(1, 13):
        du_mois = PieceCaisse.objects.filter(
            entreprise=entreprise, date_operation__year=annee, date_operation__month=mois)
        e = du_mois.filter(type_piece='entree').aggregate(t=Sum('montant'))['t'] or ZERO
        s = du_mois.filter(type_piece='sortie').aggregate(t=Sum('montant'))['t'] or ZERO
        solde_cumule += e - s
        recap_mensuel.append({'mois': date(annee, mois, 1), 'entrees': e, 'sorties': s,
                              'variation': e - s, 'solde': solde_cumule})

    export = exporter_etat(
        request, f'Situation de caisse {annee}',
        f"Au {a_date} — solde : {solde:,.0f} GNF",
        ['Mois', 'Entrées', 'Sorties', 'Variation', 'Solde cumulé'],
        [[m['mois'].strftime('%m/%Y'), m['entrees'], m['sorties'], m['variation'], m['solde']]
         for m in recap_mensuel],
        f'situation_caisse_{annee}', entreprise=entreprise,
        totaux=['TOTAL', entrees, sorties, entrees - sorties, solde])
    if export:
        return export

    return render(request, 'comptabilite/livres/situation_caisse.html', {
        'a_date': a_date,
        'entrees': entrees,
        'sorties': sorties,
        'solde': solde,
        'annee': annee,
        'recap_mensuel': recap_mensuel,
    })


# ═══════════════════════════════════════════════════════════════════════════
# 4. REÇU DE PAIEMENT / QUITTANCE
# ═══════════════════════════════════════════════════════════════════════════

@reauth_required
@login_required
@compta_required
def reglement_recu(request, pk):
    """Reçu de paiement / quittance imprimable pour un règlement."""
    reglement = get_object_or_404(Reglement, pk=pk, entreprise=request.user.entreprise)
    # Quittance si le règlement solde la facture, sinon simple reçu
    est_quittance = reglement.facture.montant_paye >= reglement.facture.montant_ttc
    return render(request, 'comptabilite/livres/recu_paiement.html', {
        'reglement': reglement,
        'facture': reglement.facture,
        'entreprise': request.user.entreprise,
        'est_quittance': est_quittance,
    })


# ═══════════════════════════════════════════════════════════════════════════
# 5. BORDEREAUX DE VERSEMENT / REMISE DE CHÈQUES
# ═══════════════════════════════════════════════════════════════════════════

class BordereauForm(forms.ModelForm):
    class Meta:
        model = BordereauRemise
        fields = ['type_bordereau', 'date_remise', 'compte_bancaire', 'deposant', 'observation']
        widgets = {
            'type_bordereau': forms.Select(attrs={'class': 'form-select'}),
            'date_remise': forms.DateInput(attrs={'type': 'date', 'class': 'form-control'}),
            'compte_bancaire': forms.Select(attrs={'class': 'form-select'}),
            'deposant': forms.TextInput(attrs={'class': 'form-control'}),
            'observation': forms.Textarea(attrs={'class': 'form-control', 'rows': 2}),
        }

    def __init__(self, *args, entreprise=None, **kwargs):
        super().__init__(*args, **kwargs)
        if entreprise:
            self.fields['compte_bancaire'].queryset = CompteBancaire.objects.filter(entreprise=entreprise)


@reauth_required
@login_required
@compta_required
def bordereau_list(request):
    """Liste des bordereaux de versement et de remise de chèques."""
    entreprise = request.user.entreprise
    bordereaux = (BordereauRemise.objects.filter(entreprise=entreprise)
                  .select_related('compte_bancaire').prefetch_related('lignes'))
    type_b = request.GET.get('type', '')
    if type_b in ('versement', 'cheques'):
        bordereaux = bordereaux.filter(type_bordereau=type_b)
    return render(request, 'comptabilite/livres/bordereau_list.html', {
        'bordereaux': bordereaux, 'type_b': type_b,
    })


@reauth_required
@login_required
@compta_required
def bordereau_create(request):
    """Créer un bordereau avec ses lignes (saisie dynamique)."""
    entreprise = request.user.entreprise
    if request.method == 'POST':
        form = BordereauForm(request.POST, entreprise=entreprise)
        descriptions = request.POST.getlist('ligne_description[]')
        banques = request.POST.getlist('ligne_banque[]')
        tireurs = request.POST.getlist('ligne_tireur[]')
        montants = request.POST.getlist('ligne_montant[]')
        lignes_valides = []
        for i, desc in enumerate(descriptions):
            desc = desc.strip()
            montant_str = (montants[i] if i < len(montants) else '').strip()
            if not desc and not montant_str:
                continue
            try:
                montant = Decimal(montant_str.replace(' ', '').replace(',', '.'))
            except Exception:
                montant = ZERO
            if desc and montant > 0:
                lignes_valides.append({
                    'description': desc,
                    'banque_emettrice': banques[i].strip() if i < len(banques) else '',
                    'tireur': tireurs[i].strip() if i < len(tireurs) else '',
                    'montant': montant,
                })
        if form.is_valid() and lignes_valides:
            with transaction.atomic():
                bordereau = form.save(commit=False)
                bordereau.entreprise = entreprise
                bordereau.cree_par = request.user
                bordereau.numero = BordereauRemise.prochain_numero(entreprise, bordereau.type_bordereau)
                bordereau.save()
                for l in lignes_valides:
                    LigneBordereau.objects.create(bordereau=bordereau, **l)
            messages.success(request, f"Bordereau {bordereau.numero} créé.")
            return redirect('comptabilite:bordereau_print', pk=bordereau.pk)
        if not lignes_valides:
            messages.error(request, "Ajoutez au moins une ligne avec description et montant.")
    else:
        form = BordereauForm(entreprise=entreprise,
                             initial={'type_bordereau': request.GET.get('type', 'versement')})
    return render(request, 'comptabilite/livres/bordereau_form.html', {'form': form})


@reauth_required
@login_required
@compta_required
def bordereau_print(request, pk):
    """Bordereau imprimable."""
    bordereau = get_object_or_404(
        BordereauRemise.objects.select_related('compte_bancaire').prefetch_related('lignes'),
        pk=pk, entreprise=request.user.entreprise)
    return render(request, 'comptabilite/livres/bordereau_print.html', {
        'bordereau': bordereau, 'entreprise': request.user.entreprise,
    })


# ═══════════════════════════════════════════════════════════════════════════
# 6. EMPRUNTS : situation + tableau d'amortissement
# ═══════════════════════════════════════════════════════════════════════════

class EmpruntForm(forms.ModelForm):
    class Meta:
        model = Emprunt
        fields = ['libelle', 'preteur', 'reference_contrat', 'capital_emprunte', 'taux_annuel',
                  'nombre_echeances', 'periodicite', 'date_deblocage', 'date_premiere_echeance',
                  'statut', 'observation']
        widgets = {
            'libelle': forms.TextInput(attrs={'class': 'form-control'}),
            'preteur': forms.TextInput(attrs={'class': 'form-control'}),
            'reference_contrat': forms.TextInput(attrs={'class': 'form-control'}),
            'capital_emprunte': forms.NumberInput(attrs={'class': 'form-control', 'min': '0', 'step': '1'}),
            'taux_annuel': forms.NumberInput(attrs={'class': 'form-control', 'min': '0', 'step': '0.001'}),
            'nombre_echeances': forms.NumberInput(attrs={'class': 'form-control', 'min': '1'}),
            'periodicite': forms.Select(attrs={'class': 'form-select'}),
            'date_deblocage': forms.DateInput(attrs={'type': 'date', 'class': 'form-control'}),
            'date_premiere_echeance': forms.DateInput(attrs={'type': 'date', 'class': 'form-control'}),
            'statut': forms.Select(attrs={'class': 'form-select'}),
            'observation': forms.Textarea(attrs={'class': 'form-control', 'rows': 2}),
        }


@reauth_required
@login_required
@compta_required
def emprunt_list(request):
    """Situation des emprunts : capital, CRD, échéances de l'année."""
    entreprise = request.user.entreprise
    emprunts = Emprunt.objects.filter(entreprise=entreprise)
    aujourd_hui = timezone.now().date()
    situation = []
    total_capital = total_crd = ZERO
    for e in emprunts:
        crd = e.capital_restant_du(aujourd_hui)
        situation.append({'emprunt': e, 'crd': crd})
        total_capital += e.capital_emprunte
        total_crd += crd
    export = exporter_etat(
        request, 'Situation des emprunts',
        f"Capital restant dû au {aujourd_hui.strftime('%d/%m/%Y')}",
        ['Libellé', 'Prêteur', 'Capital', 'Taux annuel (%)', 'Échéances', 'Périodicité',
         'Capital restant dû', 'Statut'],
        [[s['emprunt'].libelle, s['emprunt'].preteur, s['emprunt'].capital_emprunte,
          s['emprunt'].taux_annuel, s['emprunt'].nombre_echeances,
          s['emprunt'].get_periodicite_display(), s['crd'], s['emprunt'].get_statut_display()]
         for s in situation],
        'situation_emprunts', entreprise=entreprise,
        totaux=['TOTAUX', '', total_capital, '', '', '', total_crd, ''])
    if export:
        return export

    return render(request, 'comptabilite/livres/emprunt_list.html', {
        'situation': situation,
        'total_capital': total_capital,
        'total_crd': total_crd,
        'aujourd_hui': aujourd_hui,
    })


@reauth_required
@login_required
@compta_required
def emprunt_create(request):
    """Créer un emprunt."""
    if request.method == 'POST':
        form = EmpruntForm(request.POST)
        if form.is_valid():
            emprunt = form.save(commit=False)
            emprunt.entreprise = request.user.entreprise
            emprunt.cree_par = request.user
            emprunt.save()
            messages.success(request, f"Emprunt « {emprunt.libelle} » enregistré.")
            return redirect('comptabilite:emprunt_detail', pk=emprunt.pk)
    else:
        form = EmpruntForm()
    return render(request, 'comptabilite/livres/emprunt_form.html', {'form': form})


@reauth_required
@login_required
@compta_required
def emprunt_detail(request, pk):
    """Détail d'un emprunt avec tableau d'amortissement complet."""
    emprunt = get_object_or_404(Emprunt, pk=pk, entreprise=request.user.entreprise)
    echeancier = emprunt.echeancier()

    export = exporter_etat(
        request, f"Tableau d'amortissement — {emprunt.libelle}",
        f"{emprunt.preteur} — {emprunt.capital_emprunte:,.0f} GNF à {emprunt.taux_annuel} % "
        f"({emprunt.get_periodicite_display()}, {emprunt.nombre_echeances} échéances)",
        ['#', 'Date', 'Capital début', 'Intérêts', 'Amortissement', 'Annuité', 'Capital fin'],
        [[l['numero'], l['date'].strftime('%d/%m/%Y'), l['crd_debut'], l['interets'],
          l['amortissement'], l['annuite'], l['crd_fin']]
         for l in echeancier],
        f'emprunt_{emprunt.pk}_amortissement', entreprise=request.user.entreprise,
        totaux=['TOTAUX', '', '', sum((l['interets'] for l in echeancier), ZERO),
                emprunt.capital_emprunte, sum((l['annuite'] for l in echeancier), ZERO), ''])
    if export:
        return export

    return render(request, 'comptabilite/livres/emprunt_detail.html', {
        'emprunt': emprunt,
        'echeancier': echeancier,
        'total_interets': sum((l['interets'] for l in echeancier), ZERO),
        'total_annuites': sum((l['annuite'] for l in echeancier), ZERO),
        'crd_actuel': emprunt.capital_restant_du(),
    })


# ═══════════════════════════════════════════════════════════════════════════
# 7. TABLEAU DES FLUX DE TRÉSORERIE (SYSCOHADA, méthode indirecte simplifiée)
# ═══════════════════════════════════════════════════════════════════════════

def _soldes_par_prefixe(entreprise, prefixes, date_debut=None, date_fin=None, sens='solde'):
    """Somme (débit − crédit) des lignes validées dont le compte commence
    par l'un des préfixes. sens='debit'/'credit' pour un seul côté."""
    q = Q()
    for p in prefixes:
        q |= Q(compte__numero_compte__startswith=p)
    lignes = LigneEcriture.objects.filter(q, ecriture__entreprise=entreprise, ecriture__est_validee=True)
    if date_debut:
        lignes = lignes.filter(ecriture__date_ecriture__gte=date_debut)
    if date_fin:
        lignes = lignes.filter(ecriture__date_ecriture__lte=date_fin)
    aggr = lignes.aggregate(d=Sum('montant_debit'), c=Sum('montant_credit'))
    d, c = aggr['d'] or ZERO, aggr['c'] or ZERO
    if sens == 'debit':
        return d
    if sens == 'credit':
        return c
    return d - c


@reauth_required
@login_required
@compta_required
def tableau_flux_tresorerie(request):
    """Tableau des flux de trésorerie (méthode indirecte simplifiée SYSCOHADA)
    calculé à partir des écritures validées de la période."""
    entreprise = request.user.entreprise
    annee = int(request.GET.get('annee', timezone.now().year))
    d1, d2 = date(annee, 1, 1), date(annee, 12, 31)

    # Résultat net de la période : produits (7) − charges (6)
    produits = -_soldes_par_prefixe(entreprise, ['7'], d1, d2)   # crédit net
    charges = _soldes_par_prefixe(entreprise, ['6'], d1, d2)     # débit net
    resultat = produits - charges

    # CAFG : résultat + dotations (681/687/691/697/85) − reprises (781/787/791/797/86)
    dotations = _soldes_par_prefixe(entreprise, ['681', '687', '691', '697'], d1, d2)
    reprises = -_soldes_par_prefixe(entreprise, ['781', '787', '791', '797'], d1, d2)
    cafg = resultat + dotations - reprises

    # Variation du BFR : Δ (classe 3 + classe 4) sur la période
    var_stocks = _soldes_par_prefixe(entreprise, ['3'], d1, d2)
    var_creances_dettes = _soldes_par_prefixe(entreprise, ['4'], d1, d2)
    variation_bfr = var_stocks + var_creances_dettes
    flux_activite = cafg - variation_bfr

    # Flux d'investissement : acquisitions (débits classe 2 hors 28/29) − cessions (crédits)
    acquisitions = _soldes_par_prefixe(
        entreprise, ['20', '21', '22', '23', '24', '25', '26', '27'], d1, d2, sens='debit')
    cessions = _soldes_par_prefixe(
        entreprise, ['20', '21', '22', '23', '24', '25', '26', '27'], d1, d2, sens='credit')
    flux_investissement = -(acquisitions - cessions)

    # Flux de financement : capital (crédits 10/13) + emprunts reçus (crédits 16)
    # − remboursements (débits 16)
    apports = _soldes_par_prefixe(entreprise, ['10', '13'], d1, d2, sens='credit')
    emprunts_recus = _soldes_par_prefixe(entreprise, ['16'], d1, d2, sens='credit')
    remboursements = _soldes_par_prefixe(entreprise, ['16'], d1, d2, sens='debit')
    flux_financement = apports + emprunts_recus - remboursements

    variation_calculee = flux_activite + flux_investissement + flux_financement

    # Contrôle : variation réelle de trésorerie (classe 5) sur la période
    treso_debut = _soldes_par_prefixe(entreprise, ['5'], None, date(annee - 1, 12, 31))
    treso_fin = _soldes_par_prefixe(entreprise, ['5'], None, d2)
    variation_reelle = treso_fin - treso_debut

    export = exporter_etat(
        request, 'Tableau des flux de trésorerie',
        f"Exercice {annee} — méthode indirecte simplifiée (SYSCOHADA)",
        ['Rubrique', 'Montant (GNF)'],
        [['Résultat net de l\'exercice', resultat],
         ['+ Dotations aux amortissements et provisions', dotations],
         ['− Reprises', reprises],
         ['= CAFG', cafg],
         ['− Variation du BFR', variation_bfr],
         ['FLUX OPÉRATIONNELS (A)', flux_activite],
         ['− Acquisitions d\'immobilisations', acquisitions],
         ['+ Cessions d\'immobilisations', cessions],
         ['FLUX D\'INVESTISSEMENT (B)', flux_investissement],
         ['+ Apports en capital', apports],
         ['+ Emprunts contractés', emprunts_recus],
         ['− Remboursements d\'emprunts', remboursements],
         ['FLUX DE FINANCEMENT (C)', flux_financement],
         ['VARIATION DE TRÉSORERIE (A+B+C)', variation_calculee],
         ['Trésorerie à l\'ouverture', treso_debut],
         ['Trésorerie à la clôture', treso_fin],
         ['Variation réelle', variation_reelle],
         ['Écart de contrôle', variation_calculee - variation_reelle]],
        f'flux_tresorerie_{annee}', entreprise=entreprise)
    if export:
        return export

    return render(request, 'comptabilite/livres/flux_tresorerie.html', {
        'annee': annee,
        'resultat': resultat,
        'dotations': dotations,
        'reprises': reprises,
        'cafg': cafg,
        'variation_bfr': variation_bfr,
        'flux_activite': flux_activite,
        'acquisitions': acquisitions,
        'cessions': cessions,
        'flux_investissement': flux_investissement,
        'apports': apports,
        'emprunts_recus': emprunts_recus,
        'remboursements': remboursements,
        'flux_financement': flux_financement,
        'variation_calculee': variation_calculee,
        'treso_debut': treso_debut,
        'treso_fin': treso_fin,
        'variation_reelle': variation_reelle,
        'ecart': variation_calculee - variation_reelle,
    })


# ═══════════════════════════════════════════════════════════════════════════
# 8. NOTES ANNEXES (synthèse)
# ═══════════════════════════════════════════════════════════════════════════

@reauth_required
@login_required
@compta_required
def notes_annexes(request):
    """Notes annexes synthétiques : immobilisations, créances, dettes,
    emprunts, trésorerie, capitaux propres, effectifs."""
    entreprise = request.user.entreprise
    annee = int(request.GET.get('annee', timezone.now().year))
    fin = date(annee, 12, 31)

    # Note 1 — Immobilisations par catégorie (brut, amortissements, VNC)
    immobilisations = Immobilisation.objects.filter(entreprise=entreprise, est_actif=True)
    note_immos = []
    for code, libelle in Immobilisation.CATEGORIES:
        immos_cat = immobilisations.filter(categorie=code)
        if not immos_cat.exists():
            continue
        brut = immos_cat.aggregate(t=Sum('valeur_acquisition'))['t'] or ZERO
        amortissements = Amortissement.objects.filter(
            immobilisation__in=immos_cat).aggregate(t=Sum('montant_amortissement'))['t'] or ZERO
        note_immos.append({'categorie': libelle, 'nombre': immos_cat.count(),
                           'brut': brut, 'amortissements': amortissements,
                           'vnc': brut - amortissements})

    # Note 2 — Créances clients / Note 3 — Dettes fournisseurs
    creances = Facture.objects.filter(
        entreprise=entreprise, type_facture='vente'
    ).exclude(statut__in=['brouillon', 'annulee', 'payee']).aggregate(
        t=Sum('montant_ttc'), p=Sum('montant_paye'))
    note_creances = (creances['t'] or ZERO) - (creances['p'] or ZERO)
    dettes = Facture.objects.filter(
        entreprise=entreprise, type_facture='achat'
    ).exclude(statut__in=['brouillon', 'annulee', 'payee']).aggregate(
        t=Sum('montant_ttc'), p=Sum('montant_paye'))
    note_dettes = (dettes['t'] or ZERO) - (dettes['p'] or ZERO)

    # Note 4 — Emprunts (capital restant dû)
    note_emprunts = []
    for e in Emprunt.objects.filter(entreprise=entreprise, statut='en_cours'):
        note_emprunts.append({'emprunt': e, 'crd': e.capital_restant_du(fin)})

    # Note 5 — Trésorerie : soldes classe 5 + caisse (pièces)
    treso_comptable = _soldes_par_prefixe(entreprise, ['5'], None, fin)
    pieces = PieceCaisse.objects.filter(entreprise=entreprise, date_operation__lte=fin)
    caisse_entrees = pieces.filter(type_piece='entree').aggregate(t=Sum('montant'))['t'] or ZERO
    caisse_sorties = pieces.filter(type_piece='sortie').aggregate(t=Sum('montant'))['t'] or ZERO

    # Note 6 — Capitaux propres (classes 10-13, solde créditeur)
    capitaux = -_soldes_par_prefixe(entreprise, ['10', '11', '12', '13'], None, fin)

    # Note 7 — Effectifs
    try:
        from employes.models import Employe
        effectif = Employe.objects.filter(entreprise=entreprise, statut='actif').count()
    except Exception:
        effectif = None

    return render(request, 'comptabilite/livres/notes_annexes.html', {
        'annee': annee,
        'note_immos': note_immos,
        'note_creances': note_creances,
        'note_dettes': note_dettes,
        'note_emprunts': note_emprunts,
        'treso_comptable': treso_comptable,
        'solde_caisse': caisse_entrees - caisse_sorties,
        'capitaux': capitaux,
        'effectif': effectif,
        'entreprise': entreprise,
    })


# ═══════════════════════════════════════════════════════════════════════════
# 9. REGISTRE DES IMMOBILISATIONS (état imprimable)
# ═══════════════════════════════════════════════════════════════════════════

@reauth_required
@login_required
@compta_required
def registre_immobilisations(request):
    """Registre des immobilisations : brut, amortissements cumulés, VNC."""
    entreprise = request.user.entreprise
    immobilisations = (Immobilisation.objects.filter(entreprise=entreprise)
                       .select_related('fournisseur').order_by('categorie', 'date_acquisition'))
    lignes = []
    total_brut = total_amort = ZERO
    for immo in immobilisations:
        amort = Amortissement.objects.filter(immobilisation=immo).aggregate(
            t=Sum('montant_amortissement'))['t'] or ZERO
        lignes.append({'immo': immo, 'amortissements': amort, 'vnc': immo.valeur_acquisition - amort})
        total_brut += immo.valeur_acquisition
        total_amort += amort
    export = exporter_etat(
        request, 'Registre des immobilisations', '',
        ['N°', 'Désignation', 'Catégorie', 'Date acquisition', 'Fournisseur',
         'Durée (ans)', 'Valeur brute', 'Amort. cumulés', 'VNC'],
        [[l['immo'].numero, l['immo'].designation, l['immo'].get_categorie_display(),
          l['immo'].date_acquisition.strftime('%d/%m/%Y'),
          l['immo'].fournisseur.raison_sociale if l['immo'].fournisseur else '',
          l['immo'].duree_vie_ans, l['immo'].valeur_acquisition, l['amortissements'], l['vnc']]
         for l in lignes],
        'registre_immobilisations', entreprise=entreprise,
        totaux=['TOTAUX', '', '', '', '', '', total_brut, total_amort, total_brut - total_amort])
    if export:
        return export

    return render(request, 'comptabilite/livres/registre_immobilisations.html', {
        'lignes': lignes,
        'total_brut': total_brut,
        'total_amort': total_amort,
        'total_vnc': total_brut - total_amort,
        'entreprise': entreprise,
    })


# ═══════════════════════════════════════════════════════════════════════════
# 10. ARRÊTÉ DE CAISSE (billetage GNF)
# ═══════════════════════════════════════════════════════════════════════════

class ArreteCaisseForm(forms.ModelForm):
    class Meta:
        model = ArreteCaisse
        fields = ['date_arrete', 'nb_20000', 'nb_10000', 'nb_5000', 'nb_2000',
                  'nb_1000', 'nb_500', 'autres_valeurs', 'caissier', 'observation']
        widgets = {
            'date_arrete': forms.DateInput(attrs={'type': 'date', 'class': 'form-control'}),
            **{f'nb_{c}': forms.NumberInput(attrs={'class': 'form-control billet-qte', 'min': '0',
                                                   'data-coupure': str(c)})
               for c in ArreteCaisse.COUPURES},
            'autres_valeurs': forms.NumberInput(attrs={'class': 'form-control', 'min': '0', 'step': '1'}),
            'caissier': forms.TextInput(attrs={'class': 'form-control'}),
            'observation': forms.Textarea(attrs={'class': 'form-control', 'rows': 2}),
        }


def _solde_caisse_au(entreprise, a_date):
    """Solde théorique du livre de caisse à une date incluse."""
    pieces = PieceCaisse.objects.filter(entreprise=entreprise, date_operation__lte=a_date)
    entrees = pieces.filter(type_piece='entree').aggregate(t=Sum('montant'))['t'] or ZERO
    sorties = pieces.filter(type_piece='sortie').aggregate(t=Sum('montant'))['t'] or ZERO
    return entrees - sorties


@reauth_required
@login_required
@compta_required
def arrete_caisse_list(request):
    """Liste des arrêtés de caisse."""
    arretes = ArreteCaisse.objects.filter(entreprise=request.user.entreprise)
    return render(request, 'comptabilite/livres/arrete_caisse_list.html', {'arretes': arretes})


@reauth_required
@login_required
@compta_required
def arrete_caisse_create(request):
    """Saisir un arrêté de caisse (comptage physique par coupure)."""
    entreprise = request.user.entreprise
    if request.method == 'POST':
        form = ArreteCaisseForm(request.POST)
        if form.is_valid():
            arrete = form.save(commit=False)
            arrete.entreprise = entreprise
            arrete.cree_par = request.user
            arrete.numero = ArreteCaisse.prochain_numero(entreprise)
            arrete.solde_theorique = _solde_caisse_au(entreprise, arrete.date_arrete)
            arrete.save()
            messages.success(request, f"Arrêté de caisse {arrete.numero} enregistré "
                                      f"(écart : {arrete.ecart:,.0f} GNF).")
            return redirect('comptabilite:arrete_caisse_print', pk=arrete.pk)
    else:
        form = ArreteCaisseForm()
    solde_actuel = _solde_caisse_au(entreprise, timezone.now().date())
    return render(request, 'comptabilite/livres/arrete_caisse_form.html', {
        'form': form, 'solde_actuel': solde_actuel,
    })


@reauth_required
@login_required
@compta_required
def arrete_caisse_print(request, pk):
    """Arrêté de caisse imprimable (rapport de caisse)."""
    arrete = get_object_or_404(ArreteCaisse, pk=pk, entreprise=request.user.entreprise)
    return render(request, 'comptabilite/livres/arrete_caisse_print.html', {
        'arrete': arrete, 'entreprise': request.user.entreprise,
    })


# ═══════════════════════════════════════════════════════════════════════════
# 11. RELEVÉ DE TIERS, LETTRE DE RELANCE, ÉCHÉANCIERS
# ═══════════════════════════════════════════════════════════════════════════

TYPES_VENTE = ['vente', 'acompte']
TYPES_ACHAT = ['achat']


def _mouvements_tiers(entreprise, tiers, date_debut='', date_fin=''):
    """Relevé chronologique d'un tiers : factures (débit) et règlements
    (crédit) — avoirs en sens inverse — avec solde progressif."""
    factures = Facture.objects.filter(entreprise=entreprise, tiers=tiers).exclude(
        statut__in=['brouillon', 'annulee'])
    reglements = Reglement.objects.filter(entreprise=entreprise, facture__tiers=tiers).exclude(
        facture__statut__in=['brouillon', 'annulee'])
    if date_debut:
        factures = factures.filter(date_facture__gte=date_debut)
        reglements = reglements.filter(date_reglement__gte=date_debut)
    if date_fin:
        factures = factures.filter(date_facture__lte=date_fin)
        reglements = reglements.filter(date_reglement__lte=date_fin)

    mouvements = []
    for f in factures:
        est_avoir = f.type_facture.startswith('avoir')
        mouvements.append({
            'date': f.date_facture,
            'libelle': f"{f.get_type_facture_display()} {f.numero}",
            'debit': ZERO if est_avoir else f.montant_ttc,
            'credit': f.montant_ttc if est_avoir else ZERO,
        })
    for r in reglements:
        mouvements.append({
            'date': r.date_reglement,
            'libelle': f"Règlement {r.numero} (fact. {r.facture.numero})",
            'debit': ZERO,
            'credit': r.montant,
        })
    mouvements.sort(key=lambda m: m['date'])
    solde = ZERO
    for m in mouvements:
        solde += m['debit'] - m['credit']
        m['solde'] = solde
    return mouvements, solde


@reauth_required
@login_required
@compta_required
def releve_tiers(request, pk):
    """Relevé de compte d'un tiers (client ou fournisseur), imprimable."""
    entreprise = request.user.entreprise
    tiers = get_object_or_404(Tiers, pk=pk, entreprise=entreprise)
    date_debut, date_fin = _periode(request)
    mouvements, solde = _mouvements_tiers(entreprise, tiers, date_debut, date_fin)
    return render(request, 'comptabilite/livres/releve_tiers.html', {
        'tiers': tiers,
        'mouvements': mouvements,
        'solde': solde,
        'date_debut': date_debut,
        'date_fin': date_fin,
        'entreprise': entreprise,
    })


@reauth_required
@login_required
@compta_required
def lettre_relance(request, pk):
    """Lettre de relance imprimable pour un client : factures échues impayées."""
    entreprise = request.user.entreprise
    tiers = get_object_or_404(Tiers, pk=pk, entreprise=entreprise)
    aujourd_hui = timezone.now().date()
    factures = (Facture.objects
                .filter(entreprise=entreprise, tiers=tiers, type_facture__in=TYPES_VENTE,
                        date_echeance__lt=aujourd_hui)
                .exclude(statut__in=['brouillon', 'annulee', 'payee'])
                .order_by('date_echeance'))
    factures_data = []
    total_du = ZERO
    for f in factures:
        reste = f.reste_a_payer
        if reste <= 0:
            continue
        factures_data.append({'facture': f, 'reste': reste,
                              'retard': (aujourd_hui - f.date_echeance).days})
        total_du += reste
    return render(request, 'comptabilite/livres/lettre_relance.html', {
        'tiers': tiers,
        'factures': factures_data,
        'total_du': total_du,
        'aujourd_hui': aujourd_hui,
        'entreprise': entreprise,
    })


@reauth_required
@login_required
@compta_required
def echeancier_tiers(request, categorie):
    """Tableau des échéances clients ou fournisseurs : factures non soldées
    classées par date d'échéance (échu / à venir)."""
    if categorie not in ('clients', 'fournisseurs'):
        return redirect('comptabilite:dashboard')
    entreprise = request.user.entreprise
    types = TYPES_VENTE if categorie == 'clients' else TYPES_ACHAT
    aujourd_hui = timezone.now().date()

    factures = (Facture.objects
                .filter(entreprise=entreprise, type_facture__in=types)
                .exclude(statut__in=['brouillon', 'annulee', 'payee'])
                .select_related('tiers')
                .order_by('date_echeance', 'date_facture'))
    echues, a_venir, sans_echeance = [], [], []
    total_echu = total_a_venir = ZERO
    for f in factures:
        reste = f.reste_a_payer
        if reste <= 0:
            continue
        item = {'facture': f, 'reste': reste}
        if f.date_echeance is None:
            sans_echeance.append(item)
            total_echu += reste
        elif f.date_echeance < aujourd_hui:
            item['retard'] = (aujourd_hui - f.date_echeance).days
            echues.append(item)
            total_echu += reste
        else:
            item['dans'] = (f.date_echeance - aujourd_hui).days
            a_venir.append(item)
            total_a_venir += reste

    export = exporter_etat(
        request, f"Échéancier {'clients' if categorie == 'clients' else 'fournisseurs'}",
        f"Au {aujourd_hui.strftime('%d/%m/%Y')} — échu : {total_echu:,.0f} GNF / à échoir : {total_a_venir:,.0f} GNF",
        ['Statut', 'Tiers', 'Facture', 'Échéance', 'Retard/Dans (j)', 'Reste dû'],
        [['ÉCHU', e['facture'].tiers.raison_sociale, e['facture'].numero,
          e['facture'].date_echeance.strftime('%d/%m/%Y'), e['retard'], e['reste']]
         for e in echues] +
        [['Sans échéance', e['facture'].tiers.raison_sociale, e['facture'].numero, '', '', e['reste']]
         for e in sans_echeance] +
        [['À échoir', e['facture'].tiers.raison_sociale, e['facture'].numero,
          e['facture'].date_echeance.strftime('%d/%m/%Y'), e['dans'], e['reste']]
         for e in a_venir],
        f'echeancier_{categorie}', entreprise=entreprise,
        totaux=['TOTAL', '', '', '', '', total_echu + total_a_venir])
    if export:
        return export

    return render(request, 'comptabilite/livres/echeancier_tiers.html', {
        'categorie': categorie,
        'titre': f"Échéancier {'clients' if categorie == 'clients' else 'fournisseurs'}",
        'echues': echues,
        'a_venir': a_venir,
        'sans_echeance': sans_echeance,
        'total_echu': total_echu,
        'total_a_venir': total_a_venir,
        'aujourd_hui': aujourd_hui,
    })


# ═══════════════════════════════════════════════════════════════════════════
# 12. JOURNAL DE TVA, JOURNAL DES IMMOBILISATIONS, LIVRE DES AMORTISSEMENTS
# ═══════════════════════════════════════════════════════════════════════════

@reauth_required
@login_required
@compta_required
def journal_tva(request):
    """Journal de TVA : mouvements des comptes 443*/445* (TVA collectée,
    déductible, à payer) sur la période."""
    entreprise = request.user.entreprise
    date_debut, date_fin = _periode(request)
    lignes = (LigneEcriture.objects
              .filter(Q(compte__numero_compte__startswith='443') |
                      Q(compte__numero_compte__startswith='445'),
                      ecriture__entreprise=entreprise, ecriture__est_validee=True)
              .select_related('compte', 'ecriture', 'ecriture__journal')
              .order_by('ecriture__date_ecriture'))
    if date_debut:
        lignes = lignes.filter(ecriture__date_ecriture__gte=date_debut)
    if date_fin:
        lignes = lignes.filter(ecriture__date_ecriture__lte=date_fin)

    total_debit = total_credit = ZERO
    for l in lignes:
        total_debit += l.montant_debit or ZERO
        total_credit += l.montant_credit or ZERO

    export = exporter_etat(
        request, 'Journal de TVA',
        f"Comptes 443*/445* — période : {date_debut or '…'} au {date_fin or '…'}",
        ['Date', 'Écriture', 'Journal', 'Compte', 'Libellé', 'Débit (déductible)', 'Crédit (collectée)'],
        [[l.ecriture.date_ecriture.strftime('%d/%m/%Y'), l.ecriture.numero_ecriture,
          l.ecriture.journal.code, f"{l.compte.numero_compte} {l.compte.intitule}",
          l.libelle or l.ecriture.libelle, l.montant_debit or ZERO, l.montant_credit or ZERO]
         for l in lignes],
        'journal_tva', entreprise=entreprise,
        totaux=['TOTAUX', '', '', '', f'TVA nette : {total_credit - total_debit:,.0f}',
                total_debit, total_credit])
    if export:
        return export

    return render(request, 'comptabilite/livres/journal_tva.html', {
        'lignes': lignes,
        'total_debit': total_debit,
        'total_credit': total_credit,
        'solde_tva': total_credit - total_debit,
        'date_debut': date_debut,
        'date_fin': date_fin,
    })


@reauth_required
@login_required
@compta_required
def journal_immobilisations(request):
    """Journal des immobilisations : acquisitions et sorties chronologiques."""
    entreprise = request.user.entreprise
    annee = request.GET.get('annee', '')
    immobilisations = Immobilisation.objects.filter(entreprise=entreprise)
    cessions = CessionImmobilisation.objects.filter(immobilisation__entreprise=entreprise)
    if annee:
        immobilisations = immobilisations.filter(date_acquisition__year=annee)
        cessions = cessions.filter(date_sortie__year=annee)

    evenements = []
    for i in immobilisations:
        evenements.append({'date': i.date_acquisition, 'type': 'Acquisition', 'immo': i,
                           'montant': i.valeur_acquisition})
    for c in cessions:
        evenements.append({'date': c.date_sortie, 'type': c.get_type_sortie_display(),
                           'immo': c.immobilisation, 'montant': c.prix_vente or ZERO,
                           'resultat': c.resultat_cession})
    evenements.sort(key=lambda e: e['date'])

    export = exporter_etat(
        request, 'Journal des immobilisations',
        f"Année : {annee or 'toutes'}",
        ['Date', 'Opération', 'N° Immo', 'Désignation', 'Catégorie', 'Montant', '± Value'],
        [[e['date'].strftime('%d/%m/%Y'), e['type'], e['immo'].numero, e['immo'].designation,
          e['immo'].get_categorie_display(), e['montant'],
          e.get('resultat', '') if e['type'] != 'Acquisition' else '']
         for e in evenements],
        'journal_immobilisations', entreprise=request.user.entreprise,
        totaux=['TOTAL ACQUISITIONS', '', '', '', '',
                sum((e['montant'] for e in evenements if e['type'] == 'Acquisition'), ZERO), ''])
    if export:
        return export

    return render(request, 'comptabilite/livres/journal_immobilisations.html', {
        'evenements': evenements,
        'annee': annee,
        'total_acquisitions': sum((e['montant'] for e in evenements if e['type'] == 'Acquisition'), ZERO),
    })


@reauth_required
@login_required
@compta_required
def livre_amortissements(request):
    """Livre des amortissements : dotations par immobilisation et exercice."""
    entreprise = request.user.entreprise
    exercice_id = request.GET.get('exercice', '')
    amortissements = (Amortissement.objects
                      .filter(immobilisation__entreprise=entreprise)
                      .select_related('immobilisation', 'exercice')
                      .order_by('exercice__date_debut', 'immobilisation__numero'))
    if exercice_id:
        amortissements = amortissements.filter(exercice_id=exercice_id)
    exercices = ExerciceComptable.objects.filter(entreprise=entreprise).order_by('-date_debut')
    total_dotations = amortissements.aggregate(t=Sum('montant_amortissement'))['t'] or ZERO
    export = exporter_etat(
        request, 'Livre des amortissements', '',
        ['Exercice', 'N° Immo', 'Désignation', 'Taux (%)', 'Dotation', 'Cumul'],
        [[a.exercice.libelle, a.immobilisation.numero, a.immobilisation.designation,
          a.taux_amortissement, a.montant_amortissement, a.montant_cumule]
         for a in amortissements],
        'livre_amortissements', entreprise=entreprise,
        totaux=['TOTAL', '', '', '', total_dotations, ''])
    if export:
        return export

    return render(request, 'comptabilite/livres/livre_amortissements.html', {
        'amortissements': amortissements,
        'exercices': exercices,
        'exercice_id': exercice_id,
        'total_dotations': total_dotations,
    })


# ═══════════════════════════════════════════════════════════════════════════
# 13. ANALYSES DE GESTION (charges, produits, marges) + CAPITAUX + BANQUE
# ═══════════════════════════════════════════════════════════════════════════

def _totaux_par_compte(entreprise, prefixe, date_debut='', date_fin=''):
    """Totaux (débit, crédit, solde) par compte commençant par `prefixe`."""
    lignes = (LigneEcriture.objects
              .filter(compte__numero_compte__startswith=prefixe,
                      ecriture__entreprise=entreprise, ecriture__est_validee=True))
    if date_debut:
        lignes = lignes.filter(ecriture__date_ecriture__gte=date_debut)
    if date_fin:
        lignes = lignes.filter(ecriture__date_ecriture__lte=date_fin)
    aggr = (lignes.values('compte__numero_compte', 'compte__intitule')
            .annotate(debit=Sum('montant_debit'), credit=Sum('montant_credit'))
            .order_by('compte__numero_compte'))
    return list(aggr)


@reauth_required
@login_required
@compta_required
def analyse_charges_produits(request):
    """Analyse des charges (classe 6) et des produits (classe 7) par compte,
    avec marge globale (tableau des dépenses / recettes)."""
    entreprise = request.user.entreprise
    date_debut, date_fin = _periode(request)

    charges = _totaux_par_compte(entreprise, '6', date_debut, date_fin)
    produits = _totaux_par_compte(entreprise, '7', date_debut, date_fin)
    for c in charges:
        c['solde'] = (c['debit'] or ZERO) - (c['credit'] or ZERO)
    for p in produits:
        p['solde'] = (p['credit'] or ZERO) - (p['debit'] or ZERO)
    total_charges = sum((c['solde'] for c in charges), ZERO)
    total_produits = sum((p['solde'] for p in produits), ZERO)
    resultat = total_produits - total_charges
    marge_pct = (resultat / total_produits * 100) if total_produits else None

    export = exporter_etat(
        request, 'Analyse des charges et des produits',
        f"Période : {date_debut or '…'} au {date_fin or '…'} — résultat : {resultat:,.0f} GNF",
        ['Nature', 'Compte', 'Intitulé', 'Montant'],
        [['CHARGE', c['compte__numero_compte'], c['compte__intitule'], c['solde']] for c in charges] +
        [['PRODUIT', p['compte__numero_compte'], p['compte__intitule'], p['solde']] for p in produits],
        'analyse_charges_produits', entreprise=entreprise,
        totaux=['RÉSULTAT', '', f'Charges {total_charges:,.0f} / Produits {total_produits:,.0f}', resultat])
    if export:
        return export

    return render(request, 'comptabilite/livres/analyse_charges_produits.html', {
        'charges': charges,
        'produits': produits,
        'total_charges': total_charges,
        'total_produits': total_produits,
        'resultat': resultat,
        'marge_pct': marge_pct,
        'date_debut': date_debut,
        'date_fin': date_fin,
    })


@reauth_required
@login_required
@compta_required
def variation_capitaux_propres(request):
    """Tableau de variation des capitaux propres (classes 10-13) :
    situation d'ouverture, mouvements de l'exercice, résultat, clôture."""
    entreprise = request.user.entreprise
    annee = int(request.GET.get('annee', timezone.now().year))
    d1, d2 = date(annee, 1, 1), date(annee, 12, 31)
    fin_precedent = date(annee - 1, 12, 31)

    rubriques = [
        ('10', 'Capital'),
        ('11', 'Réserves'),
        ('12', 'Report à nouveau'),
        ('13', 'Résultat / Subventions'),
    ]
    lignes = []
    total_ouverture = total_mouvements = ZERO
    for prefixe, libelle in rubriques:
        ouverture = -_soldes_par_prefixe(entreprise, [prefixe], None, fin_precedent)
        mouvements = -_soldes_par_prefixe(entreprise, [prefixe], d1, d2)
        lignes.append({'libelle': libelle, 'ouverture': ouverture,
                       'mouvements': mouvements, 'cloture': ouverture + mouvements})
        total_ouverture += ouverture
        total_mouvements += mouvements

    # Résultat de l'exercice (classes 7 − 6) non encore affecté
    produits = -_soldes_par_prefixe(entreprise, ['7'], d1, d2)
    charges = _soldes_par_prefixe(entreprise, ['6'], d1, d2)
    resultat_exercice = produits - charges

    export = exporter_etat(
        request, 'Tableau de variation des capitaux propres',
        f"Exercice {annee}",
        ['Rubrique', 'Ouverture', 'Mouvements', 'Clôture'],
        [[l['libelle'], l['ouverture'], l['mouvements'], l['cloture']] for l in lignes] +
        [[f'Résultat de l\'exercice {annee}', '', resultat_exercice, resultat_exercice]],
        f'variation_capitaux_{annee}', entreprise=entreprise,
        totaux=['TOTAL CAPITAUX PROPRES', total_ouverture, '', total_ouverture + total_mouvements + resultat_exercice])
    if export:
        return export

    return render(request, 'comptabilite/livres/variation_capitaux.html', {
        'annee': annee,
        'lignes': lignes,
        'total_ouverture': total_ouverture,
        'total_mouvements': total_mouvements,
        'total_cloture': total_ouverture + total_mouvements,
        'resultat_exercice': resultat_exercice,
        'total_final': total_ouverture + total_mouvements + resultat_exercice,
    })


@reauth_required
@login_required
@compta_required
def situation_bancaire(request):
    """Situation bancaire : solde comptable de chaque compte bancaire
    (solde initial + mouvements du compte comptable associé) et dernier
    rapprochement."""
    entreprise = request.user.entreprise
    comptes = CompteBancaire.objects.filter(entreprise=entreprise, est_actif=True)
    situation = []
    total = ZERO
    for cb in comptes:
        mouvements = ZERO
        if cb.compte_comptable:
            aggr = (LigneEcriture.objects
                    .filter(compte=cb.compte_comptable, ecriture__entreprise=entreprise,
                            ecriture__est_validee=True)
                    .aggregate(d=Sum('montant_debit'), c=Sum('montant_credit')))
            mouvements = (aggr['d'] or ZERO) - (aggr['c'] or ZERO)
        solde = cb.solde_initial + mouvements
        dernier_rappro = (RapprochementBancaire.objects
                          .filter(compte_bancaire=cb).order_by('-date_rapprochement').first())
        situation.append({'compte': cb, 'solde': solde, 'rapprochement': dernier_rappro})
        total += solde
    export = exporter_etat(
        request, 'Situation bancaire', '',
        ['Compte', 'Banque', 'Solde comptable', 'Dernier rapprochement', 'Écart rapprochement'],
        [[f"{s['compte'].code} {s['compte'].libelle}", s['compte'].banque, s['solde'],
          s['rapprochement'].date_rapprochement.strftime('%d/%m/%Y') if s['rapprochement'] else 'Jamais',
          s['rapprochement'].ecart if s['rapprochement'] else '']
         for s in situation],
        'situation_bancaire', entreprise=entreprise,
        totaux=['TOTAL', '', total, '', ''])
    if export:
        return export

    return render(request, 'comptabilite/livres/situation_bancaire.html', {
        'situation': situation, 'total': total,
    })


# ═══════════════════════════════════════════════════════════════════════════
# 14. CHÈQUES ÉMIS (suivi + impression fac-similé)
# ═══════════════════════════════════════════════════════════════════════════

class ChequeForm(forms.ModelForm):
    class Meta:
        model = ChequeEmis
        fields = ['compte_bancaire', 'numero_cheque', 'beneficiaire', 'montant',
                  'date_emission', 'lieu_emission', 'motif', 'barre', 'statut']
        widgets = {
            'compte_bancaire': forms.Select(attrs={'class': 'form-select'}),
            'numero_cheque': forms.TextInput(attrs={'class': 'form-control'}),
            'beneficiaire': forms.TextInput(attrs={'class': 'form-control'}),
            'montant': forms.NumberInput(attrs={'class': 'form-control', 'min': '0', 'step': '1'}),
            'date_emission': forms.DateInput(attrs={'type': 'date', 'class': 'form-control'}),
            'lieu_emission': forms.TextInput(attrs={'class': 'form-control'}),
            'motif': forms.TextInput(attrs={'class': 'form-control'}),
            'barre': forms.CheckboxInput(attrs={'class': 'form-check-input'}),
            'statut': forms.Select(attrs={'class': 'form-select'}),
        }

    def __init__(self, *args, entreprise=None, **kwargs):
        super().__init__(*args, **kwargs)
        if entreprise:
            self.fields['compte_bancaire'].queryset = CompteBancaire.objects.filter(entreprise=entreprise)

    def clean_montant(self):
        montant = self.cleaned_data['montant']
        if montant is None or montant <= 0:
            raise forms.ValidationError("Le montant doit être supérieur à zéro.")
        return montant


@reauth_required
@login_required
@compta_required
def cheque_list(request):
    """Registre des chèques émis."""
    entreprise = request.user.entreprise
    cheques = ChequeEmis.objects.filter(entreprise=entreprise).select_related('compte_bancaire')
    statut = request.GET.get('statut', '')
    if statut:
        cheques = cheques.filter(statut=statut)
    total = cheques.aggregate(t=Sum('montant'))['t'] or ZERO

    export = exporter_etat(
        request, 'Registre des chèques émis', '',
        ['N° Chèque', 'Date', 'Compte', 'Bénéficiaire', 'Motif', 'Montant', 'Statut'],
        [[ch.numero_cheque, ch.date_emission.strftime('%d/%m/%Y'),
          f"{ch.compte_bancaire.code} {ch.compte_bancaire.banque}", ch.beneficiaire,
          ch.motif, ch.montant, ch.get_statut_display()]
         for ch in cheques],
        'registre_cheques', entreprise=entreprise,
        totaux=['TOTAL', '', '', '', '', total, ''])
    if export:
        return export

    return render(request, 'comptabilite/livres/cheque_list.html', {
        'cheques': cheques, 'statut': statut, 'total': total,
    })


@reauth_required
@login_required
@compta_required
def cheque_create(request):
    """Émettre un chèque."""
    entreprise = request.user.entreprise
    if request.method == 'POST':
        form = ChequeForm(request.POST, entreprise=entreprise)
        if form.is_valid():
            cheque = form.save(commit=False)
            cheque.entreprise = entreprise
            cheque.cree_par = request.user
            cheque.save()
            messages.success(request, f"Chèque {cheque.numero_cheque} enregistré.")
            return redirect('comptabilite:cheque_print', pk=cheque.pk)
    else:
        form = ChequeForm(entreprise=entreprise)
    return render(request, 'comptabilite/livres/cheque_form.html', {'form': form})


@reauth_required
@login_required
@compta_required
def cheque_print(request, pk):
    """Chèque imprimable (fac-similé) avec montant en toutes lettres."""
    cheque = get_object_or_404(ChequeEmis, pk=pk, entreprise=request.user.entreprise)
    return render(request, 'comptabilite/livres/cheque_print.html', {
        'cheque': cheque,
        'entreprise': request.user.entreprise,
        'montant_lettres': montant_en_lettres(cheque.montant),
    })


# ═══════════════════════════════════════════════════════════════════════════
# 15. DÉCLARATION DE PATENTE
# ═══════════════════════════════════════════════════════════════════════════

class PatenteForm(forms.ModelForm):
    class Meta:
        model = DeclarationPatente
        fields = ['annee', 'activite', 'reference_tarif', 'chiffre_affaires', 'valeur_locative',
                  'droit_fixe', 'taux_proportionnel', 'droit_proportionnel',
                  'date_declaration', 'statut', 'observation']
        widgets = {
            'annee': forms.NumberInput(attrs={'class': 'form-control', 'min': '2000', 'max': '2100'}),
            'activite': forms.TextInput(attrs={'class': 'form-control'}),
            'reference_tarif': forms.TextInput(attrs={'class': 'form-control'}),
            'chiffre_affaires': forms.NumberInput(attrs={'class': 'form-control', 'min': '0', 'step': '1'}),
            'valeur_locative': forms.NumberInput(attrs={'class': 'form-control', 'min': '0', 'step': '1'}),
            'droit_fixe': forms.NumberInput(attrs={'class': 'form-control', 'min': '0', 'step': '1'}),
            'taux_proportionnel': forms.NumberInput(attrs={'class': 'form-control', 'min': '0', 'step': '0.01'}),
            'droit_proportionnel': forms.NumberInput(attrs={'class': 'form-control', 'min': '0', 'step': '1'}),
            'date_declaration': forms.DateInput(attrs={'type': 'date', 'class': 'form-control'}),
            'statut': forms.Select(attrs={'class': 'form-select'}),
            'observation': forms.Textarea(attrs={'class': 'form-control', 'rows': 2}),
        }


@reauth_required
@login_required
@compta_required
def patente_list(request):
    """Liste des déclarations de patente."""
    entreprise = request.user.entreprise
    declarations = DeclarationPatente.objects.filter(entreprise=entreprise)

    export = exporter_etat(
        request, 'Déclarations de patente', '',
        ['Année', 'Activité', 'Chiffre d\'affaires', 'Valeur locative', 'Droit fixe',
         'Droit proportionnel', 'Total patente', 'Statut'],
        [[d.annee, d.activite, d.chiffre_affaires, d.valeur_locative, d.droit_fixe,
          d.droit_proportionnel, d.total_patente, d.get_statut_display()]
         for d in declarations],
        'declarations_patente', entreprise=entreprise)
    if export:
        return export

    return render(request, 'comptabilite/livres/patente_list.html', {'declarations': declarations})


@reauth_required
@login_required
@compta_required
def patente_create(request, pk=None):
    """Créer ou modifier une déclaration de patente."""
    entreprise = request.user.entreprise
    instance = get_object_or_404(DeclarationPatente, pk=pk, entreprise=entreprise) if pk else None
    if request.method == 'POST':
        form = PatenteForm(request.POST, instance=instance)
        if form.is_valid():
            declaration = form.save(commit=False)
            declaration.entreprise = entreprise
            if not declaration.cree_par:
                declaration.cree_par = request.user
            existe = (DeclarationPatente.objects
                      .filter(entreprise=entreprise, annee=declaration.annee)
                      .exclude(pk=declaration.pk).exists())
            if existe:
                messages.error(request, f"Une déclaration existe déjà pour l'année {declaration.annee}.")
            else:
                declaration.save()
                messages.success(request, f"Déclaration de patente {declaration.annee} enregistrée.")
                return redirect('comptabilite:patente_print', pk=declaration.pk)
    else:
        form = PatenteForm(instance=instance,
                           initial=None if instance else {'annee': timezone.now().year})
    return render(request, 'comptabilite/livres/patente_form.html', {'form': form, 'instance': instance})


# ═══════════════════════════════════════════════════════════════════════════
# 17. TABLEAU DE BORD COMPTABLE MÉTIER
# ═══════════════════════════════════════════════════════════════════════════

@reauth_required
@login_required
@compta_required
def tableau_bord_compta(request):
    """Tableau de bord orienté métier : indicateurs utiles, pas de comptes.
    CA, trésorerie, clients/fournisseurs, TVA, résultat, alertes."""
    entreprise = request.user.entreprise
    aujourd_hui = timezone.now().date()
    annee = aujourd_hui.year
    d1 = date(annee, 1, 1)

    # Chiffre d'affaires (classe 70, crédits nets) et résultat (7 − 6)
    ca = -_soldes_par_prefixe(entreprise, ['70'], d1, aujourd_hui)
    produits = -_soldes_par_prefixe(entreprise, ['7'], d1, aujourd_hui)
    charges = _soldes_par_prefixe(entreprise, ['6'], d1, aujourd_hui)
    resultat = produits - charges

    # Trésorerie : banque (52) et caisse (57 comptable + pièces de caisse)
    banque = _soldes_par_prefixe(entreprise, ['52'], None, aujourd_hui)
    caisse_comptable = _soldes_par_prefixe(entreprise, ['57'], None, aujourd_hui)
    pieces = PieceCaisse.objects.filter(entreprise=entreprise, date_operation__lte=aujourd_hui)
    caisse_pieces = ((pieces.filter(type_piece='entree').aggregate(t=Sum('montant'))['t'] or ZERO) -
                     (pieces.filter(type_piece='sortie').aggregate(t=Sum('montant'))['t'] or ZERO))

    # Clients débiteurs / fournisseurs créditeurs (factures non soldées)
    def _encours(type_facture):
        aggr = (Facture.objects
                .filter(entreprise=entreprise, type_facture=type_facture)
                .exclude(statut__in=['brouillon', 'annulee', 'payee'])
                .aggregate(t=Sum('montant_ttc'), p=Sum('montant_paye')))
        return (aggr['t'] or ZERO) - (aggr['p'] or ZERO)
    creances = _encours('vente')
    dettes = _encours('achat')

    # TVA à payer (collectée 443 − déductible 445)
    tva_collectee = -_soldes_par_prefixe(entreprise, ['443'], d1, aujourd_hui)
    tva_deductible = _soldes_par_prefixe(entreprise, ['445'], d1, aujourd_hui)
    tva_a_payer = tva_collectee - tva_deductible

    # Alertes métier
    alertes = []
    factures_echues = (Facture.objects
                       .filter(entreprise=entreprise, type_facture__in=['vente', 'acompte'],
                               date_echeance__lt=aujourd_hui)
                       .exclude(statut__in=['brouillon', 'annulee', 'payee']).count())
    if factures_echues:
        alertes.append({'niveau': 'danger', 'icone': 'exclamation-triangle',
                        'texte': f"{factures_echues} facture(s) client échue(s) impayée(s)",
                        'url': 'comptabilite:echeancier_tiers', 'arg': 'clients'})
    dettes_echues = (Facture.objects
                     .filter(entreprise=entreprise, type_facture='achat',
                             date_echeance__lt=aujourd_hui)
                     .exclude(statut__in=['brouillon', 'annulee', 'payee']).count())
    if dettes_echues:
        alertes.append({'niveau': 'warning', 'icone': 'clock-history',
                        'texte': f"{dettes_echues} facture(s) fournisseur échue(s) à payer",
                        'url': 'comptabilite:echeancier_tiers', 'arg': 'fournisseurs'})
    if caisse_pieces < 0:
        alertes.append({'niveau': 'danger', 'icone': 'cash-coin',
                        'texte': 'Le solde de caisse est négatif : vérifiez les pièces de caisse',
                        'url': 'comptabilite:livre_caisse', 'arg': None})
    if tva_a_payer > 0:
        alertes.append({'niveau': 'info', 'icone': 'percent',
                        'texte': f"TVA nette à reverser : {tva_a_payer:,.0f} GNF",
                        'url': 'comptabilite:journal_tva', 'arg': None})
    if not PlanComptable.objects.filter(entreprise=entreprise).exists():
        alertes.append({'niveau': 'warning', 'icone': 'stars',
                        'texte': 'Plan comptable vide : initialisez le plan SYSCOHADA',
                        'url': 'comptabilite:plan_comptable_list', 'arg': None})

    # Résultat du mois courant
    debut_mois = date(annee, aujourd_hui.month, 1)
    produits_mois = -_soldes_par_prefixe(entreprise, ['7'], debut_mois, aujourd_hui)
    charges_mois = _soldes_par_prefixe(entreprise, ['6'], debut_mois, aujourd_hui)
    resultat_mois = produits_mois - charges_mois

    # Comparaison N-1 (même période : 1er janvier → même jour)
    n1 = annee - 1
    fin_n1 = date(n1, aujourd_hui.month, min(aujourd_hui.day, 28))
    ca_n1 = -_soldes_par_prefixe(entreprise, ['70'], date(n1, 1, 1), fin_n1)
    produits_n1 = -_soldes_par_prefixe(entreprise, ['7'], date(n1, 1, 1), fin_n1)
    charges_n1 = _soldes_par_prefixe(entreprise, ['6'], date(n1, 1, 1), fin_n1)
    resultat_n1 = produits_n1 - charges_n1
    evolution_ca = ((ca - ca_n1) / ca_n1 * 100) if ca_n1 else None

    # Top 5 clients / fournisseurs de l'année (factures hors brouillon/annulée)
    def _top(type_facture):
        return (Facture.objects
                .filter(entreprise=entreprise, type_facture=type_facture,
                        date_facture__gte=d1)
                .exclude(statut__in=['brouillon', 'annulee'])
                .values('tiers__raison_sociale')
                .annotate(total=Sum('montant_ttc'))
                .order_by('-total')[:5])
    top_clients = _top('vente')
    top_fournisseurs = _top('achat')

    # Charges par nature (préfixes classe 6)
    natures = [('60', 'Achats et variations de stocks'), ('61', 'Transports'),
               ('62', 'Services extérieurs A'), ('63', 'Services extérieurs B'),
               ('64', 'Impôts et taxes'), ('65', 'Autres charges'),
               ('66', 'Charges de personnel'), ('67', 'Frais financiers'),
               ('68', 'Dotations')]
    charges_nature = []
    for prefixe, libelle in natures:
        montant = _soldes_par_prefixe(entreprise, [prefixe], d1, aujourd_hui)
        if montant:
            charges_nature.append({'libelle': libelle, 'montant': montant,
                                   'pct': (montant / charges * 100) if charges else ZERO})

    # Dernières écritures générées
    dernieres_ecritures = (EcritureComptable.objects
                           .filter(entreprise=entreprise, est_validee=True)
                           .select_related('journal')
                           .order_by('-date_ecriture', '-date_creation')[:8])

    return render(request, 'comptabilite/livres/tableau_bord.html', {
        'annee': annee,
        'ca': ca,
        'resultat': resultat,
        'resultat_mois': resultat_mois,
        'ca_n1': ca_n1,
        'resultat_n1': resultat_n1,
        'evolution_ca': evolution_ca,
        'top_clients': top_clients,
        'top_fournisseurs': top_fournisseurs,
        'charges_nature': charges_nature,
        'banque': banque,
        'caisse': caisse_comptable if caisse_comptable else caisse_pieces,
        'caisse_pieces': caisse_pieces,
        'creances': creances,
        'dettes': dettes,
        'tva_a_payer': tva_a_payer,
        'alertes': alertes,
        'dernieres_ecritures': dernieres_ecritures,
    })


@reauth_required
@login_required
@compta_required
def initialiser_plan_syscohada(request):
    """Initialise le plan comptable SYSCOHADA de l'entreprise (idempotent)."""
    if request.method == 'POST':
        from .management.commands.seed_plan_syscohada import PLAN_SYSCOHADA
        entreprise = request.user.entreprise
        crees = 0
        for numero, intitule in PLAN_SYSCOHADA:
            _, cree = PlanComptable.objects.get_or_create(
                entreprise=entreprise, numero_compte=numero,
                defaults={'intitule': intitule, 'classe': numero[0], 'est_actif': True})
            if cree:
                crees += 1
        if crees:
            messages.success(request, f"Plan SYSCOHADA initialisé : {crees} compte(s) créé(s).")
        else:
            messages.info(request, "Le plan SYSCOHADA est déjà en place : aucun compte à créer.")
    return redirect('comptabilite:plan_comptable_list')


@reauth_required
@login_required
@compta_required
def importer_plan_excel(request):
    """Importe un plan comptable depuis Excel (colonnes : numéro, intitulé).
    Les comptes existants ne sont pas modifiés."""
    if request.method == 'POST' and request.FILES.get('fichier'):
        entreprise = request.user.entreprise
        try:
            from openpyxl import load_workbook
            wb = load_workbook(request.FILES['fichier'], read_only=True, data_only=True)
            ws = wb.active
            crees = ignores = erreurs = 0
            for ligne in ws.iter_rows(min_row=1, values_only=True):
                if not ligne or ligne[0] is None:
                    continue
                numero = str(ligne[0]).strip().replace('.0', '')
                if not numero or not numero[0].isdigit():
                    continue  # en-tête ou ligne invalide
                intitule = str(ligne[1]).strip() if len(ligne) > 1 and ligne[1] else f'Compte {numero}'
                if len(numero) > 20:
                    erreurs += 1
                    continue
                _, cree = PlanComptable.objects.get_or_create(
                    entreprise=entreprise, numero_compte=numero,
                    defaults={'intitule': intitule[:200], 'classe': numero[0], 'est_actif': True})
                crees += cree
                ignores += (not cree)
            messages.success(request,
                             f"Import terminé : {crees} compte(s) créé(s), "
                             f"{ignores} déjà existant(s)"
                             f"{f', {erreurs} ligne(s) invalide(s)' if erreurs else ''}.")
        except Exception as exc:
            messages.error(request, f"Import impossible : {exc}")
    else:
        messages.error(request, "Sélectionnez un fichier Excel (.xlsx).")
    return redirect('comptabilite:plan_comptable_list')


# ═══════════════════════════════════════════════════════════════════════════
# 18. CLÔTURE DE PÉRIODE — DOTATIONS D'AMORTISSEMENT AUTOMATIQUES
# ═══════════════════════════════════════════════════════════════════════════

def _dotations_a_generer(entreprise, exercice):
    """Dotations linéaires de l'exercice par immobilisation active :
    dotation = valeur / durée, plafonnée à la VNC restante.
    Ignore les immobilisations déjà dotées pour cet exercice."""
    propositions = []
    immobilisations = Immobilisation.objects.filter(
        entreprise=entreprise, est_actif=True,
        date_acquisition__lte=exercice.date_fin)
    for immo in immobilisations:
        if Amortissement.objects.filter(immobilisation=immo, exercice=exercice).exists():
            continue
        if not immo.duree_vie_ans:
            continue
        cumul = Amortissement.objects.filter(immobilisation=immo).aggregate(
            t=Sum('montant_amortissement'))['t'] or ZERO
        vnc = immo.valeur_acquisition - cumul
        if vnc <= 0:
            continue
        dotation_annuelle = (immo.valeur_acquisition / Decimal(immo.duree_vie_ans)).quantize(Decimal('1'))
        dotation = min(dotation_annuelle, vnc)
        taux = (Decimal('100') / Decimal(immo.duree_vie_ans)).quantize(Decimal('0.01'))
        propositions.append({'immo': immo, 'dotation': dotation, 'cumul': cumul,
                             'vnc': vnc, 'taux': taux})
    return propositions


@reauth_required
@login_required
@compta_required
def cloture_periode(request):
    """Écritures automatiques de fin de période : calcule et comptabilise
    les dotations aux amortissements de l'exercice (linéaire)."""
    from .moteur_comptable import generer_ecriture, obtenir_compte, ErreurComptabilisation
    entreprise = request.user.entreprise
    exercices = ExerciceComptable.objects.filter(entreprise=entreprise).order_by('-date_debut')
    exercice_id = request.GET.get('exercice') or request.POST.get('exercice')
    exercice = (exercices.filter(pk=exercice_id).first() if exercice_id
                else exercices.filter(statut='ouvert').first())

    propositions = _dotations_a_generer(entreprise, exercice) if exercice else []

    if request.method == 'POST' and exercice and propositions:
        if exercice.statut != 'ouvert':
            messages.error(request, f"L'exercice {exercice.libelle} est clôturé.")
            return redirect('comptabilite:cloture_periode')
        compte_dotation = obtenir_compte(entreprise, '6811',
                                         'Dotations aux amortissements d\'exploitation')
        nb = 0
        total = ZERO
        for p in propositions:
            immo = p['immo']
            compte_amort = immo.compte_amortissement or obtenir_compte(
                entreprise, '2841', 'Amortissements du matériel')
            lib = f"Dotation {exercice.libelle} - {immo.numero} {immo.designation}"
            try:
                ecriture = generer_ecriture(
                    entreprise, request.user, 'OD', exercice.date_fin, lib,
                    [(compte_dotation, lib, p['dotation'], ZERO),
                     (compte_amort, lib, ZERO, p['dotation'])],
                    verifier_doublon=False)
            except ErreurComptabilisation as exc:
                messages.warning(request, f"{immo.numero} : {exc}")
                continue
            Amortissement.objects.create(
                immobilisation=immo, exercice=exercice,
                taux_amortissement=p['taux'], montant_amortissement=p['dotation'],
                montant_cumule=p['cumul'] + p['dotation'], ecriture=ecriture)
            nb += 1
            total += p['dotation']
        messages.success(request,
                         f"Clôture : {nb} dotation(s) comptabilisée(s) pour "
                         f"{total:,.0f} GNF (exercice {exercice.libelle}).")
        return redirect('comptabilite:livre_amortissements')

    return render(request, 'comptabilite/livres/cloture_periode.html', {
        'exercices': exercices,
        'exercice': exercice,
        'propositions': propositions,
        'total_dotations': sum((p['dotation'] for p in propositions), ZERO),
    })


# ═══════════════════════════════════════════════════════════════════════════
# 19. CLÔTURE D'EXERCICE ASSISTÉE (contrôles → résultat → à-nouveaux → ouverture)
# ═══════════════════════════════════════════════════════════════════════════

def _controles_pre_cloture(entreprise, exercice):
    """Contrôles bloquants et avertissements avant clôture d'exercice."""
    controles = []
    # 1. Écritures non validées sur l'exercice
    brouillons = EcritureComptable.objects.filter(
        entreprise=entreprise, exercice=exercice, est_validee=False).count()
    controles.append({
        'libelle': 'Écritures non validées', 'valeur': brouillons,
        'ok': brouillons == 0, 'bloquant': True,
        'detail': "Validez ou supprimez les brouillons avant de clôturer."})
    # 2. Journaux déséquilibrés (somme débits ≠ somme crédits par journal)
    desequilibres = []
    lignes = LigneEcriture.objects.filter(
        ecriture__entreprise=entreprise, ecriture__exercice=exercice,
        ecriture__est_validee=True)
    par_journal = (lignes.values('ecriture__journal__code')
                   .annotate(d=Sum('montant_debit'), c=Sum('montant_credit')))
    for j in par_journal:
        if (j['d'] or ZERO) != (j['c'] or ZERO):
            desequilibres.append(j['ecriture__journal__code'])
    controles.append({
        'libelle': 'Journaux déséquilibrés', 'valeur': ', '.join(desequilibres) or 0,
        'ok': not desequilibres, 'bloquant': True,
        'detail': "Chaque journal doit être équilibré (débit = crédit)."})
    # 3. Comptes d'attente (47x) non soldés
    attente = lignes.filter(compte__numero_compte__startswith='47').aggregate(
        d=Sum('montant_debit'), c=Sum('montant_credit'))
    solde_attente = (attente['d'] or ZERO) - (attente['c'] or ZERO)
    controles.append({
        'libelle': "Comptes d'attente (47x) non soldés",
        'valeur': f"{solde_attente:,.0f} GNF" if solde_attente else 0,
        'ok': solde_attente == 0, 'bloquant': False,
        'detail': "Reclassez les montants en attente vers leurs comptes définitifs."})
    # 4. Dotations d'amortissement de l'exercice
    dotations_restantes = len(_dotations_a_generer(entreprise, exercice))
    controles.append({
        'libelle': 'Dotations aux amortissements non comptabilisées',
        'valeur': dotations_restantes,
        'ok': dotations_restantes == 0, 'bloquant': False,
        'detail': "Passez par « Clôture de période » pour générer les dotations."})
    return controles


def _resultat_exercice(entreprise, exercice):
    """(produits, charges, résultat) de l'exercice sur écritures validées."""
    produits = -_soldes_par_prefixe(entreprise, ['7'], exercice.date_debut, exercice.date_fin)
    charges = _soldes_par_prefixe(entreprise, ['6'], exercice.date_debut, exercice.date_fin)
    return produits, charges, produits - charges


@reauth_required
@login_required
@compta_required
def cloture_exercice(request):
    """Clôture d'exercice assistée : contrôles, affectation du résultat,
    à-nouveaux et ouverture automatique de l'exercice suivant."""
    from .moteur_comptable import (generer_ecriture, obtenir_compte, obtenir_journal,
                                   ErreurComptabilisation)
    entreprise = request.user.entreprise
    exercices = ExerciceComptable.objects.filter(entreprise=entreprise).order_by('-date_debut')
    exercice_id = request.GET.get('exercice') or request.POST.get('exercice')
    exercice = (exercices.filter(pk=exercice_id).first() if exercice_id
                else exercices.filter(statut='ouvert').order_by('date_debut').first())

    if exercice is None:
        messages.warning(request, "Aucun exercice à clôturer.")
        return redirect('comptabilite:exercice_list')

    controles = _controles_pre_cloture(entreprise, exercice)
    bloquants = [c for c in controles if c['bloquant'] and not c['ok']]
    produits, charges, resultat = _resultat_exercice(entreprise, exercice)

    if request.method == 'POST' and exercice.statut == 'ouvert':
        if bloquants:
            messages.error(request, "Clôture impossible : corrigez d'abord les contrôles bloquants.")
            return redirect(f"{request.path}?exercice={exercice.pk}")
        try:
            with transaction.atomic():
                # ── 1. Écriture d'affectation du résultat : solder les classes 6 et 7 ──
                lignes_soldes = (LigneEcriture.objects
                                 .filter(ecriture__entreprise=entreprise, ecriture__exercice=exercice,
                                         ecriture__est_validee=True)
                                 .filter(Q(compte__numero_compte__startswith='6') |
                                         Q(compte__numero_compte__startswith='7'))
                                 .values('compte').annotate(d=Sum('montant_debit'), c=Sum('montant_credit')))
                lignes_affectation = []
                for l in lignes_soldes:
                    solde = (l['d'] or ZERO) - (l['c'] or ZERO)
                    if solde == 0:
                        continue
                    compte = PlanComptable.objects.get(pk=l['compte'])
                    if solde > 0:   # solde débiteur → on crédite pour solder
                        lignes_affectation.append((compte, 'Solde pour affectation du résultat', ZERO, solde))
                    else:           # solde créditeur → on débite pour solder
                        lignes_affectation.append((compte, 'Solde pour affectation du résultat', -solde, ZERO))
                compte_resultat = obtenir_compte(entreprise, '131' if resultat >= 0 else '139',
                                                 'Résultat net : bénéfice' if resultat >= 0 else 'Résultat net : perte')
                if resultat >= 0:
                    lignes_affectation.append((compte_resultat, f'Résultat {exercice.libelle} (bénéfice)', ZERO, resultat))
                else:
                    lignes_affectation.append((compte_resultat, f'Résultat {exercice.libelle} (perte)', -resultat, ZERO))
                ecriture_resultat = None
                if lignes_affectation and (produits or charges):
                    ecriture_resultat = generer_ecriture(
                        entreprise, request.user, 'OD', exercice.date_fin,
                        f"Affectation du résultat {exercice.libelle}",
                        lignes_affectation, verifier_doublon=False)

                # ── 2. Exercice suivant (créé/ouvert automatiquement) ──
                an_suivant = exercice.date_fin.year + 1
                exercice_suivant, _ = ExerciceComptable.objects.get_or_create(
                    entreprise=entreprise,
                    date_debut=date(an_suivant, 1, 1), date_fin=date(an_suivant, 12, 31),
                    defaults={'libelle': f'Exercice {an_suivant}', 'statut': 'ouvert'})
                if exercice_suivant.statut != 'ouvert':
                    exercice_suivant.statut = 'ouvert'
                    exercice_suivant.save(update_fields=['statut'])

                # ── 3. À-nouveaux : reprise des soldes de bilan (classes 1-5) au 1er jour ──
                soldes_bilan = (LigneEcriture.objects
                                .filter(ecriture__entreprise=entreprise, ecriture__est_validee=True,
                                        ecriture__date_ecriture__lte=exercice.date_fin)
                                .exclude(Q(compte__numero_compte__startswith='6') |
                                         Q(compte__numero_compte__startswith='7') |
                                         Q(compte__numero_compte__startswith='8'))
                                .values('compte').annotate(d=Sum('montant_debit'), c=Sum('montant_credit')))
                lignes_an = []
                for l in soldes_bilan:
                    solde = (l['d'] or ZERO) - (l['c'] or ZERO)
                    if solde == 0:
                        continue
                    compte = PlanComptable.objects.get(pk=l['compte'])
                    if solde > 0:
                        lignes_an.append((compte, 'À-nouveaux', solde, ZERO))
                    else:
                        lignes_an.append((compte, 'À-nouveaux', ZERO, -solde))
                ecriture_an = None
                if lignes_an:
                    deja_an = EcritureComptable.objects.filter(
                        entreprise=entreprise, exercice=exercice_suivant,
                        journal__type_journal='AN').exists()
                    if not deja_an:
                        ecriture_an = generer_ecriture(
                            entreprise, request.user, 'AN', exercice_suivant.date_debut,
                            f"À-nouveaux {exercice_suivant.libelle} (reprise {exercice.libelle})",
                            lignes_an, verifier_doublon=False)

                # ── 4. Clôturer l'exercice ──
                exercice.statut = 'cloture'
                exercice.est_courant = False
                exercice.save(update_fields=['statut', 'est_courant'])
                exercice_suivant.est_courant = True
                exercice_suivant.save(update_fields=['est_courant'])

            resume = [f"Exercice {exercice.libelle} clôturé."]
            if ecriture_resultat:
                resume.append(f"Affectation du résultat : {ecriture_resultat.numero_ecriture} "
                              f"({'bénéfice' if resultat >= 0 else 'perte'} de {abs(resultat):,.0f} GNF).")
            if ecriture_an:
                resume.append(f"À-nouveaux : {ecriture_an.numero_ecriture} "
                              f"({len(lignes_an)} compte(s) repris).")
            resume.append(f"Exercice {exercice_suivant.libelle} ouvert.")
            messages.success(request, ' '.join(resume))
            return redirect('comptabilite:exercice_list')
        except ErreurComptabilisation as exc:
            messages.error(request, f"Clôture interrompue : {exc}")
        except Exception as exc:
            messages.error(request, f"Erreur lors de la clôture : {exc}")

    return render(request, 'comptabilite/livres/cloture_exercice.html', {
        'exercices': exercices,
        'exercice': exercice,
        'controles': controles,
        'bloquants': bloquants,
        'produits': produits,
        'charges': charges,
        'resultat': resultat,
    })


# ═══════════════════════════════════════════════════════════════════════════
# 20. TABLEAU DE BORD FISCAL AUTOMATIQUE
# ═══════════════════════════════════════════════════════════════════════════

@reauth_required
@login_required
@compta_required
def situation_fiscale(request):
    """Situation fiscale automatique : TVA par mois, retenues, charges
    sociales, estimation d'acompte IS — calculée depuis les écritures."""
    entreprise = request.user.entreprise
    annee = int(request.GET.get('annee', timezone.now().year))

    tva_mensuelle = []
    total_collectee = total_deductible = ZERO
    for mois in range(1, 13):
        debut = date(annee, mois, 1)
        fin = date(annee, mois + 1, 1) - timedelta(days=1) if mois < 12 else date(annee, 12, 31)
        collectee = _soldes_par_prefixe(entreprise, ['443'], debut, fin, sens='credit')
        deductible = _soldes_par_prefixe(entreprise, ['445'], debut, fin, sens='debit')
        if collectee or deductible:
            tva_mensuelle.append({'mois': debut, 'collectee': collectee,
                                  'deductible': deductible, 'nette': collectee - deductible})
        total_collectee += collectee
        total_deductible += deductible

    d1, d2 = date(annee, 1, 1), date(annee, 12, 31)
    retenues = -_soldes_par_prefixe(entreprise, ['447'], d1, d2)
    charges_sociales = -_soldes_par_prefixe(entreprise, ['43'], d1, d2)
    produits, charges, resultat = (
        -_soldes_par_prefixe(entreprise, ['7'], d1, d2),
        _soldes_par_prefixe(entreprise, ['6'], d1, d2), None)
    resultat = produits - charges
    # IS Guinée : 25 % du bénéfice imposable (estimation sur résultat comptable)
    is_estime = (resultat * Decimal('0.25')).quantize(Decimal('1')) if resultat > 0 else ZERO

    return render(request, 'comptabilite/livres/situation_fiscale.html', {
        'annee': annee,
        'tva_mensuelle': tva_mensuelle,
        'total_collectee': total_collectee,
        'total_deductible': total_deductible,
        'tva_nette': total_collectee - total_deductible,
        'retenues': retenues,
        'charges_sociales': charges_sociales,
        'resultat': resultat,
        'is_estime': is_estime,
    })


# ═══════════════════════════════════════════════════════════════════════════
# 21. APPROBATIONS (moteur de validation configurable)
# ═══════════════════════════════════════════════════════════════════════════

def _niveau_acces(user):
    if user.is_superuser or getattr(user, 'est_admin_entreprise', False):
        return 5
    return getattr(getattr(user, 'profil', None), 'niveau_acces', 0) or 0


def _peut_approuver(user, demande):
    """Moteur d'autorisations : permission '<type>.approuver' (rôle ou
    délégation active) ; à défaut, niveau d'accès (compatibilité)."""
    if user.has_permission(f"{demande.type_document}.approuver", demande.entreprise):
        return True
    regle = demande.regle
    return regle is not None and _niveau_acces(user) >= regle.niveau_acces_min


def _finaliser_approbation_facture(demande, utilisateur, request):
    """Quorum atteint sur une facture : validation + comptabilisation auto."""
    from .moteur_comptable import comptabiliser_facture
    facture = Facture.objects.filter(pk=demande.objet_id,
                                     entreprise=demande.entreprise).first()
    if facture is None or facture.statut != 'brouillon':
        return
    facture.statut = 'validee'
    facture.save()
    try:
        ecriture = comptabiliser_facture(facture, utilisateur)
        messages.success(request,
                         f"Facture {facture.numero} validée et comptabilisée "
                         f"({ecriture.numero_ecriture}).")
    except Exception as exc:
        messages.warning(request, f"Facture validée, écriture non générée : {exc}")


@reauth_required
@login_required
@compta_required
def approbations(request):
    """File d'approbation : demandes en attente à traiter + historique."""
    from .models import DemandeApprobation, DecisionApprobation
    entreprise = request.user.entreprise
    niveau = _niveau_acces(request.user)

    if request.method == 'POST':
        demande = get_object_or_404(DemandeApprobation, pk=request.POST.get('demande'),
                                    entreprise=entreprise, statut='en_attente')
        decision = request.POST.get('decision')
        if decision not in ('approuve', 'rejete'):
            return redirect('comptabilite:approbations')
        if not _peut_approuver(request.user, demande):
            messages.error(request,
                           f"Autorisation insuffisante : cette demande requiert la permission "
                           f"« {demande.type_document}.approuver » (rôle chef comptable, DAF, "
                           f"DG ou administrateur) ou le niveau "
                           f"{demande.regle.niveau_acces_min if demande.regle else 4}+.")
            return redirect('comptabilite:approbations')
        if demande.demandeur_id == request.user.id and not request.user.is_superuser:
            messages.error(request, "Vous ne pouvez pas approuver votre propre demande.")
            return redirect('comptabilite:approbations')
        if demande.decisions.filter(approbateur=request.user).exists():
            messages.warning(request, "Vous avez déjà donné votre décision sur cette demande.")
            return redirect('comptabilite:approbations')

        DecisionApprobation.objects.create(
            demande=demande, approbateur=request.user, decision=decision,
            commentaire=request.POST.get('commentaire', '').strip())

        if decision == 'rejete':
            demande.statut = 'rejetee'
            demande.date_decision = timezone.now()
            demande.save(update_fields=['statut', 'date_decision'])
            messages.info(request, f"Demande rejetée : {demande.libelle}.")
        elif demande.nb_approbations_recues >= demande.nb_approbations_requises:
            demande.statut = 'approuvee'
            demande.date_decision = timezone.now()
            demande.save(update_fields=['statut', 'date_decision'])
            if demande.type_document == 'facture':
                _finaliser_approbation_facture(demande, request.user, request)
            else:
                messages.success(request, f"Demande approuvée : {demande.libelle}.")
        else:
            messages.success(request,
                             f"Approbation enregistrée "
                             f"({demande.nb_approbations_recues}/{demande.nb_approbations_requises}).")
        return redirect('comptabilite:approbations')

    en_attente = (DemandeApprobation.objects
                  .filter(entreprise=entreprise, statut='en_attente')
                  .select_related('regle', 'demandeur').prefetch_related('decisions'))
    historique = (DemandeApprobation.objects
                  .filter(entreprise=entreprise)
                  .exclude(statut='en_attente')
                  .select_related('regle', 'demandeur')[:15])
    from .models import RegleValidation
    regles = RegleValidation.objects.filter(
        Q(entreprise=entreprise) | Q(entreprise__isnull=True), est_active=True)
    return render(request, 'comptabilite/livres/approbations.html', {
        'en_attente': en_attente,
        'historique': historique,
        'regles': regles,
        'niveau': niveau,
    })


# ═══════════════════════════════════════════════════════════════════════════
# 16. ASSISTANT « NOUVELLE OPÉRATION » (moteur comptable)
# ═══════════════════════════════════════════════════════════════════════════

OPERATIONS_ASSISTANT = [
    ('vente', 'Vente (facturée à un client)', 'client'),
    ('achat', 'Achat (auprès d\'un fournisseur)', 'fournisseur'),
    ('encaissement_client', 'Encaissement client', 'client'),
    ('paiement_fournisseur', 'Paiement fournisseur', 'fournisseur'),
    ('entree_caisse', 'Entrée de caisse (recette diverse)', ''),
    ('sortie_caisse', 'Sortie de caisse (dépense diverse)', ''),
    ('salaire', 'Paiement de salaires', ''),
    ('achat_immobilisation', 'Achat d\'immobilisation', 'fournisseur'),
    ('avoir_client', 'Avoir client (retour/remise)', 'client'),
    ('avoir_fournisseur', 'Avoir fournisseur', 'fournisseur'),
]


class OperationForm(forms.Form):
    type_operation = forms.ChoiceField(
        label='Opération', choices=[(c, l) for c, l, _ in OPERATIONS_ASSISTANT],
        widget=forms.Select(attrs={'class': 'form-select'}))
    tiers = forms.ModelChoiceField(
        label='Tiers (client / fournisseur)', queryset=Tiers.objects.none(), required=False,
        widget=forms.Select(attrs={'class': 'form-select'}))
    date_operation = forms.DateField(
        label='Date', initial=timezone.now,
        widget=forms.DateInput(attrs={'type': 'date', 'class': 'form-control'}))
    libelle = forms.CharField(
        label='Libellé', max_length=200, required=False,
        widget=forms.TextInput(attrs={'class': 'form-control',
                                      'placeholder': 'Ex. : Vente de marchandises au comptant'}))
    montant_ht = forms.DecimalField(
        label='Montant HT (GNF)', min_value=Decimal('1'),
        widget=forms.NumberInput(attrs={'class': 'form-control', 'step': '1'}))
    taux_tva = forms.DecimalField(
        label='TVA (%)', initial=Decimal('18'), min_value=ZERO, max_value=Decimal('100'),
        required=False,
        widget=forms.NumberInput(attrs={'class': 'form-control', 'step': '0.01'}))
    mode_paiement = forms.ChoiceField(
        label='Règlement par', choices=[('banque', 'Banque'), ('especes', 'Caisse (espèces)')],
        widget=forms.Select(attrs={'class': 'form-select'}))
    centre_analyse = forms.ModelChoiceField(
        label='Dimension analytique (projet, agence, centre de coût…)',
        queryset=None, required=False,
        widget=forms.Select(attrs={'class': 'form-select'}))
    piece_jointe = forms.FileField(
        label='Pièce justificative (PDF, scan, photo…)', required=False,
        widget=forms.ClearableFileInput(attrs={'class': 'form-control'}))

    def __init__(self, *args, entreprise=None, **kwargs):
        super().__init__(*args, **kwargs)
        from .models import CentreAnalyse
        if entreprise:
            self.fields['tiers'].queryset = Tiers.objects.filter(
                entreprise=entreprise, est_actif=True)
            self.fields['centre_analyse'].queryset = CentreAnalyse.objects.filter(
                entreprise=entreprise, est_actif=True)
        else:
            self.fields['centre_analyse'].queryset = CentreAnalyse.objects.none()


@reauth_required
@login_required
@compta_required
def nouvelle_operation(request):
    """Assistant : l'utilisateur saisit une opération métier, le moteur
    comptable génère l'écriture — aucun compte à connaître."""
    from .moteur_comptable import operation_simple, ErreurComptabilisation
    entreprise = request.user.entreprise
    if request.method == 'POST':
        form = OperationForm(request.POST, request.FILES, entreprise=entreprise)
        if form.is_valid():
            d = form.cleaned_data
            try:
                ecriture = operation_simple(
                    entreprise, request.user, d['type_operation'], d['date_operation'],
                    d['libelle'], d['montant_ht'], d['taux_tva'] or ZERO,
                    tiers=d['tiers'], mode_paiement=d['mode_paiement'],
                    centre_analyse=d['centre_analyse'], piece_jointe=d['piece_jointe'])
            except ErreurComptabilisation as exc:
                messages.error(request, str(exc))
            else:
                messages.success(request,
                                 f"Opération enregistrée : écriture {ecriture.numero_ecriture} "
                                 f"générée et validée automatiquement.")
                return redirect('comptabilite:ecriture_detail', pk=ecriture.pk)
    else:
        form = OperationForm(entreprise=entreprise,
                             initial={'type_operation': request.GET.get('type', 'vente')})
    return render(request, 'comptabilite/livres/nouvelle_operation.html', {
        'form': form,
        'operations': OPERATIONS_ASSISTANT,
    })


@reauth_required
@login_required
@compta_required
def patente_print(request, pk):
    """Déclaration de patente imprimable."""
    declaration = get_object_or_404(DeclarationPatente, pk=pk, entreprise=request.user.entreprise)
    return render(request, 'comptabilite/livres/patente_print.html', {
        'declaration': declaration,
        'entreprise': request.user.entreprise,
        'total_lettres': montant_en_lettres(declaration.total_patente),
    })
