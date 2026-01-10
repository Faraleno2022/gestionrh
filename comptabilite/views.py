from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db.models import Sum, Q, F, Count
from django.utils import timezone
from django.http import JsonResponse, HttpResponse
from django.core.paginator import Paginator
from decimal import Decimal
from collections import OrderedDict
import io

from .models import (
    PlanComptable, Journal, ExerciceComptable, EcritureComptable,
    LigneEcriture, Tiers, Facture, LigneFacture, Reglement, TauxTVA, PieceComptable
)
from .forms import (
    PlanComptableForm, JournalForm, ExerciceForm, EcritureForm,
    TiersForm, FactureForm, ReglementForm
)


def compta_required(view_func):
    """Décorateur pour vérifier l'accès au module comptabilité"""
    def wrapper(request, *args, **kwargs):
        if not request.user.is_authenticated:
            return redirect('core:login')
        if not request.user.entreprise:
            messages.error(request, "Vous devez être associé à une entreprise.")
            return redirect('core:index')
        if not request.user.entreprise.has_compta:
            messages.error(request, "Votre entreprise n'a pas accès au module comptabilité.")
            return redirect('dashboard:index')
        return view_func(request, *args, **kwargs)
    return wrapper


@login_required
@compta_required
def dashboard(request):
    """Dashboard principal comptabilité"""
    entreprise = request.user.entreprise
    
    # Statistiques
    exercice_courant = ExerciceComptable.objects.filter(
        entreprise=entreprise, est_courant=True
    ).first()
    
    stats = {
        'nb_ecritures': EcritureComptable.objects.filter(entreprise=entreprise).count(),
        'nb_factures_impayees': Facture.objects.filter(
            entreprise=entreprise, statut__in=['validee']
        ).count(),
        'nb_tiers': Tiers.objects.filter(entreprise=entreprise, est_actif=True).count(),
        'nb_comptes': PlanComptable.objects.filter(entreprise=entreprise, est_actif=True).count(),
    }
    
    # Dernières écritures
    dernieres_ecritures = EcritureComptable.objects.filter(
        entreprise=entreprise
    ).select_related('journal').order_by('-date_creation')[:5]
    
    # Factures en attente
    factures_en_attente = Facture.objects.filter(
        entreprise=entreprise, statut='validee'
    ).select_related('tiers').order_by('date_echeance')[:5]
    
    context = {
        'stats': stats,
        'exercice_courant': exercice_courant,
        'dernieres_ecritures': dernieres_ecritures,
        'factures_en_attente': factures_en_attente,
    }
    return render(request, 'comptabilite/dashboard.html', context)


# ==================== PLAN COMPTABLE ====================

@login_required
@compta_required
def plan_comptable_list(request):
    """Liste du plan comptable"""
    entreprise = request.user.entreprise
    classe = request.GET.get('classe', '')
    search = request.GET.get('q', '')
    
    comptes = PlanComptable.objects.filter(entreprise=entreprise)
    
    if classe:
        comptes = comptes.filter(classe=classe)
    if search:
        comptes = comptes.filter(
            Q(numero_compte__icontains=search) | Q(intitule__icontains=search)
        )
    
    comptes = comptes.order_by('numero_compte')
    
    paginator = Paginator(comptes, 50)
    page = request.GET.get('page', 1)
    comptes = paginator.get_page(page)
    
    context = {
        'comptes': comptes,
        'classes': PlanComptable.CLASSES,
        'classe_selectionnee': classe,
        'search': search,
    }
    return render(request, 'comptabilite/plan_comptable/list.html', context)


@login_required
@compta_required
def plan_comptable_create(request):
    """Créer un compte comptable"""
    if request.method == 'POST':
        form = PlanComptableForm(request.POST, entreprise=request.user.entreprise)
        if form.is_valid():
            compte = form.save(commit=False)
            compte.entreprise = request.user.entreprise
            compte.classe = compte.numero_compte[0] if compte.numero_compte else '1'
            compte.save()
            messages.success(request, f"Compte {compte.numero_compte} créé avec succès.")
            return redirect('comptabilite:plan_comptable_list')
    else:
        form = PlanComptableForm(entreprise=request.user.entreprise)
    
    return render(request, 'comptabilite/plan_comptable/form.html', {'form': form})


@login_required
@compta_required
def plan_comptable_detail(request, pk):
    """Détail d'un compte comptable"""
    compte = get_object_or_404(PlanComptable, pk=pk, entreprise=request.user.entreprise)
    
    # Mouvements du compte
    lignes = LigneEcriture.objects.filter(
        compte=compte
    ).select_related('ecriture', 'ecriture__journal').order_by('-ecriture__date_ecriture')[:50]
    
    context = {
        'compte': compte,
        'lignes': lignes,
    }
    return render(request, 'comptabilite/plan_comptable/detail.html', context)


@login_required
@compta_required
def plan_comptable_update(request, pk):
    """Modifier un compte comptable"""
    compte = get_object_or_404(PlanComptable, pk=pk, entreprise=request.user.entreprise)
    
    if request.method == 'POST':
        form = PlanComptableForm(request.POST, instance=compte, entreprise=request.user.entreprise)
        if form.is_valid():
            form.save()
            messages.success(request, "Compte modifié avec succès.")
            return redirect('comptabilite:plan_comptable_list')
    else:
        form = PlanComptableForm(instance=compte, entreprise=request.user.entreprise)
    
    return render(request, 'comptabilite/plan_comptable/form.html', {'form': form, 'compte': compte})


# ==================== JOURNAUX ====================

@login_required
@compta_required
def journal_list(request):
    """Liste des journaux"""
    journaux = Journal.objects.filter(entreprise=request.user.entreprise)
    return render(request, 'comptabilite/journaux/list.html', {'journaux': journaux})


@login_required
@compta_required
def journal_create(request):
    """Créer un journal"""
    if request.method == 'POST':
        form = JournalForm(request.POST, entreprise=request.user.entreprise)
        if form.is_valid():
            journal = form.save(commit=False)
            journal.entreprise = request.user.entreprise
            journal.save()
            messages.success(request, f"Journal {journal.code} créé avec succès.")
            return redirect('comptabilite:journal_list')
    else:
        form = JournalForm(entreprise=request.user.entreprise)
    
    return render(request, 'comptabilite/journaux/form.html', {'form': form})


@login_required
@compta_required
def journal_update(request, pk):
    """Modifier un journal"""
    journal = get_object_or_404(Journal, pk=pk, entreprise=request.user.entreprise)
    
    if request.method == 'POST':
        form = JournalForm(request.POST, instance=journal, entreprise=request.user.entreprise)
        if form.is_valid():
            form.save()
            messages.success(request, "Journal modifié avec succès.")
            return redirect('comptabilite:journal_list')
    else:
        form = JournalForm(instance=journal, entreprise=request.user.entreprise)
    
    return render(request, 'comptabilite/journaux/form.html', {'form': form, 'journal': journal})


# ==================== EXERCICES ====================

@login_required
@compta_required
def exercice_list(request):
    """Liste des exercices comptables"""
    exercices = ExerciceComptable.objects.filter(entreprise=request.user.entreprise)
    return render(request, 'comptabilite/exercices/list.html', {'exercices': exercices})


@login_required
@compta_required
def exercice_create(request):
    """Créer un exercice"""
    if request.method == 'POST':
        form = ExerciceForm(request.POST)
        if form.is_valid():
            exercice = form.save(commit=False)
            exercice.entreprise = request.user.entreprise
            exercice.save()
            messages.success(request, f"Exercice {exercice.libelle} créé avec succès.")
            return redirect('comptabilite:exercice_list')
    else:
        form = ExerciceForm()
    
    return render(request, 'comptabilite/exercices/form.html', {'form': form})


@login_required
@compta_required
def exercice_update(request, pk):
    """Modifier un exercice"""
    exercice = get_object_or_404(ExerciceComptable, pk=pk, entreprise=request.user.entreprise)
    
    if request.method == 'POST':
        form = ExerciceForm(request.POST, instance=exercice)
        if form.is_valid():
            form.save()
            messages.success(request, "Exercice modifié avec succès.")
            return redirect('comptabilite:exercice_list')
    else:
        form = ExerciceForm(instance=exercice)
    
    return render(request, 'comptabilite/exercices/form.html', {'form': form, 'exercice': exercice})


# ==================== ÉCRITURES ====================

@login_required
@compta_required
def ecriture_list(request):
    """Liste des écritures comptables"""
    entreprise = request.user.entreprise
    journal_id = request.GET.get('journal', '')
    date_debut = request.GET.get('date_debut', '')
    date_fin = request.GET.get('date_fin', '')
    
    ecritures = EcritureComptable.objects.filter(
        entreprise=entreprise
    ).select_related('journal', 'exercice')
    
    if journal_id:
        ecritures = ecritures.filter(journal_id=journal_id)
    if date_debut:
        ecritures = ecritures.filter(date_ecriture__gte=date_debut)
    if date_fin:
        ecritures = ecritures.filter(date_ecriture__lte=date_fin)
    
    ecritures = ecritures.order_by('-date_ecriture', '-numero_ecriture')
    
    paginator = Paginator(ecritures, 25)
    page = request.GET.get('page', 1)
    ecritures = paginator.get_page(page)
    
    journaux = Journal.objects.filter(entreprise=entreprise, est_actif=True)
    
    context = {
        'ecritures': ecritures,
        'journaux': journaux,
    }
    return render(request, 'comptabilite/ecritures/list.html', context)


@login_required
@compta_required
def ecriture_create(request):
    """Créer une écriture comptable"""
    entreprise = request.user.entreprise
    
    if request.method == 'POST':
        form = EcritureForm(request.POST, entreprise=entreprise)
        if form.is_valid():
            ecriture = form.save(commit=False)
            ecriture.entreprise = entreprise
            ecriture.save()
            
            # Traiter les lignes d'écriture
            comptes_ids = request.POST.getlist('compte[]')
            libelles = request.POST.getlist('libelle_ligne[]')
            debits = request.POST.getlist('debit[]')
            credits = request.POST.getlist('credit[]')
            
            for i, compte_id in enumerate(comptes_ids):
                if compte_id:
                    debit = Decimal(debits[i]) if debits[i] else Decimal('0')
                    credit = Decimal(credits[i]) if credits[i] else Decimal('0')
                    
                    if debit > 0 or credit > 0:
                        LigneEcriture.objects.create(
                            ecriture=ecriture,
                            compte_id=compte_id,
                            libelle=libelles[i] if i < len(libelles) else '',
                            montant_debit=debit,
                            montant_credit=credit
                        )
            
            messages.success(request, f"Écriture {ecriture.numero_ecriture} créée avec {ecriture.lignes.count()} lignes.")
            return redirect('comptabilite:ecriture_detail', pk=ecriture.pk)
    else:
        form = EcritureForm(entreprise=entreprise)
    
    comptes = PlanComptable.objects.filter(entreprise=entreprise, est_actif=True).order_by('numero_compte')
    
    context = {
        'form': form,
        'comptes': comptes,
    }
    return render(request, 'comptabilite/ecritures/form.html', context)


@login_required
@compta_required
def ecriture_detail(request, pk):
    """Détail d'une écriture"""
    ecriture = get_object_or_404(
        EcritureComptable.objects.prefetch_related('lignes__compte'),
        pk=pk, entreprise=request.user.entreprise
    )
    return render(request, 'comptabilite/ecritures/detail.html', {'ecriture': ecriture})


@login_required
@compta_required
def ecriture_update(request, pk):
    """Modifier une écriture"""
    ecriture = get_object_or_404(EcritureComptable, pk=pk, entreprise=request.user.entreprise)
    
    if ecriture.est_validee:
        messages.error(request, "Impossible de modifier une écriture validée.")
        return redirect('comptabilite:ecriture_detail', pk=pk)
    
    if request.method == 'POST':
        form = EcritureForm(request.POST, instance=ecriture, entreprise=request.user.entreprise)
        if form.is_valid():
            form.save()
            
            # Supprimer les anciennes lignes et recréer
            ecriture.lignes.all().delete()
            
            comptes_ids = request.POST.getlist('compte[]')
            libelles = request.POST.getlist('libelle_ligne[]')
            debits = request.POST.getlist('debit[]')
            credits = request.POST.getlist('credit[]')
            
            for i, compte_id in enumerate(comptes_ids):
                if compte_id:
                    debit = Decimal(debits[i]) if debits[i] else Decimal('0')
                    credit = Decimal(credits[i]) if credits[i] else Decimal('0')
                    
                    if debit > 0 or credit > 0:
                        LigneEcriture.objects.create(
                            ecriture=ecriture,
                            compte_id=compte_id,
                            libelle=libelles[i] if i < len(libelles) else '',
                            montant_debit=debit,
                            montant_credit=credit
                        )
            
            messages.success(request, "Écriture modifiée avec succès.")
            return redirect('comptabilite:ecriture_detail', pk=pk)
    else:
        form = EcritureForm(instance=ecriture, entreprise=request.user.entreprise)
    
    comptes = PlanComptable.objects.filter(entreprise=request.user.entreprise, est_actif=True).order_by('numero_compte')
    
    context = {
        'form': form,
        'ecriture': ecriture,
        'comptes': comptes,
    }
    return render(request, 'comptabilite/ecritures/form.html', context)


@login_required
@compta_required
def ecriture_valider(request, pk):
    """Valider une écriture"""
    ecriture = get_object_or_404(EcritureComptable, pk=pk, entreprise=request.user.entreprise)
    
    if not ecriture.est_equilibree:
        messages.error(request, "L'écriture n'est pas équilibrée.")
        return redirect('comptabilite:ecriture_detail', pk=pk)
    
    ecriture.est_validee = True
    ecriture.date_validation = timezone.now()
    ecriture.validee_par = request.user
    ecriture.save()
    
    messages.success(request, "Écriture validée avec succès.")
    return redirect('comptabilite:ecriture_detail', pk=pk)


# ==================== TIERS ====================

@login_required
@compta_required
def tiers_list(request):
    """Liste des tiers"""
    entreprise = request.user.entreprise
    type_tiers = request.GET.get('type', '')
    search = request.GET.get('q', '')
    
    tiers = Tiers.objects.filter(entreprise=entreprise)
    
    if type_tiers:
        tiers = tiers.filter(type_tiers=type_tiers)
    if search:
        tiers = tiers.filter(
            Q(code__icontains=search) | Q(raison_sociale__icontains=search)
        )
    
    paginator = Paginator(tiers, 25)
    page = request.GET.get('page', 1)
    tiers = paginator.get_page(page)
    
    context = {
        'tiers': tiers,
        'types_tiers': Tiers.TYPES_TIERS,
    }
    return render(request, 'comptabilite/tiers/list.html', context)


@login_required
@compta_required
def tiers_create(request):
    """Créer un tiers"""
    if request.method == 'POST':
        form = TiersForm(request.POST, entreprise=request.user.entreprise)
        if form.is_valid():
            tiers = form.save(commit=False)
            tiers.entreprise = request.user.entreprise
            tiers.save()
            messages.success(request, f"Tiers {tiers.raison_sociale} créé avec succès.")
            return redirect('comptabilite:tiers_list')
    else:
        form = TiersForm(entreprise=request.user.entreprise)
    
    return render(request, 'comptabilite/tiers/form.html', {'form': form})


@login_required
@compta_required
def tiers_detail(request, pk):
    """Détail d'un tiers"""
    tiers = get_object_or_404(Tiers, pk=pk, entreprise=request.user.entreprise)
    factures = Facture.objects.filter(tiers=tiers).order_by('-date_facture')[:10]
    
    context = {
        'tiers': tiers,
        'factures': factures,
    }
    return render(request, 'comptabilite/tiers/detail.html', context)


@login_required
@compta_required
def tiers_update(request, pk):
    """Modifier un tiers"""
    tiers = get_object_or_404(Tiers, pk=pk, entreprise=request.user.entreprise)
    
    if request.method == 'POST':
        form = TiersForm(request.POST, instance=tiers, entreprise=request.user.entreprise)
        if form.is_valid():
            form.save()
            messages.success(request, "Tiers modifié avec succès.")
            return redirect('comptabilite:tiers_detail', pk=pk)
    else:
        form = TiersForm(instance=tiers, entreprise=request.user.entreprise)
    
    return render(request, 'comptabilite/tiers/form.html', {'form': form, 'tiers': tiers})


# ==================== FACTURES ====================

@login_required
@compta_required
def facture_list(request):
    """Liste des factures"""
    entreprise = request.user.entreprise
    type_facture = request.GET.get('type', '')
    statut = request.GET.get('statut', '')
    
    factures = Facture.objects.filter(entreprise=entreprise).select_related('tiers')
    
    if type_facture:
        factures = factures.filter(type_facture=type_facture)
    if statut:
        factures = factures.filter(statut=statut)
    
    paginator = Paginator(factures, 25)
    page = request.GET.get('page', 1)
    factures = paginator.get_page(page)
    
    context = {
        'factures': factures,
        'types_facture': Facture.TYPES_FACTURE,
        'statuts': Facture.STATUTS,
    }
    return render(request, 'comptabilite/factures/list.html', context)


@login_required
@compta_required
def facture_create(request):
    """Créer une facture"""
    type_facture = request.GET.get('type', 'vente')
    
    if request.method == 'POST':
        form = FactureForm(request.POST, entreprise=request.user.entreprise)
        if form.is_valid():
            facture = form.save(commit=False)
            facture.entreprise = request.user.entreprise
            facture.save()
            messages.success(request, f"Facture {facture.numero} créée.")
            return redirect('comptabilite:facture_detail', pk=facture.pk)
    else:
        form = FactureForm(entreprise=request.user.entreprise, initial={'type_facture': type_facture})
    
    return render(request, 'comptabilite/factures/form.html', {'form': form})


@login_required
@compta_required
def facture_detail(request, pk):
    """Détail d'une facture"""
    facture = get_object_or_404(
        Facture.objects.prefetch_related('lignes', 'reglements'),
        pk=pk, entreprise=request.user.entreprise
    )
    return render(request, 'comptabilite/factures/detail.html', {'facture': facture})


@login_required
@compta_required
def facture_update(request, pk):
    """Modifier une facture"""
    facture = get_object_or_404(Facture, pk=pk, entreprise=request.user.entreprise)
    
    if facture.statut != 'brouillon':
        messages.error(request, "Seules les factures en brouillon peuvent être modifiées.")
        return redirect('comptabilite:facture_detail', pk=pk)
    
    if request.method == 'POST':
        form = FactureForm(request.POST, instance=facture, entreprise=request.user.entreprise)
        if form.is_valid():
            form.save()
            messages.success(request, "Facture modifiée avec succès.")
            return redirect('comptabilite:facture_detail', pk=pk)
    else:
        form = FactureForm(instance=facture, entreprise=request.user.entreprise)
    
    return render(request, 'comptabilite/factures/form.html', {'form': form, 'facture': facture})


@login_required
@compta_required
def facture_valider(request, pk):
    """Valider une facture"""
    facture = get_object_or_404(Facture, pk=pk, entreprise=request.user.entreprise)
    
    if facture.statut != 'brouillon':
        messages.error(request, "Cette facture ne peut pas être validée.")
        return redirect('comptabilite:facture_detail', pk=pk)
    
    facture.statut = 'validee'
    facture.save()
    
    messages.success(request, "Facture validée avec succès.")
    return redirect('comptabilite:facture_detail', pk=pk)


@login_required
@compta_required
def facture_print(request, pk):
    """Imprimer une facture"""
    facture = get_object_or_404(
        Facture.objects.prefetch_related('lignes'),
        pk=pk, entreprise=request.user.entreprise
    )
    return render(request, 'comptabilite/factures/print.html', {'facture': facture})


# ==================== RÈGLEMENTS ====================

@login_required
@compta_required
def reglement_list(request):
    """Liste des règlements"""
    reglements = Reglement.objects.filter(
        entreprise=request.user.entreprise
    ).select_related('facture', 'facture__tiers')
    
    paginator = Paginator(reglements, 25)
    page = request.GET.get('page', 1)
    reglements = paginator.get_page(page)
    
    return render(request, 'comptabilite/reglements/list.html', {'reglements': reglements})


@login_required
@compta_required
def reglement_create(request):
    """Créer un règlement"""
    facture_id = request.GET.get('facture', '')
    
    if request.method == 'POST':
        form = ReglementForm(request.POST, entreprise=request.user.entreprise)
        if form.is_valid():
            reglement = form.save(commit=False)
            reglement.entreprise = request.user.entreprise
            reglement.save()
            
            # Mettre à jour le montant payé sur la facture
            facture = reglement.facture
            facture.montant_paye += reglement.montant
            if facture.montant_paye >= facture.montant_ttc:
                facture.statut = 'payee'
            facture.save()
            
            messages.success(request, "Règlement enregistré avec succès.")
            return redirect('comptabilite:facture_detail', pk=facture.pk)
    else:
        initial = {}
        if facture_id:
            initial['facture'] = facture_id
        form = ReglementForm(entreprise=request.user.entreprise, initial=initial)
    
    return render(request, 'comptabilite/reglements/form.html', {'form': form})


@login_required
@compta_required
def reglement_detail(request, pk):
    """Détail d'un règlement"""
    reglement = get_object_or_404(
        Reglement.objects.select_related('facture', 'facture__tiers'),
        pk=pk, entreprise=request.user.entreprise
    )
    return render(request, 'comptabilite/reglements/detail.html', {'reglement': reglement})


# ==================== ÉTATS FINANCIERS ====================

@login_required
@compta_required
def grand_livre(request):
    """Grand livre comptable"""
    entreprise = request.user.entreprise
    compte_id = request.GET.get('compte', '')
    date_debut = request.GET.get('date_debut', '')
    date_fin = request.GET.get('date_fin', '')
    
    lignes = LigneEcriture.objects.filter(
        ecriture__entreprise=entreprise,
        ecriture__est_validee=True
    ).select_related('compte', 'ecriture', 'ecriture__journal')
    
    if compte_id:
        lignes = lignes.filter(compte_id=compte_id)
    if date_debut:
        lignes = lignes.filter(ecriture__date_ecriture__gte=date_debut)
    if date_fin:
        lignes = lignes.filter(ecriture__date_ecriture__lte=date_fin)
    
    lignes = lignes.order_by('compte__numero_compte', 'ecriture__date_ecriture')
    
    # Regrouper les lignes par compte avec totaux
    from collections import OrderedDict
    comptes_groupes = OrderedDict()
    for ligne in lignes:
        compte = ligne.compte
        if compte.pk not in comptes_groupes:
            comptes_groupes[compte.pk] = {
                'compte': compte,
                'lignes': [],
                'total_debit': Decimal('0'),
                'total_credit': Decimal('0'),
            }
        comptes_groupes[compte.pk]['lignes'].append(ligne)
        comptes_groupes[compte.pk]['total_debit'] += ligne.montant_debit or Decimal('0')
        comptes_groupes[compte.pk]['total_credit'] += ligne.montant_credit or Decimal('0')
    
    comptes = PlanComptable.objects.filter(entreprise=entreprise, est_actif=True).order_by('numero_compte')
    
    context = {
        'comptes_groupes': comptes_groupes.values(),
        'comptes': comptes,
        'compte_id': compte_id,
        'date_debut': date_debut,
        'date_fin': date_fin,
    }
    return render(request, 'comptabilite/etats/grand_livre.html', context)


def _get_grand_livre_data(request):
    """Helper pour récupérer les données du grand livre"""
    entreprise = request.user.entreprise
    compte_id = request.GET.get('compte', '')
    date_debut = request.GET.get('date_debut', '')
    date_fin = request.GET.get('date_fin', '')
    
    lignes = LigneEcriture.objects.filter(
        ecriture__entreprise=entreprise,
        ecriture__est_validee=True
    ).select_related('compte', 'ecriture', 'ecriture__journal')
    
    if compte_id:
        lignes = lignes.filter(compte_id=compte_id)
    if date_debut:
        lignes = lignes.filter(ecriture__date_ecriture__gte=date_debut)
    if date_fin:
        lignes = lignes.filter(ecriture__date_ecriture__lte=date_fin)
    
    lignes = lignes.order_by('compte__numero_compte', 'ecriture__date_ecriture')
    
    comptes_groupes = OrderedDict()
    for ligne in lignes:
        compte = ligne.compte
        if compte.pk not in comptes_groupes:
            comptes_groupes[compte.pk] = {
                'compte': compte,
                'lignes': [],
                'total_debit': Decimal('0'),
                'total_credit': Decimal('0'),
            }
        comptes_groupes[compte.pk]['lignes'].append(ligne)
        comptes_groupes[compte.pk]['total_debit'] += ligne.montant_debit or Decimal('0')
        comptes_groupes[compte.pk]['total_credit'] += ligne.montant_credit or Decimal('0')
    
    return comptes_groupes, entreprise, date_debut, date_fin


@login_required
@compta_required
def grand_livre_excel(request):
    """Export Excel du Grand Livre"""
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter
    
    comptes_groupes, entreprise, date_debut, date_fin = _get_grand_livre_data(request)
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Grand Livre"
    
    # Styles
    header_font = Font(bold=True, size=14)
    title_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="FFD699", end_color="FFD699", fill_type="solid")
    total_fill = PatternFill(start_color="E8E8E8", end_color="E8E8E8", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    # En-tête
    ws.merge_cells('A1:F1')
    ws['A1'] = entreprise.nom_entreprise
    ws['A1'].font = header_font
    ws['A1'].alignment = Alignment(horizontal='center')
    
    ws.merge_cells('A2:F2')
    periode = f"GRAND LIVRE - Du {date_debut or '--'} au {date_fin or '--'}"
    ws['A2'] = periode
    ws['A2'].alignment = Alignment(horizontal='center')
    
    row = 4
    for groupe in comptes_groupes.values():
        compte = groupe['compte']
        
        # En-tête du compte
        ws.merge_cells(f'A{row}:F{row}')
        ws[f'A{row}'] = f"{compte.numero_compte} - {compte.intitule}"
        ws[f'A{row}'].font = title_font
        ws[f'A{row}'].fill = header_fill
        row += 1
        
        # En-têtes colonnes
        headers = ['Date', 'Journal', 'N° Pièce', 'Libellé', 'Débit', 'Crédit']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = Font(bold=True)
            cell.border = thin_border
        row += 1
        
        # Lignes
        for ligne in groupe['lignes']:
            ws.cell(row=row, column=1, value=ligne.ecriture.date_ecriture.strftime('%d/%m/%Y')).border = thin_border
            ws.cell(row=row, column=2, value=ligne.ecriture.journal.code).border = thin_border
            ws.cell(row=row, column=3, value=ligne.ecriture.numero_ecriture).border = thin_border
            ws.cell(row=row, column=4, value=ligne.libelle or ligne.ecriture.libelle).border = thin_border
            cell_d = ws.cell(row=row, column=5, value=float(ligne.montant_debit) if ligne.montant_debit else None)
            cell_d.border = thin_border
            cell_d.number_format = '#,##0'
            cell_c = ws.cell(row=row, column=6, value=float(ligne.montant_credit) if ligne.montant_credit else None)
            cell_c.border = thin_border
            cell_c.number_format = '#,##0'
            row += 1
        
        # Total compte
        ws.merge_cells(f'A{row}:D{row}')
        ws[f'A{row}'] = f"Total compte {compte.numero_compte}"
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'A{row}'].fill = total_fill
        ws[f'A{row}'].alignment = Alignment(horizontal='right')
        cell_td = ws.cell(row=row, column=5, value=float(groupe['total_debit']))
        cell_td.font = Font(bold=True)
        cell_td.fill = total_fill
        cell_td.number_format = '#,##0'
        cell_tc = ws.cell(row=row, column=6, value=float(groupe['total_credit']))
        cell_tc.font = Font(bold=True)
        cell_tc.fill = total_fill
        cell_tc.number_format = '#,##0'
        row += 2
    
    # Ajuster largeurs colonnes
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 40
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15
    
    # Réponse
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="grand_livre_{timezone.now().strftime("%Y%m%d")}.xlsx"'
    wb.save(response)
    return response


@login_required
@compta_required
def grand_livre_pdf(request):
    """Export PDF du Grand Livre"""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    
    comptes_groupes, entreprise, date_debut, date_fin = _get_grand_livre_data(request)
    
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), leftMargin=1*cm, rightMargin=1*cm, topMargin=1*cm, bottomMargin=1*cm)
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('Title', parent=styles['Heading1'], fontSize=16, alignment=1, spaceAfter=10)
    subtitle_style = ParagraphStyle('Subtitle', parent=styles['Normal'], fontSize=10, alignment=1, spaceAfter=20)
    compte_style = ParagraphStyle('Compte', parent=styles['Heading3'], fontSize=11, backColor=colors.HexColor('#FFD699'), spaceAfter=5)
    
    elements = []
    
    # En-tête
    elements.append(Paragraph(entreprise.nom_entreprise, title_style))
    periode = f"GRAND LIVRE COMPTABLE - Du {date_debut or '--/--/----'} au {date_fin or '--/--/----'}"
    elements.append(Paragraph(periode, subtitle_style))
    
    for groupe in comptes_groupes.values():
        compte = groupe['compte']
        
        # Titre du compte
        elements.append(Paragraph(f"<b>{compte.numero_compte}</b> - {compte.intitule}", compte_style))
        
        # Table des lignes
        data = [['Date', 'Journal', 'N° Pièce', 'Libellé', 'Débit', 'Crédit']]
        for ligne in groupe['lignes']:
            data.append([
                ligne.ecriture.date_ecriture.strftime('%d/%m/%Y'),
                ligne.ecriture.journal.code,
                ligne.ecriture.numero_ecriture,
                (ligne.libelle or ligne.ecriture.libelle)[:50],
                f"{ligne.montant_debit:,.0f}" if ligne.montant_debit else '',
                f"{ligne.montant_credit:,.0f}" if ligne.montant_credit else '',
            ])
        
        # Ligne total
        data.append([
            '', '', '', f"Total compte {compte.numero_compte}",
            f"{groupe['total_debit']:,.0f}",
            f"{groupe['total_credit']:,.0f}"
        ])
        
        table = Table(data, colWidths=[2.5*cm, 2*cm, 3*cm, 10*cm, 3.5*cm, 3.5*cm])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#EF7707')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('ALIGN', (4, 0), (5, -1), 'RIGHT'),
            ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#E8E8E8')),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))
        
        elements.append(table)
        elements.append(Spacer(1, 0.5*cm))
    
    doc.build(elements)
    buffer.seek(0)
    
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="grand_livre_{timezone.now().strftime("%Y%m%d")}.pdf"'
    return response


@login_required
@compta_required
def balance(request):
    """Balance générale"""
    entreprise = request.user.entreprise
    
    comptes = PlanComptable.objects.filter(
        entreprise=entreprise, est_actif=True
    ).annotate(
        total_debit=Sum('lignes_ecritures__montant_debit'),
        total_credit=Sum('lignes_ecritures__montant_credit')
    ).order_by('numero_compte')
    
    # Calculer les totaux
    total_debit = Decimal('0')
    total_credit = Decimal('0')
    total_solde_debit = Decimal('0')
    total_solde_credit = Decimal('0')
    
    for c in comptes:
        d = c.total_debit or Decimal('0')
        cr = c.total_credit or Decimal('0')
        total_debit += d
        total_credit += cr
        if d > cr:
            total_solde_debit += (d - cr)
        else:
            total_solde_credit += (cr - d)
    
    context = {
        'comptes': comptes,
        'total_debit': total_debit,
        'total_credit': total_credit,
        'total_solde_debit': total_solde_debit,
        'total_solde_credit': total_solde_credit,
        'today': timezone.now().date(),
    }
    return render(request, 'comptabilite/etats/balance.html', context)


def _get_balance_data(request):
    """Helper pour récupérer les données de la balance"""
    entreprise = request.user.entreprise
    
    comptes = PlanComptable.objects.filter(
        entreprise=entreprise, est_actif=True
    ).annotate(
        total_debit=Sum('lignes_ecritures__montant_debit'),
        total_credit=Sum('lignes_ecritures__montant_credit')
    ).order_by('numero_compte')
    
    total_debit = Decimal('0')
    total_credit = Decimal('0')
    total_solde_debit = Decimal('0')
    total_solde_credit = Decimal('0')
    
    comptes_avec_mvt = []
    for c in comptes:
        d = c.total_debit or Decimal('0')
        cr = c.total_credit or Decimal('0')
        if d or cr:
            solde_d = (d - cr) if d > cr else Decimal('0')
            solde_c = (cr - d) if cr > d else Decimal('0')
            comptes_avec_mvt.append({
                'numero': c.numero_compte,
                'intitule': c.intitule,
                'mvt_debit': d,
                'mvt_credit': cr,
                'solde_debit': solde_d,
                'solde_credit': solde_c,
            })
            total_debit += d
            total_credit += cr
            total_solde_debit += solde_d
            total_solde_credit += solde_c
    
    return comptes_avec_mvt, entreprise, total_debit, total_credit, total_solde_debit, total_solde_credit


@login_required
@compta_required
def balance_excel(request):
    """Export Excel de la Balance"""
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    
    comptes, entreprise, total_debit, total_credit, total_solde_debit, total_solde_credit = _get_balance_data(request)
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Balance"
    
    header_font = Font(bold=True, size=14)
    header_fill = PatternFill(start_color="FFD699", end_color="FFD699", fill_type="solid")
    total_fill = PatternFill(start_color="FFD699", end_color="FFD699", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    # En-tête
    ws.merge_cells('A1:F1')
    ws['A1'] = entreprise.nom_entreprise
    ws['A1'].font = header_font
    ws['A1'].alignment = Alignment(horizontal='center')
    
    ws.merge_cells('A2:F2')
    ws['A2'] = f"BALANCE GÉNÉRALE - Arrêtée au {timezone.now().strftime('%d/%m/%Y')}"
    ws['A2'].alignment = Alignment(horizontal='center')
    
    # En-têtes colonnes
    headers = ['N° Compte', 'Intitulé', 'Mvt Débit', 'Mvt Crédit', 'Solde Débit', 'Solde Crédit']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.border = thin_border
    
    # Données
    row = 5
    for c in comptes:
        ws.cell(row=row, column=1, value=c['numero']).border = thin_border
        ws.cell(row=row, column=2, value=c['intitule']).border = thin_border
        cell_md = ws.cell(row=row, column=3, value=float(c['mvt_debit']))
        cell_md.border = thin_border
        cell_md.number_format = '#,##0'
        cell_mc = ws.cell(row=row, column=4, value=float(c['mvt_credit']))
        cell_mc.border = thin_border
        cell_mc.number_format = '#,##0'
        cell_sd = ws.cell(row=row, column=5, value=float(c['solde_debit']) if c['solde_debit'] else None)
        cell_sd.border = thin_border
        cell_sd.number_format = '#,##0'
        cell_sc = ws.cell(row=row, column=6, value=float(c['solde_credit']) if c['solde_credit'] else None)
        cell_sc.border = thin_border
        cell_sc.number_format = '#,##0'
        row += 1
    
    # Totaux
    ws.cell(row=row, column=1, value='').border = thin_border
    ws.cell(row=row, column=2, value='TOTAUX').border = thin_border
    ws[f'B{row}'].font = Font(bold=True)
    ws[f'B{row}'].fill = total_fill
    for col in range(1, 7):
        ws.cell(row=row, column=col).fill = total_fill
        ws.cell(row=row, column=col).font = Font(bold=True)
    ws.cell(row=row, column=3, value=float(total_debit)).number_format = '#,##0'
    ws.cell(row=row, column=4, value=float(total_credit)).number_format = '#,##0'
    ws.cell(row=row, column=5, value=float(total_solde_debit)).number_format = '#,##0'
    ws.cell(row=row, column=6, value=float(total_solde_credit)).number_format = '#,##0'
    
    # Largeurs colonnes
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="balance_{timezone.now().strftime("%Y%m%d")}.xlsx"'
    wb.save(response)
    return response


@login_required
@compta_required
def balance_pdf(request):
    """Export PDF de la Balance"""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    
    comptes, entreprise, total_debit, total_credit, total_solde_debit, total_solde_credit = _get_balance_data(request)
    
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), leftMargin=1*cm, rightMargin=1*cm, topMargin=1*cm, bottomMargin=1*cm)
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('Title', parent=styles['Heading1'], fontSize=16, alignment=1, spaceAfter=10)
    subtitle_style = ParagraphStyle('Subtitle', parent=styles['Normal'], fontSize=10, alignment=1, spaceAfter=20)
    
    elements = []
    
    # En-tête
    elements.append(Paragraph(entreprise.nom_entreprise, title_style))
    elements.append(Paragraph(f"BALANCE GÉNÉRALE - Arrêtée au {timezone.now().strftime('%d/%m/%Y')}", subtitle_style))
    
    # Table
    data = [['N° Compte', 'Intitulé', 'Mvt Débit', 'Mvt Crédit', 'Solde Débit', 'Solde Crédit']]
    for c in comptes:
        data.append([
            c['numero'],
            c['intitule'][:50],
            f"{c['mvt_debit']:,.0f}",
            f"{c['mvt_credit']:,.0f}",
            f"{c['solde_debit']:,.0f}" if c['solde_debit'] else '',
            f"{c['solde_credit']:,.0f}" if c['solde_credit'] else '',
        ])
    
    # Ligne totaux
    data.append([
        '', 'TOTAUX',
        f"{total_debit:,.0f}",
        f"{total_credit:,.0f}",
        f"{total_solde_debit:,.0f}",
        f"{total_solde_credit:,.0f}",
    ])
    
    table = Table(data, colWidths=[2.5*cm, 10*cm, 3.5*cm, 3.5*cm, 3.5*cm, 3.5*cm])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#EF7707')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('ALIGN', (2, 0), (-1, -1), 'RIGHT'),
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#FFD699')),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    
    elements.append(table)
    
    doc.build(elements)
    buffer.seek(0)
    
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="balance_{timezone.now().strftime("%Y%m%d")}.pdf"'
    return response


@login_required
@compta_required
def journal_general(request):
    """Journal général"""
    entreprise = request.user.entreprise
    journal_id = request.GET.get('journal', '')
    date_debut = request.GET.get('date_debut', '')
    date_fin = request.GET.get('date_fin', '')
    
    ecritures = EcritureComptable.objects.filter(
        entreprise=entreprise,
        est_validee=True
    ).prefetch_related('lignes__compte').select_related('journal')
    
    if journal_id:
        ecritures = ecritures.filter(journal_id=journal_id)
    if date_debut:
        ecritures = ecritures.filter(date_ecriture__gte=date_debut)
    if date_fin:
        ecritures = ecritures.filter(date_ecriture__lte=date_fin)
    
    ecritures = ecritures.order_by('date_ecriture', 'numero_ecriture')
    
    journaux = Journal.objects.filter(entreprise=entreprise, est_actif=True)
    
    context = {
        'ecritures': ecritures,
        'journaux': journaux,
    }
    return render(request, 'comptabilite/etats/journal_general.html', context)


@login_required
@compta_required
def bilan(request):
    """Bilan comptable"""
    entreprise = request.user.entreprise
    
    exercice = ExerciceComptable.objects.filter(entreprise=entreprise, est_courant=True).first()
    
    def get_comptes_avec_solde(classes):
        comptes = PlanComptable.objects.filter(
            entreprise=entreprise,
            classe__in=classes,
            est_actif=True
        ).annotate(
            total_debit=Sum('lignes_ecritures__montant_debit'),
            total_credit=Sum('lignes_ecritures__montant_credit')
        ).order_by('numero_compte')
        
        result = []
        for c in comptes:
            d = c.total_debit or Decimal('0')
            cr = c.total_credit or Decimal('0')
            solde = d - cr if d > cr else cr - d
            if solde > 0:
                result.append({'numero_compte': c.numero_compte, 'intitule': c.intitule, 'solde': solde, 'debiteur': d > cr})
        return result
    
    # Actif immobilisé (classe 2)
    actif_immobilise = get_comptes_avec_solde(['2'])
    
    # Actif circulant (classes 3, 4 débiteur, 5)
    actif_circulant = get_comptes_avec_solde(['3', '5'])
    # Ajouter comptes classe 4 débiteurs
    comptes_4 = PlanComptable.objects.filter(entreprise=entreprise, classe='4', est_actif=True).annotate(
        total_debit=Sum('lignes_ecritures__montant_debit'),
        total_credit=Sum('lignes_ecritures__montant_credit')
    )
    for c in comptes_4:
        d = c.total_debit or Decimal('0')
        cr = c.total_credit or Decimal('0')
        if d > cr:
            actif_circulant.append({'numero_compte': c.numero_compte, 'intitule': c.intitule, 'solde': d - cr, 'debiteur': True})
    
    # Capitaux propres (classe 1)
    capitaux_propres = get_comptes_avec_solde(['1'])
    
    # Dettes (classe 4 créditeur)
    dettes = []
    for c in comptes_4:
        d = c.total_debit or Decimal('0')
        cr = c.total_credit or Decimal('0')
        if cr > d:
            dettes.append({'numero_compte': c.numero_compte, 'intitule': c.intitule, 'solde': cr - d, 'debiteur': False})
    
    # Totaux
    total_actif = sum(c['solde'] for c in actif_immobilise) + sum(c['solde'] for c in actif_circulant)
    total_passif = sum(c['solde'] for c in capitaux_propres) + sum(c['solde'] for c in dettes)
    ecart = total_actif - total_passif
    
    context = {
        'actif_immobilise': actif_immobilise,
        'actif_circulant': actif_circulant,
        'capitaux_propres': capitaux_propres,
        'dettes': dettes,
        'total_actif': total_actif,
        'total_passif': total_passif,
        'ecart': ecart,
        'exercice': exercice,
        'today': timezone.now().date(),
    }
    return render(request, 'comptabilite/etats/bilan.html', context)


@login_required
@compta_required
def compte_resultat(request):
    """Compte de résultat"""
    entreprise = request.user.entreprise
    exercice = ExerciceComptable.objects.filter(entreprise=entreprise, est_courant=True).first()
    
    def get_comptes_classe(classe_prefix):
        comptes = PlanComptable.objects.filter(
            entreprise=entreprise,
            numero_compte__startswith=classe_prefix,
            est_actif=True
        ).annotate(
            total_debit=Sum('lignes_ecritures__montant_debit'),
            total_credit=Sum('lignes_ecritures__montant_credit')
        ).order_by('numero_compte')
        
        result = []
        for c in comptes:
            d = c.total_debit or Decimal('0')
            cr = c.total_credit or Decimal('0')
            solde = d - cr if d > cr else cr - d
            if solde > 0:
                result.append({'numero_compte': c.numero_compte, 'intitule': c.intitule, 'solde': solde})
        return result
    
    # Charges d'exploitation (60-65)
    charges_exploitation = get_comptes_classe('60') + get_comptes_classe('61') + get_comptes_classe('62') + get_comptes_classe('63') + get_comptes_classe('64') + get_comptes_classe('65')
    
    # Charges financières (66-67)
    charges_financieres = get_comptes_classe('66') + get_comptes_classe('67')
    
    # Produits d'exploitation (70-75)
    produits_exploitation = get_comptes_classe('70') + get_comptes_classe('71') + get_comptes_classe('72') + get_comptes_classe('73') + get_comptes_classe('74') + get_comptes_classe('75')
    
    # Produits financiers (76-77)
    produits_financiers = get_comptes_classe('76') + get_comptes_classe('77')
    
    # Totaux
    total_charges = sum(c['solde'] for c in charges_exploitation) + sum(c['solde'] for c in charges_financieres)
    total_produits = sum(c['solde'] for c in produits_exploitation) + sum(c['solde'] for c in produits_financiers)
    resultat = total_produits - total_charges
    
    context = {
        'charges_exploitation': charges_exploitation,
        'charges_financieres': charges_financieres,
        'produits_exploitation': produits_exploitation,
        'produits_financiers': produits_financiers,
        'total_charges': total_charges,
        'total_produits': total_produits,
        'resultat': resultat,
        'exercice': exercice,
        'today': timezone.now().date(),
    }
    return render(request, 'comptabilite/etats/compte_resultat.html', context)


# ============ EXPORTS JOURNAL GENERAL ============

def _get_journal_general_data(request):
    """Helper pour récupérer les données du journal général"""
    entreprise = request.user.entreprise
    journal_id = request.GET.get('journal', '')
    date_debut = request.GET.get('date_debut', '')
    date_fin = request.GET.get('date_fin', '')
    
    ecritures = EcritureComptable.objects.filter(
        entreprise=entreprise, est_validee=True
    ).prefetch_related('lignes__compte').select_related('journal')
    
    if journal_id:
        ecritures = ecritures.filter(journal_id=journal_id)
    if date_debut:
        ecritures = ecritures.filter(date_ecriture__gte=date_debut)
    if date_fin:
        ecritures = ecritures.filter(date_ecriture__lte=date_fin)
    
    return ecritures.order_by('date_ecriture', 'numero_ecriture'), entreprise, journal_id, date_debut, date_fin


@login_required
@compta_required
def journal_general_excel(request):
    """Export Excel du Journal Général"""
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    
    ecritures, entreprise, journal_id, date_debut, date_fin = _get_journal_general_data(request)
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Journal Général"
    
    header_font = Font(bold=True, size=14)
    header_fill = PatternFill(start_color="FFD699", end_color="FFD699", fill_type="solid")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    ws.merge_cells('A1:F1')
    ws['A1'] = entreprise.nom_entreprise
    ws['A1'].font = header_font
    ws['A1'].alignment = Alignment(horizontal='center')
    
    ws.merge_cells('A2:F2')
    ws['A2'] = f"JOURNAL GÉNÉRAL"
    ws['A2'].alignment = Alignment(horizontal='center')
    
    row = 4
    for ecriture in ecritures:
        # En-tête écriture
        ws.merge_cells(f'A{row}:F{row}')
        ws[f'A{row}'] = f"{ecriture.numero_ecriture} - {ecriture.libelle} ({ecriture.journal.code} - {ecriture.date_ecriture.strftime('%d/%m/%Y')})"
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'A{row}'].fill = header_fill
        row += 1
        
        # En-têtes colonnes
        for col, header in enumerate(['Compte', 'Intitulé', 'Libellé', 'Débit', 'Crédit'], 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = Font(bold=True)
            cell.border = thin_border
        row += 1
        
        # Lignes
        for ligne in ecriture.lignes.all():
            ws.cell(row=row, column=1, value=ligne.compte.numero_compte).border = thin_border
            ws.cell(row=row, column=2, value=ligne.compte.intitule).border = thin_border
            ws.cell(row=row, column=3, value=ligne.libelle or '').border = thin_border
            cell_d = ws.cell(row=row, column=4, value=float(ligne.montant_debit) if ligne.montant_debit else None)
            cell_d.border = thin_border
            cell_d.number_format = '#,##0'
            cell_c = ws.cell(row=row, column=5, value=float(ligne.montant_credit) if ligne.montant_credit else None)
            cell_c.border = thin_border
            cell_c.number_format = '#,##0'
            row += 1
        row += 1
    
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="journal_general_{timezone.now().strftime("%Y%m%d")}.xlsx"'
    wb.save(response)
    return response


@login_required
@compta_required
def journal_general_pdf(request):
    """Export PDF du Journal Général"""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    
    ecritures, entreprise, journal_id, date_debut, date_fin = _get_journal_general_data(request)
    
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), leftMargin=1*cm, rightMargin=1*cm, topMargin=1*cm, bottomMargin=1*cm)
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('Title', parent=styles['Heading1'], fontSize=16, alignment=1, spaceAfter=10)
    subtitle_style = ParagraphStyle('Subtitle', parent=styles['Normal'], fontSize=10, alignment=1, spaceAfter=20)
    
    elements = []
    elements.append(Paragraph(entreprise.nom_entreprise, title_style))
    elements.append(Paragraph("JOURNAL GÉNÉRAL", subtitle_style))
    
    for ecriture in ecritures:
        elements.append(Paragraph(f"<b>{ecriture.numero_ecriture}</b> - {ecriture.libelle} ({ecriture.journal.code} - {ecriture.date_ecriture.strftime('%d/%m/%Y')})", styles['Heading4']))
        
        data = [['Compte', 'Intitulé', 'Libellé', 'Débit', 'Crédit']]
        for ligne in ecriture.lignes.all():
            data.append([
                ligne.compte.numero_compte,
                ligne.compte.intitule[:35],
                (ligne.libelle or '')[:25],
                f"{ligne.montant_debit:,.0f}" if ligne.montant_debit else '',
                f"{ligne.montant_credit:,.0f}" if ligne.montant_credit else '',
            ])
        
        table = Table(data, colWidths=[2.5*cm, 8*cm, 6*cm, 3.5*cm, 3.5*cm])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#EF7707')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('ALIGN', (3, 0), (4, -1), 'RIGHT'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ]))
        elements.append(table)
        elements.append(Spacer(1, 0.3*cm))
    
    doc.build(elements)
    buffer.seek(0)
    
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="journal_general_{timezone.now().strftime("%Y%m%d")}.pdf"'
    return response


# ============ EXPORTS BILAN ============

def _get_bilan_data(request):
    """Helper pour récupérer les données du bilan"""
    entreprise = request.user.entreprise
    exercice = ExerciceComptable.objects.filter(entreprise=entreprise, est_courant=True).first()
    
    def get_comptes_avec_solde(classes):
        comptes = PlanComptable.objects.filter(
            entreprise=entreprise, classe__in=classes, est_actif=True
        ).annotate(
            total_debit=Sum('lignes_ecritures__montant_debit'),
            total_credit=Sum('lignes_ecritures__montant_credit')
        ).order_by('numero_compte')
        
        result = []
        for c in comptes:
            d = c.total_debit or Decimal('0')
            cr = c.total_credit or Decimal('0')
            solde = d - cr if d > cr else cr - d
            if solde > 0:
                result.append({'numero_compte': c.numero_compte, 'intitule': c.intitule, 'solde': solde, 'debiteur': d > cr})
        return result
    
    actif_immobilise = get_comptes_avec_solde(['2'])
    actif_circulant = get_comptes_avec_solde(['3', '5'])
    
    comptes_4 = PlanComptable.objects.filter(entreprise=entreprise, classe='4', est_actif=True).annotate(
        total_debit=Sum('lignes_ecritures__montant_debit'),
        total_credit=Sum('lignes_ecritures__montant_credit')
    )
    for c in comptes_4:
        d = c.total_debit or Decimal('0')
        cr = c.total_credit or Decimal('0')
        if d > cr:
            actif_circulant.append({'numero_compte': c.numero_compte, 'intitule': c.intitule, 'solde': d - cr, 'debiteur': True})
    
    capitaux_propres = get_comptes_avec_solde(['1'])
    dettes = []
    for c in comptes_4:
        d = c.total_debit or Decimal('0')
        cr = c.total_credit or Decimal('0')
        if cr > d:
            dettes.append({'numero_compte': c.numero_compte, 'intitule': c.intitule, 'solde': cr - d, 'debiteur': False})
    
    total_actif = sum(c['solde'] for c in actif_immobilise) + sum(c['solde'] for c in actif_circulant)
    total_passif = sum(c['solde'] for c in capitaux_propres) + sum(c['solde'] for c in dettes)
    
    return actif_immobilise, actif_circulant, capitaux_propres, dettes, total_actif, total_passif, entreprise, exercice


@login_required
@compta_required
def bilan_excel(request):
    """Export Excel du Bilan"""
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    
    actif_immobilise, actif_circulant, capitaux_propres, dettes, total_actif, total_passif, entreprise, exercice = _get_bilan_data(request)
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Bilan"
    
    header_font = Font(bold=True, size=14)
    actif_fill = PatternFill(start_color="CFE2FF", end_color="CFE2FF", fill_type="solid")
    passif_fill = PatternFill(start_color="D1E7DD", end_color="D1E7DD", fill_type="solid")
    section_fill = PatternFill(start_color="E8E8E8", end_color="E8E8E8", fill_type="solid")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    ws.merge_cells('A1:F1')
    ws['A1'] = entreprise.nom_entreprise
    ws['A1'].font = header_font
    ws['A1'].alignment = Alignment(horizontal='center')
    
    ws.merge_cells('A2:F2')
    ws['A2'] = f"BILAN COMPTABLE - Arrêté au {timezone.now().strftime('%d/%m/%Y')}"
    ws['A2'].alignment = Alignment(horizontal='center')
    
    # ACTIF
    row = 4
    ws.merge_cells(f'A{row}:C{row}')
    ws[f'A{row}'] = 'ACTIF'
    ws[f'A{row}'].font = Font(bold=True, color='FFFFFF')
    ws[f'A{row}'].fill = PatternFill(start_color="0D6EFD", end_color="0D6EFD", fill_type="solid")
    
    ws.merge_cells(f'D{row}:F{row}')
    ws[f'D{row}'] = 'PASSIF'
    ws[f'D{row}'].font = Font(bold=True, color='FFFFFF')
    ws[f'D{row}'].fill = PatternFill(start_color="198754", end_color="198754", fill_type="solid")
    row += 1
    
    # Headers
    for col, header in enumerate(['Compte', 'Intitulé', 'Montant'], 1):
        ws.cell(row=row, column=col, value=header).font = Font(bold=True)
        ws.cell(row=row, column=col).fill = actif_fill
    for col, header in enumerate(['Compte', 'Intitulé', 'Montant'], 4):
        ws.cell(row=row, column=col, value=header).font = Font(bold=True)
        ws.cell(row=row, column=col).fill = passif_fill
    row += 1
    
    # Actif immobilisé / Capitaux propres
    ws.merge_cells(f'A{row}:C{row}')
    ws[f'A{row}'] = 'ACTIF IMMOBILISÉ'
    ws[f'A{row}'].fill = section_fill
    ws.merge_cells(f'D{row}:F{row}')
    ws[f'D{row}'] = 'CAPITAUX PROPRES'
    ws[f'D{row}'].fill = section_fill
    row += 1
    
    max_rows = max(len(actif_immobilise), len(capitaux_propres))
    for i in range(max_rows):
        if i < len(actif_immobilise):
            ws.cell(row=row, column=1, value=actif_immobilise[i]['numero_compte'])
            ws.cell(row=row, column=2, value=actif_immobilise[i]['intitule'])
            ws.cell(row=row, column=3, value=float(actif_immobilise[i]['solde'])).number_format = '#,##0'
        if i < len(capitaux_propres):
            ws.cell(row=row, column=4, value=capitaux_propres[i]['numero_compte'])
            ws.cell(row=row, column=5, value=capitaux_propres[i]['intitule'])
            ws.cell(row=row, column=6, value=float(capitaux_propres[i]['solde'])).number_format = '#,##0'
        row += 1
    
    # Actif circulant / Dettes
    ws.merge_cells(f'A{row}:C{row}')
    ws[f'A{row}'] = 'ACTIF CIRCULANT'
    ws[f'A{row}'].fill = section_fill
    ws.merge_cells(f'D{row}:F{row}')
    ws[f'D{row}'] = 'DETTES'
    ws[f'D{row}'].fill = section_fill
    row += 1
    
    max_rows = max(len(actif_circulant), len(dettes))
    for i in range(max_rows):
        if i < len(actif_circulant):
            ws.cell(row=row, column=1, value=actif_circulant[i]['numero_compte'])
            ws.cell(row=row, column=2, value=actif_circulant[i]['intitule'])
            ws.cell(row=row, column=3, value=float(actif_circulant[i]['solde'])).number_format = '#,##0'
        if i < len(dettes):
            ws.cell(row=row, column=4, value=dettes[i]['numero_compte'])
            ws.cell(row=row, column=5, value=dettes[i]['intitule'])
            ws.cell(row=row, column=6, value=float(dettes[i]['solde'])).number_format = '#,##0'
        row += 1
    
    # Totaux
    row += 1
    ws.cell(row=row, column=2, value='TOTAL ACTIF').font = Font(bold=True)
    ws.cell(row=row, column=3, value=float(total_actif)).number_format = '#,##0'
    ws.cell(row=row, column=3).font = Font(bold=True)
    ws.cell(row=row, column=5, value='TOTAL PASSIF').font = Font(bold=True)
    ws.cell(row=row, column=6, value=float(total_passif)).number_format = '#,##0'
    ws.cell(row=row, column=6).font = Font(bold=True)
    
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 30
    ws.column_dimensions['F'].width = 15
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="bilan_{timezone.now().strftime("%Y%m%d")}.xlsx"'
    wb.save(response)
    return response


@login_required
@compta_required
def bilan_pdf(request):
    """Export PDF du Bilan"""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    
    actif_immobilise, actif_circulant, capitaux_propres, dettes, total_actif, total_passif, entreprise, exercice = _get_bilan_data(request)
    
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), leftMargin=1*cm, rightMargin=1*cm, topMargin=1*cm, bottomMargin=1*cm)
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('Title', parent=styles['Heading1'], fontSize=16, alignment=1, spaceAfter=10)
    subtitle_style = ParagraphStyle('Subtitle', parent=styles['Normal'], fontSize=10, alignment=1, spaceAfter=20)
    
    elements = []
    elements.append(Paragraph(entreprise.nom_entreprise, title_style))
    elements.append(Paragraph(f"BILAN COMPTABLE - Arrêté au {timezone.now().strftime('%d/%m/%Y')}", subtitle_style))
    
    # Build table with ACTIF and PASSIF side by side
    data = [['ACTIF', '', '', 'PASSIF', '', ''],
            ['Compte', 'Intitulé', 'Montant', 'Compte', 'Intitulé', 'Montant'],
            ['ACTIF IMMOBILISÉ', '', '', 'CAPITAUX PROPRES', '', '']]
    
    max_rows = max(len(actif_immobilise), len(capitaux_propres))
    for i in range(max_rows):
        row = ['', '', '', '', '', '']
        if i < len(actif_immobilise):
            row[0] = actif_immobilise[i]['numero_compte']
            row[1] = actif_immobilise[i]['intitule'][:25]
            row[2] = f"{actif_immobilise[i]['solde']:,.0f}"
        if i < len(capitaux_propres):
            row[3] = capitaux_propres[i]['numero_compte']
            row[4] = capitaux_propres[i]['intitule'][:25]
            row[5] = f"{capitaux_propres[i]['solde']:,.0f}"
        data.append(row)
    
    data.append(['ACTIF CIRCULANT', '', '', 'DETTES', '', ''])
    
    max_rows = max(len(actif_circulant), len(dettes))
    for i in range(max_rows):
        row = ['', '', '', '', '', '']
        if i < len(actif_circulant):
            row[0] = actif_circulant[i]['numero_compte']
            row[1] = actif_circulant[i]['intitule'][:25]
            row[2] = f"{actif_circulant[i]['solde']:,.0f}"
        if i < len(dettes):
            row[3] = dettes[i]['numero_compte']
            row[4] = dettes[i]['intitule'][:25]
            row[5] = f"{dettes[i]['solde']:,.0f}"
        data.append(row)
    
    data.append(['', 'TOTAL ACTIF', f"{total_actif:,.0f}", '', 'TOTAL PASSIF', f"{total_passif:,.0f}"])
    
    table = Table(data, colWidths=[2*cm, 5*cm, 3*cm, 2*cm, 5*cm, 3*cm])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (2, 0), colors.HexColor('#0D6EFD')),
        ('BACKGROUND', (3, 0), (5, 0), colors.HexColor('#198754')),
        ('TEXTCOLOR', (0, 0), (5, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('ALIGN', (2, 0), (2, -1), 'RIGHT'),
        ('ALIGN', (5, 0), (5, -1), 'RIGHT'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#E8E8E8')),
    ]))
    
    elements.append(table)
    doc.build(elements)
    buffer.seek(0)
    
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="bilan_{timezone.now().strftime("%Y%m%d")}.pdf"'
    return response


# ============ EXPORTS COMPTE DE RESULTAT ============

def _get_compte_resultat_data(request):
    """Helper pour récupérer les données du compte de résultat"""
    entreprise = request.user.entreprise
    exercice = ExerciceComptable.objects.filter(entreprise=entreprise, est_courant=True).first()
    
    def get_comptes_classe(classe_prefix):
        comptes = PlanComptable.objects.filter(
            entreprise=entreprise, numero_compte__startswith=classe_prefix, est_actif=True
        ).annotate(
            total_debit=Sum('lignes_ecritures__montant_debit'),
            total_credit=Sum('lignes_ecritures__montant_credit')
        ).order_by('numero_compte')
        
        result = []
        for c in comptes:
            d = c.total_debit or Decimal('0')
            cr = c.total_credit or Decimal('0')
            solde = d - cr if d > cr else cr - d
            if solde > 0:
                result.append({'numero_compte': c.numero_compte, 'intitule': c.intitule, 'solde': solde})
        return result
    
    charges_exploitation = get_comptes_classe('60') + get_comptes_classe('61') + get_comptes_classe('62') + get_comptes_classe('63') + get_comptes_classe('64') + get_comptes_classe('65')
    charges_financieres = get_comptes_classe('66') + get_comptes_classe('67')
    produits_exploitation = get_comptes_classe('70') + get_comptes_classe('71') + get_comptes_classe('72') + get_comptes_classe('73') + get_comptes_classe('74') + get_comptes_classe('75')
    produits_financiers = get_comptes_classe('76') + get_comptes_classe('77')
    
    total_charges = sum(c['solde'] for c in charges_exploitation) + sum(c['solde'] for c in charges_financieres)
    total_produits = sum(c['solde'] for c in produits_exploitation) + sum(c['solde'] for c in produits_financiers)
    resultat = total_produits - total_charges
    
    return charges_exploitation, charges_financieres, produits_exploitation, produits_financiers, total_charges, total_produits, resultat, entreprise, exercice


@login_required
@compta_required
def compte_resultat_excel(request):
    """Export Excel du Compte de Résultat"""
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    
    charges_exploitation, charges_financieres, produits_exploitation, produits_financiers, total_charges, total_produits, resultat, entreprise, exercice = _get_compte_resultat_data(request)
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Compte de Résultat"
    
    header_font = Font(bold=True, size=14)
    charges_fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
    produits_fill = PatternFill(start_color="D1E7DD", end_color="D1E7DD", fill_type="solid")
    section_fill = PatternFill(start_color="E8E8E8", end_color="E8E8E8", fill_type="solid")
    
    ws.merge_cells('A1:F1')
    ws['A1'] = entreprise.nom_entreprise
    ws['A1'].font = header_font
    ws['A1'].alignment = Alignment(horizontal='center')
    
    ws.merge_cells('A2:F2')
    ws['A2'] = f"COMPTE DE RÉSULTAT - Au {timezone.now().strftime('%d/%m/%Y')}"
    ws['A2'].alignment = Alignment(horizontal='center')
    
    row = 4
    # Headers
    ws.merge_cells(f'A{row}:C{row}')
    ws[f'A{row}'] = 'CHARGES'
    ws[f'A{row}'].font = Font(bold=True, color='FFFFFF')
    ws[f'A{row}'].fill = PatternFill(start_color="DC3545", end_color="DC3545", fill_type="solid")
    
    ws.merge_cells(f'D{row}:F{row}')
    ws[f'D{row}'] = 'PRODUITS'
    ws[f'D{row}'].font = Font(bold=True, color='FFFFFF')
    ws[f'D{row}'].fill = PatternFill(start_color="198754", end_color="198754", fill_type="solid")
    row += 1
    
    for col, header in enumerate(['Compte', 'Intitulé', 'Montant'], 1):
        ws.cell(row=row, column=col, value=header).font = Font(bold=True)
        ws.cell(row=row, column=col).fill = charges_fill
    for col, header in enumerate(['Compte', 'Intitulé', 'Montant'], 4):
        ws.cell(row=row, column=col, value=header).font = Font(bold=True)
        ws.cell(row=row, column=col).fill = produits_fill
    row += 1
    
    # Exploitation
    ws.merge_cells(f'A{row}:C{row}')
    ws[f'A{row}'] = "Charges d'exploitation"
    ws[f'A{row}'].fill = section_fill
    ws.merge_cells(f'D{row}:F{row}')
    ws[f'D{row}'] = "Produits d'exploitation"
    ws[f'D{row}'].fill = section_fill
    row += 1
    
    max_rows = max(len(charges_exploitation), len(produits_exploitation))
    for i in range(max_rows):
        if i < len(charges_exploitation):
            ws.cell(row=row, column=1, value=charges_exploitation[i]['numero_compte'])
            ws.cell(row=row, column=2, value=charges_exploitation[i]['intitule'])
            ws.cell(row=row, column=3, value=float(charges_exploitation[i]['solde'])).number_format = '#,##0'
        if i < len(produits_exploitation):
            ws.cell(row=row, column=4, value=produits_exploitation[i]['numero_compte'])
            ws.cell(row=row, column=5, value=produits_exploitation[i]['intitule'])
            ws.cell(row=row, column=6, value=float(produits_exploitation[i]['solde'])).number_format = '#,##0'
        row += 1
    
    # Financiers
    ws.merge_cells(f'A{row}:C{row}')
    ws[f'A{row}'] = "Charges financières"
    ws[f'A{row}'].fill = section_fill
    ws.merge_cells(f'D{row}:F{row}')
    ws[f'D{row}'] = "Produits financiers"
    ws[f'D{row}'].fill = section_fill
    row += 1
    
    max_rows = max(len(charges_financieres), len(produits_financiers))
    for i in range(max_rows):
        if i < len(charges_financieres):
            ws.cell(row=row, column=1, value=charges_financieres[i]['numero_compte'])
            ws.cell(row=row, column=2, value=charges_financieres[i]['intitule'])
            ws.cell(row=row, column=3, value=float(charges_financieres[i]['solde'])).number_format = '#,##0'
        if i < len(produits_financiers):
            ws.cell(row=row, column=4, value=produits_financiers[i]['numero_compte'])
            ws.cell(row=row, column=5, value=produits_financiers[i]['intitule'])
            ws.cell(row=row, column=6, value=float(produits_financiers[i]['solde'])).number_format = '#,##0'
        row += 1
    
    # Totaux
    row += 1
    ws.cell(row=row, column=2, value='TOTAL CHARGES').font = Font(bold=True)
    ws.cell(row=row, column=3, value=float(total_charges)).number_format = '#,##0'
    ws.cell(row=row, column=3).font = Font(bold=True)
    ws.cell(row=row, column=5, value='TOTAL PRODUITS').font = Font(bold=True)
    ws.cell(row=row, column=6, value=float(total_produits)).number_format = '#,##0'
    ws.cell(row=row, column=6).font = Font(bold=True)
    
    row += 2
    ws.merge_cells(f'A{row}:F{row}')
    ws[f'A{row}'] = f"RÉSULTAT: {resultat:,.0f} GNF ({'Bénéfice' if resultat >= 0 else 'Perte'})"
    ws[f'A{row}'].font = Font(bold=True, size=12)
    ws[f'A{row}'].alignment = Alignment(horizontal='center')
    
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 30
    ws.column_dimensions['F'].width = 15
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="compte_resultat_{timezone.now().strftime("%Y%m%d")}.xlsx"'
    wb.save(response)
    return response


@login_required
@compta_required
def compte_resultat_pdf(request):
    """Export PDF du Compte de Résultat"""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    
    charges_exploitation, charges_financieres, produits_exploitation, produits_financiers, total_charges, total_produits, resultat, entreprise, exercice = _get_compte_resultat_data(request)
    
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), leftMargin=1*cm, rightMargin=1*cm, topMargin=1*cm, bottomMargin=1*cm)
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('Title', parent=styles['Heading1'], fontSize=16, alignment=1, spaceAfter=10)
    subtitle_style = ParagraphStyle('Subtitle', parent=styles['Normal'], fontSize=10, alignment=1, spaceAfter=20)
    
    elements = []
    elements.append(Paragraph(entreprise.nom_entreprise, title_style))
    elements.append(Paragraph(f"COMPTE DE RÉSULTAT - Au {timezone.now().strftime('%d/%m/%Y')}", subtitle_style))
    
    data = [['CHARGES', '', '', 'PRODUITS', '', ''],
            ['Compte', 'Intitulé', 'Montant', 'Compte', 'Intitulé', 'Montant'],
            ["Charges d'exploitation", '', '', "Produits d'exploitation", '', '']]
    
    max_rows = max(len(charges_exploitation), len(produits_exploitation))
    for i in range(max_rows):
        row = ['', '', '', '', '', '']
        if i < len(charges_exploitation):
            row[0] = charges_exploitation[i]['numero_compte']
            row[1] = charges_exploitation[i]['intitule'][:25]
            row[2] = f"{charges_exploitation[i]['solde']:,.0f}"
        if i < len(produits_exploitation):
            row[3] = produits_exploitation[i]['numero_compte']
            row[4] = produits_exploitation[i]['intitule'][:25]
            row[5] = f"{produits_exploitation[i]['solde']:,.0f}"
        data.append(row)
    
    data.append(['Charges financières', '', '', 'Produits financiers', '', ''])
    
    max_rows = max(len(charges_financieres), len(produits_financiers))
    for i in range(max_rows):
        row = ['', '', '', '', '', '']
        if i < len(charges_financieres):
            row[0] = charges_financieres[i]['numero_compte']
            row[1] = charges_financieres[i]['intitule'][:25]
            row[2] = f"{charges_financieres[i]['solde']:,.0f}"
        if i < len(produits_financiers):
            row[3] = produits_financiers[i]['numero_compte']
            row[4] = produits_financiers[i]['intitule'][:25]
            row[5] = f"{produits_financiers[i]['solde']:,.0f}"
        data.append(row)
    
    data.append(['', 'TOTAL CHARGES', f"{total_charges:,.0f}", '', 'TOTAL PRODUITS', f"{total_produits:,.0f}"])
    
    table = Table(data, colWidths=[2*cm, 5*cm, 3*cm, 2*cm, 5*cm, 3*cm])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (2, 0), colors.HexColor('#DC3545')),
        ('BACKGROUND', (3, 0), (5, 0), colors.HexColor('#198754')),
        ('TEXTCOLOR', (0, 0), (5, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('ALIGN', (2, 0), (2, -1), 'RIGHT'),
        ('ALIGN', (5, 0), (5, -1), 'RIGHT'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#E8E8E8')),
    ]))
    
    elements.append(table)
    elements.append(Spacer(1, 1*cm))
    
    resultat_text = f"RÉSULTAT: {resultat:,.0f} GNF ({'Bénéfice' if resultat >= 0 else 'Perte'})"
    resultat_style = ParagraphStyle('Resultat', parent=styles['Heading2'], fontSize=14, alignment=1)
    elements.append(Paragraph(resultat_text, resultat_style))
    
    doc.build(elements)
    buffer.seek(0)
    
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="compte_resultat_{timezone.now().strftime("%Y%m%d")}.pdf"'
    return response


# ============ MODULE CLIENTS & FOURNISSEURS DÉTAILLÉS ============

@login_required
@compta_required
def compte_client_list(request):
    """Liste des comptes clients"""
    entreprise = request.user.entreprise
    
    clients = Tiers.objects.filter(
        entreprise=entreprise,
        type_tiers__in=['client', 'mixte'],
        est_actif=True
    ).annotate(
        total_factures=Sum('factures__montant_ttc', filter=Q(factures__type_facture='vente', factures__statut='validee')),
        total_paye=Sum('factures__montant_paye', filter=Q(factures__type_facture='vente', factures__statut='validee')),
        nb_factures=Count('factures', filter=Q(factures__type_facture='vente'))
    ).order_by('raison_sociale')
    
    # Calculer le solde pour chaque client
    clients_list = list(clients)
    for client in clients_list:
        client.solde_calcule = (client.total_factures or Decimal('0')) - (client.total_paye or Decimal('0'))
    
    total_creances = sum(c.solde_calcule for c in clients_list)
    nb_clients = len(clients_list)
    creance_moyenne = total_creances / nb_clients if nb_clients > 0 else Decimal('0')
    
    context = {
        'clients': clients_list,
        'total_creances': total_creances,
        'nb_clients': nb_clients,
        'creance_moyenne': creance_moyenne,
    }
    return render(request, 'comptabilite/clients_fournisseurs/compte_client_list.html', context)


@login_required
@compta_required
def compte_client_detail(request, pk):
    """Détail du compte client avec historique des mouvements"""
    entreprise = request.user.entreprise
    client = get_object_or_404(Tiers, pk=pk, entreprise=entreprise, type_tiers__in=['client', 'mixte'])
    
    factures = Facture.objects.filter(
        tiers=client,
        type_facture='vente',
        statut__in=['validee', 'payee']
    ).order_by('-date_facture')
    
    reglements = Reglement.objects.filter(
        facture__tiers=client,
        facture__type_facture='vente'
    ).order_by('-date_reglement')
    
    # Construire l'historique des mouvements
    mouvements = []
    solde_cumule = Decimal('0')
    
    for facture in factures:
        solde_cumule += facture.montant_ttc
        mouvements.append({
            'date': facture.date_facture,
            'type': 'Facture',
            'reference': facture.numero,
            'libelle': f"Facture {facture.numero}",
            'debit': facture.montant_ttc,
            'credit': Decimal('0'),
            'solde': solde_cumule,
        })
    
    for reglement in reglements:
        solde_cumule -= reglement.montant
        mouvements.append({
            'date': reglement.date_reglement,
            'type': 'Règlement',
            'reference': reglement.numero,
            'libelle': f"Règlement {reglement.mode_paiement}",
            'debit': Decimal('0'),
            'credit': reglement.montant,
            'solde': solde_cumule,
        })
    
    # Trier par date
    mouvements.sort(key=lambda x: x['date'])
    
    # Recalculer le solde cumulé dans l'ordre
    solde_cumule = Decimal('0')
    for mvt in mouvements:
        solde_cumule += mvt['debit'] - mvt['credit']
        mvt['solde'] = solde_cumule
    
    total_debit = sum(m['debit'] for m in mouvements)
    total_credit = sum(m['credit'] for m in mouvements)
    
    context = {
        'client': client,
        'mouvements': mouvements,
        'total_debit': total_debit,
        'total_credit': total_credit,
        'solde_final': solde_cumule,
    }
    return render(request, 'comptabilite/clients_fournisseurs/compte_client_detail.html', context)


@login_required
@compta_required
def compte_fournisseur_list(request):
    """Liste des comptes fournisseurs"""
    entreprise = request.user.entreprise
    
    fournisseurs = Tiers.objects.filter(
        entreprise=entreprise,
        type_tiers__in=['fournisseur', 'mixte'],
        est_actif=True
    ).annotate(
        total_factures=Sum('factures__montant_ttc', filter=Q(factures__type_facture='achat', factures__statut='validee')),
        total_paye=Sum('factures__montant_paye', filter=Q(factures__type_facture='achat', factures__statut='validee')),
        nb_factures=Count('factures', filter=Q(factures__type_facture='achat'))
    ).order_by('raison_sociale')
    
    fournisseurs_list = list(fournisseurs)
    for fournisseur in fournisseurs_list:
        fournisseur.solde_calcule = (fournisseur.total_factures or Decimal('0')) - (fournisseur.total_paye or Decimal('0'))
    
    total_dettes = sum(f.solde_calcule for f in fournisseurs_list)
    nb_fournisseurs = len(fournisseurs_list)
    dette_moyenne = total_dettes / nb_fournisseurs if nb_fournisseurs > 0 else Decimal('0')
    
    context = {
        'fournisseurs': fournisseurs_list,
        'total_dettes': total_dettes,
        'nb_fournisseurs': nb_fournisseurs,
        'dette_moyenne': dette_moyenne,
    }
    return render(request, 'comptabilite/clients_fournisseurs/compte_fournisseur_list.html', context)


@login_required
@compta_required
def compte_fournisseur_detail(request, pk):
    """Détail du compte fournisseur avec historique des mouvements"""
    entreprise = request.user.entreprise
    fournisseur = get_object_or_404(Tiers, pk=pk, entreprise=entreprise, type_tiers__in=['fournisseur', 'mixte'])
    
    factures = Facture.objects.filter(
        tiers=fournisseur,
        type_facture='achat',
        statut__in=['validee', 'payee']
    ).order_by('-date_facture')
    
    reglements = Reglement.objects.filter(
        facture__tiers=fournisseur,
        facture__type_facture='achat'
    ).order_by('-date_reglement')
    
    mouvements = []
    solde_cumule = Decimal('0')
    
    for facture in factures:
        solde_cumule += facture.montant_ttc
        mouvements.append({
            'date': facture.date_facture,
            'type': 'Facture',
            'reference': facture.numero,
            'libelle': f"Facture {facture.numero}",
            'debit': Decimal('0'),
            'credit': facture.montant_ttc,
            'solde': solde_cumule,
        })
    
    for reglement in reglements:
        solde_cumule -= reglement.montant
        mouvements.append({
            'date': reglement.date_reglement,
            'type': 'Règlement',
            'reference': reglement.numero,
            'libelle': f"Règlement {reglement.mode_paiement}",
            'debit': reglement.montant,
            'credit': Decimal('0'),
            'solde': solde_cumule,
        })
    
    mouvements.sort(key=lambda x: x['date'])
    
    solde_cumule = Decimal('0')
    for mvt in mouvements:
        solde_cumule += mvt['credit'] - mvt['debit']
        mvt['solde'] = solde_cumule
    
    total_debit = sum(m['debit'] for m in mouvements)
    total_credit = sum(m['credit'] for m in mouvements)
    
    context = {
        'fournisseur': fournisseur,
        'mouvements': mouvements,
        'total_debit': total_debit,
        'total_credit': total_credit,
        'solde_final': solde_cumule,
    }
    return render(request, 'comptabilite/clients_fournisseurs/compte_fournisseur_detail.html', context)


@login_required
@compta_required
def vieillissement_creances(request):
    """Analyse du vieillissement des créances clients"""
    entreprise = request.user.entreprise
    today = timezone.now().date()
    
    # Récupérer les factures clients non totalement payées
    factures = Facture.objects.filter(
        entreprise=entreprise,
        type_facture='vente',
        statut='validee'
    ).exclude(
        montant_paye__gte=F('montant_ttc')
    ).select_related('tiers').order_by('tiers__raison_sociale', 'date_echeance')
    
    # Classifier par tranche d'âge
    tranches = {
        'non_echu': {'label': 'Non échu', 'factures': [], 'total': Decimal('0')},
        '0_30': {'label': '0-30 jours', 'factures': [], 'total': Decimal('0')},
        '31_60': {'label': '31-60 jours', 'factures': [], 'total': Decimal('0')},
        '61_90': {'label': '61-90 jours', 'factures': [], 'total': Decimal('0')},
        'plus_90': {'label': '> 90 jours', 'factures': [], 'total': Decimal('0')},
    }
    
    for facture in factures:
        reste = facture.montant_ttc - facture.montant_paye
        date_ref = facture.date_echeance or facture.date_facture
        jours_retard = (today - date_ref).days
        
        facture.reste_a_payer = reste
        facture.jours_retard = jours_retard
        
        if jours_retard <= 0:
            tranches['non_echu']['factures'].append(facture)
            tranches['non_echu']['total'] += reste
        elif jours_retard <= 30:
            tranches['0_30']['factures'].append(facture)
            tranches['0_30']['total'] += reste
        elif jours_retard <= 60:
            tranches['31_60']['factures'].append(facture)
            tranches['31_60']['total'] += reste
        elif jours_retard <= 90:
            tranches['61_90']['factures'].append(facture)
            tranches['61_90']['total'] += reste
        else:
            tranches['plus_90']['factures'].append(facture)
            tranches['plus_90']['total'] += reste
    
    total_creances = sum(t['total'] for t in tranches.values())
    
    context = {
        'tranches': tranches,
        'total_creances': total_creances,
        'today': today,
    }
    return render(request, 'comptabilite/clients_fournisseurs/vieillissement_creances.html', context)


@login_required
@compta_required
def vieillissement_dettes(request):
    """Analyse du vieillissement des dettes fournisseurs"""
    entreprise = request.user.entreprise
    today = timezone.now().date()
    
    factures = Facture.objects.filter(
        entreprise=entreprise,
        type_facture='achat',
        statut='validee'
    ).exclude(
        montant_paye__gte=F('montant_ttc')
    ).select_related('tiers').order_by('tiers__raison_sociale', 'date_echeance')
    
    tranches = {
        'non_echu': {'label': 'Non échu', 'factures': [], 'total': Decimal('0')},
        '0_30': {'label': '0-30 jours', 'factures': [], 'total': Decimal('0')},
        '31_60': {'label': '31-60 jours', 'factures': [], 'total': Decimal('0')},
        '61_90': {'label': '61-90 jours', 'factures': [], 'total': Decimal('0')},
        'plus_90': {'label': '> 90 jours', 'factures': [], 'total': Decimal('0')},
    }
    
    for facture in factures:
        reste = facture.montant_ttc - facture.montant_paye
        date_ref = facture.date_echeance or facture.date_facture
        jours_retard = (today - date_ref).days
        
        facture.reste_a_payer = reste
        facture.jours_retard = jours_retard
        
        if jours_retard <= 0:
            tranches['non_echu']['factures'].append(facture)
            tranches['non_echu']['total'] += reste
        elif jours_retard <= 30:
            tranches['0_30']['factures'].append(facture)
            tranches['0_30']['total'] += reste
        elif jours_retard <= 60:
            tranches['31_60']['factures'].append(facture)
            tranches['31_60']['total'] += reste
        elif jours_retard <= 90:
            tranches['61_90']['factures'].append(facture)
            tranches['61_90']['total'] += reste
        else:
            tranches['plus_90']['factures'].append(facture)
            tranches['plus_90']['total'] += reste
    
    total_dettes = sum(t['total'] for t in tranches.values())
    
    context = {
        'tranches': tranches,
        'total_dettes': total_dettes,
        'today': today,
    }
    return render(request, 'comptabilite/clients_fournisseurs/vieillissement_dettes.html', context)


@login_required
@compta_required
def impayes_clients(request):
    """Liste des factures clients impayées"""
    entreprise = request.user.entreprise
    today = timezone.now().date()
    
    factures = Facture.objects.filter(
        entreprise=entreprise,
        type_facture='vente',
        statut='validee'
    ).exclude(
        montant_paye__gte=F('montant_ttc')
    ).select_related('tiers').order_by('-date_facture')
    
    for facture in factures:
        facture.reste_a_payer_calc = facture.montant_ttc - facture.montant_paye
        date_ref = facture.date_echeance or facture.date_facture
        facture.jours_retard = (today - date_ref).days
        facture.est_en_retard = facture.jours_retard > 0
    
    total_impayes = sum(f.reste_a_payer_calc for f in factures)
    nb_impayes = factures.count()
    
    context = {
        'factures': factures,
        'total_impayes': total_impayes,
        'nb_impayes': nb_impayes,
        'today': today,
    }
    return render(request, 'comptabilite/clients_fournisseurs/impayes_clients.html', context)


@login_required
@compta_required
def impayes_fournisseurs(request):
    """Liste des factures fournisseurs impayées"""
    entreprise = request.user.entreprise
    today = timezone.now().date()
    
    factures = Facture.objects.filter(
        entreprise=entreprise,
        type_facture='achat',
        statut='validee'
    ).exclude(
        montant_paye__gte=F('montant_ttc')
    ).select_related('tiers').order_by('-date_facture')
    
    for facture in factures:
        facture.reste_a_payer_calc = facture.montant_ttc - facture.montant_paye
        date_ref = facture.date_echeance or facture.date_facture
        facture.jours_retard = (today - date_ref).days
        facture.est_en_retard = facture.jours_retard > 0
    
    total_impayes = sum(f.reste_a_payer_calc for f in factures)
    nb_impayes = factures.count()
    
    context = {
        'factures': factures,
        'total_impayes': total_impayes,
        'nb_impayes': nb_impayes,
        'today': today,
    }
    return render(request, 'comptabilite/clients_fournisseurs/impayes_fournisseurs.html', context)
