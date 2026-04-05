"""
Vues pour les rapports de paie avancés.
- Rapport masse salariale
- État de paie (récapitulatif mensuel détaillé)
- Feuille de présence (grille journalière)
- Export Excel / PDF pour chaque rapport
"""
import io
import json
import calendar
from collections import defaultdict
from datetime import date, timedelta
from decimal import Decimal

from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse, JsonResponse
from django.db.models import Sum, Count, Avg, Q, F, Prefetch
from django.utils import timezone

from .models import BulletinPaie, PeriodePaie, LigneBulletin, RubriquePaie, AvanceSalaire
from employes.models import Employe
from core.models import Service
from core.decorators import reauth_required, entreprise_active_required

# Imports optionnels
try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False

MOIS_FR = ['', 'Janvier', 'Février', 'Mars', 'Avril', 'Mai', 'Juin',
           'Juillet', 'Août', 'Septembre', 'Octobre', 'Novembre', 'Décembre']


def _get_bulletins_filtrés(entreprise, annee, periode_type, mois=None, trimestre=None, service_id=None):
    """Retourne un queryset de bulletins selon les filtres."""
    qs = BulletinPaie.objects.filter(
        employe__entreprise=entreprise,
        annee_paie=annee,
        statut_bulletin__in=['calcule', 'valide', 'paye'],
    )
    if periode_type == 'mois' and mois:
        qs = qs.filter(mois_paie=mois)
    elif periode_type == 'trimestre' and trimestre:
        t = int(trimestre)
        mois_debut = (t - 1) * 3 + 1
        mois_fin = t * 3
        qs = qs.filter(mois_paie__gte=mois_debut, mois_paie__lte=mois_fin)
    # Filtre service
    if service_id:
        qs = qs.filter(employe__service_id=service_id)
    return qs


def _calcul_rapport(bulletins_qs):
    """Calcule les agrégats principaux à partir d'un queryset de bulletins."""
    agg = bulletins_qs.aggregate(
        nb=Count('id'),
        brut=Sum('salaire_brut'),
        cnss_sal=Sum('cnss_employe'),
        cnss_emp=Sum('cnss_employeur'),
        rts=Sum('irg'),
        net=Sum('net_a_payer'),
        vf=Sum('versement_forfaitaire'),
        ta=Sum('taxe_apprentissage'),
    )
    nb = agg['nb'] or 0
    brut = agg['brut'] or Decimal('0')
    cnss_sal = agg['cnss_sal'] or Decimal('0')
    cnss_emp = agg['cnss_emp'] or Decimal('0')
    rts = agg['rts'] or Decimal('0')
    net = agg['net'] or Decimal('0')
    vf = agg['vf'] or Decimal('0')
    ta = agg['ta'] or Decimal('0')
    charges_totales = cnss_sal + cnss_emp + rts + vf + ta
    ratio_charges = (charges_totales / brut * 100) if brut else Decimal('0')
    cout_moyen = (brut / nb) if nb else Decimal('0')
    return {
        'nb': nb,
        'brut': brut,
        'cnss_sal': cnss_sal,
        'cnss_emp': cnss_emp,
        'rts': rts,
        'net': net,
        'vf': vf,
        'ta': ta,
        'charges_totales': charges_totales,
        'ratio_charges': ratio_charges,
        'cout_moyen': cout_moyen,
    }


@login_required
@entreprise_active_required
@reauth_required
def rapport_masse_salariale(request):
    """Rapport de masse salariale avec filtres et graphiques."""
    entreprise = request.user.entreprise
    annee_courante = date.today().year

    # Paramètres de filtres
    annee = int(request.GET.get('annee', annee_courante))
    periode_type = request.GET.get('periode_type', 'annee')  # mois | trimestre | annee
    mois = request.GET.get('mois', '')
    trimestre = request.GET.get('trimestre', '')
    service_id = request.GET.get('service', '')

    # Services disponibles
    services = Service.objects.filter(
        entreprise=entreprise,
        actif=True,
    ).order_by('nom_service')

    # Bulletins filtrés
    bulletins = _get_bulletins_filtrés(
        entreprise, annee, periode_type,
        mois=mois or None,
        trimestre=trimestre or None,
        service_id=service_id or None,
    )

    # --- Tableau par service ---
    par_service = {}
    for b in bulletins.select_related('employe__service'):
        svc = b.employe.service
        svc_key = svc.pk if svc else 0
        svc_nom = str(svc) if svc else 'Sans service'
        if svc_key not in par_service:
            par_service[svc_key] = {
                'nom': svc_nom,
                'nb': 0,
                'brut': Decimal('0'),
                'cnss_sal': Decimal('0'),
                'rts': Decimal('0'),
                'net': Decimal('0'),
                'cnss_emp': Decimal('0'),
            }
        r = par_service[svc_key]
        r['nb'] += 1
        r['brut'] += b.salaire_brut or Decimal('0')
        r['cnss_sal'] += b.cnss_employe or Decimal('0')
        r['rts'] += b.irg or Decimal('0')
        r['net'] += b.net_a_payer or Decimal('0')
        r['cnss_emp'] += b.cnss_employeur or Decimal('0')

    tableau_services = sorted(par_service.values(), key=lambda x: -x['brut'])

    # --- Évolution mensuelle (tous les mois de l'année) ---
    evolution_data = []
    labels_mois = []
    for m in range(1, 13):
        bm = BulletinPaie.objects.filter(
            employe__entreprise=entreprise,
            annee_paie=annee,
            mois_paie=m,
            statut_bulletin__in=['calcule', 'valide', 'paye'],
        )
        if service_id:
            bm = bm.filter(employe__service_id=service_id)
        agg_m = bm.aggregate(brut=Sum('salaire_brut'), net=Sum('net_a_payer'), nb=Count('id'))
        evolution_data.append({
            'mois': m,
            'label': MOIS_FR[m][:3],
            'brut': float(agg_m['brut'] or 0),
            'net': float(agg_m['net'] or 0),
            'nb': agg_m['nb'] or 0,
        })
        labels_mois.append(MOIS_FR[m][:3])

    # --- Top 10 coûts salariaux ---
    top10 = bulletins.order_by('-salaire_brut').select_related('employe')[:10]
    top10_data = []
    for b in top10:
        top10_data.append({
            'employe': f"{b.employe.nom} {b.employe.prenoms}",
            'matricule': b.employe.matricule,
            'service': str(b.employe.service) if b.employe.service else '-',
            'brut': b.salaire_brut or Decimal('0'),
            'net': b.net_a_payer or Decimal('0'),
        })

    # --- Indicateurs globaux ---
    indicateurs = _calcul_rapport(bulletins)

    # --- Comparaison N vs N-1 ---
    bulletins_n1 = BulletinPaie.objects.filter(
        employe__entreprise=entreprise,
        annee_paie=annee - 1,
        statut_bulletin__in=['calcule', 'valide', 'paye'],
    )
    if service_id:
        bulletins_n1 = bulletins_n1.filter(employe__service_id=service_id)
    indic_n1 = _calcul_rapport(bulletins_n1)
    evol_brut = None
    if indic_n1['brut'] and indic_n1['brut'] > 0:
        evol_brut = float((indicateurs['brut'] - indic_n1['brut']) / indic_n1['brut'] * 100)

    # Années disponibles
    annees_dispo = BulletinPaie.objects.filter(
        employe__entreprise=entreprise
    ).values_list('annee_paie', flat=True).distinct().order_by('-annee_paie')

    # Données JSON pour Chart.js
    chart_labels = json.dumps(labels_mois)
    chart_brut = json.dumps([d['brut'] for d in evolution_data])
    chart_net = json.dumps([d['net'] for d in evolution_data])

    context = {
        'annee': annee,
        'annee_n1': annee - 1,
        'periode_type': periode_type,
        'mois': mois,
        'trimestre': trimestre,
        'service_id': service_id,
        'services': services,
        'tableau_services': tableau_services,
        'evolution_data': evolution_data,
        'top10': top10_data,
        'indicateurs': indicateurs,
        'indic_n1': indic_n1,
        'evol_brut': evol_brut,
        'annees_dispo': annees_dispo,
        'chart_labels': chart_labels,
        'chart_brut': chart_brut,
        'chart_net': chart_net,
        'mois_fr': MOIS_FR,
        'trimestres': [1, 2, 3, 4],
        'mois_liste': list(range(1, 13)),
    }
    return render(request, 'paie/rapport_masse_salariale.html', context)


@login_required
@entreprise_active_required
@reauth_required
def rapport_masse_salariale_excel(request):
    """Export Excel du rapport masse salariale."""
    if not OPENPYXL_AVAILABLE:
        return HttpResponse("openpyxl non disponible", status=500)

    entreprise = request.user.entreprise
    annee = int(request.GET.get('annee', date.today().year))
    periode_type = request.GET.get('periode_type', 'annee')
    mois = request.GET.get('mois', '')
    trimestre = request.GET.get('trimestre', '')
    service_id = request.GET.get('service', '')

    bulletins = _get_bulletins_filtrés(
        entreprise, annee, periode_type,
        mois=mois or None, trimestre=trimestre or None, service_id=service_id or None,
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Masse Salariale"

    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    center = Alignment(horizontal='center')
    right = Alignment(horizontal='right')
    thin = Side(style='thin')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def cell_style(cell, bold=False, fill_color=None, align=None):
        if bold:
            cell.font = Font(bold=True)
        if fill_color:
            cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
        if align:
            cell.alignment = align
        cell.border = border

    # Titre
    ws.merge_cells('A1:G1')
    ws['A1'] = f"RAPPORT MASSE SALARIALE - {annee}"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = center

    ws.merge_cells('A2:G2')
    ws['A2'] = f"Généré le {date.today().strftime('%d/%m/%Y')}"
    ws['A2'].alignment = center

    # Entête tableau par service
    row = 4
    headers = ['Service', 'Nb Employés', 'Masse Brute', 'CNSS Salarié', 'RTS', 'CNSS Patronal', 'Masse Nette']
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=col, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = center
        c.border = border

    # Données par service
    par_service = {}
    for b in bulletins.select_related('employe__service'):
        svc = b.employe.service
        svc_key = svc.pk if svc else 0
        svc_nom = str(svc) if svc else 'Sans service'
        if svc_key not in par_service:
            par_service[svc_key] = {'nom': svc_nom, 'nb': 0, 'brut': Decimal('0'),
                                     'cnss_sal': Decimal('0'), 'rts': Decimal('0'),
                                     'net': Decimal('0'), 'cnss_emp': Decimal('0')}
        r = par_service[svc_key]
        r['nb'] += 1
        r['brut'] += b.salaire_brut or Decimal('0')
        r['cnss_sal'] += b.cnss_employe or Decimal('0')
        r['rts'] += b.irg or Decimal('0')
        r['net'] += b.net_a_payer or Decimal('0')
        r['cnss_emp'] += b.cnss_employeur or Decimal('0')

    row = 5
    for sv in sorted(par_service.values(), key=lambda x: -x['brut']):
        ws.cell(row=row, column=1, value=sv['nom']).border = border
        ws.cell(row=row, column=2, value=sv['nb']).border = border
        ws.cell(row=row, column=3, value=float(sv['brut'])).border = border
        ws.cell(row=row, column=4, value=float(sv['cnss_sal'])).border = border
        ws.cell(row=row, column=5, value=float(sv['rts'])).border = border
        ws.cell(row=row, column=6, value=float(sv['cnss_emp'])).border = border
        ws.cell(row=row, column=7, value=float(sv['net'])).border = border
        row += 1

    # Format colonnes
    for col in range(3, 8):
        for r in range(5, row):
            ws.cell(row=r, column=col).number_format = '#,##0'

    for col_letter, width in zip('ABCDEFG', [30, 14, 18, 18, 18, 18, 18]):
        ws.column_dimensions[col_letter].width = width

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="masse_salariale_{annee}.xlsx"'
    return response


@login_required
@entreprise_active_required
@reauth_required
def rapport_masse_salariale_pdf(request):
    """Export PDF du rapport masse salariale."""
    if not REPORTLAB_AVAILABLE:
        return HttpResponse("ReportLab non disponible", status=500)

    entreprise = request.user.entreprise
    annee = int(request.GET.get('annee', date.today().year))
    periode_type = request.GET.get('periode_type', 'annee')
    mois = request.GET.get('mois', '')
    trimestre = request.GET.get('trimestre', '')
    service_id = request.GET.get('service', '')

    bulletins = _get_bulletins_filtrés(
        entreprise, annee, periode_type,
        mois=mois or None, trimestre=trimestre or None, service_id=service_id or None,
    )
    indicateurs = _calcul_rapport(bulletins)

    # Tableau par service
    par_service = {}
    for b in bulletins.select_related('employe__service'):
        svc = b.employe.service
        svc_key = svc.pk if svc else 0
        svc_nom = str(svc) if svc else 'Sans service'
        if svc_key not in par_service:
            par_service[svc_key] = {'nom': svc_nom, 'nb': 0, 'brut': Decimal('0'),
                                     'cnss_sal': Decimal('0'), 'rts': Decimal('0'),
                                     'net': Decimal('0')}
        r = par_service[svc_key]
        r['nb'] += 1
        r['brut'] += b.salaire_brut or Decimal('0')
        r['cnss_sal'] += b.cnss_employe or Decimal('0')
        r['rts'] += b.irg or Decimal('0')
        r['net'] += b.net_a_payer or Decimal('0')

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4),
                            leftMargin=1.5*cm, rightMargin=1.5*cm,
                            topMargin=2*cm, bottomMargin=2*cm)
    styles = getSampleStyleSheet()
    elements = []

    title_style = ParagraphStyle('Title', parent=styles['Title'],
                                  fontSize=16, spaceAfter=6, alignment=TA_CENTER)
    sub_style = ParagraphStyle('Sub', parent=styles['Normal'],
                                fontSize=10, spaceAfter=12, alignment=TA_CENTER)
    section_style = ParagraphStyle('Section', parent=styles['Normal'],
                                    fontSize=12, fontName='Helvetica-Bold',
                                    spaceAfter=6, spaceBefore=12)

    elements.append(Paragraph(f"RAPPORT MASSE SALARIALE - {annee}", title_style))
    elements.append(Paragraph(f"Généré le {date.today().strftime('%d/%m/%Y')}", sub_style))

    # Indicateurs globaux
    elements.append(Paragraph("Indicateurs globaux", section_style))
    indic_data = [
        ['Indicateur', 'Valeur'],
        ['Nombre de bulletins', str(indicateurs['nb'])],
        ['Masse salariale brute', f"{indicateurs['brut']:,.0f} GNF"],
        ['CNSS salarié total', f"{indicateurs['cnss_sal']:,.0f} GNF"],
        ['RTS total', f"{indicateurs['rts']:,.0f} GNF"],
        ['Masse salariale nette', f"{indicateurs['net']:,.0f} GNF"],
        ['Coût moyen / employé', f"{indicateurs['cout_moyen']:,.0f} GNF"],
        ['Ratio charges / brut', f"{indicateurs['ratio_charges']:.1f}%"],
    ]
    t_indic = Table(indic_data, colWidths=[8*cm, 7*cm])
    t_indic.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1F4E79')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#EBF3FF')]),
        ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
    ]))
    elements.append(t_indic)

    # Tableau par service
    elements.append(Paragraph("Répartition par service", section_style))
    svc_header = ['Service', 'Nb Emp.', 'Masse Brute', 'CNSS Sal.', 'RTS', 'Masse Nette']
    svc_rows = [svc_header]
    for sv in sorted(par_service.values(), key=lambda x: -x['brut']):
        svc_rows.append([
            sv['nom'],
            str(sv['nb']),
            f"{sv['brut']:,.0f}",
            f"{sv['cnss_sal']:,.0f}",
            f"{sv['rts']:,.0f}",
            f"{sv['net']:,.0f}",
        ])
    t_svc = Table(svc_rows, colWidths=[7*cm, 2.5*cm, 4*cm, 3.5*cm, 3.5*cm, 4*cm])
    t_svc.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1F4E79')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#EBF3FF')]),
        ('ALIGN', (1, 0), (-1, -1), 'RIGHT'),
        ('ALIGN', (0, 0), (0, -1), 'LEFT'),
    ]))
    elements.append(t_svc)

    doc.build(elements)
    buffer.seek(0)
    response = HttpResponse(buffer.read(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="masse_salariale_{annee}.pdf"'
    return response


# ============================================================================
# ÉTAT DE PAIE — Rapport mensuel détaillé (24 colonnes)
# ============================================================================

JOURS_FR = ['Lun', 'Mar', 'Mer', 'Jeu', 'Ven', 'Sam', 'Dim']

# ---- Classification intelligente des lignes bulletin ----
# Réutilise les patterns de paie/services.py pour une détection robuste

_PATTERNS_TRANSPORT = {'TRANSPORT', 'DEPLACEMENT'}
_PATTERNS_LOGEMENT = {'LOGEMENT', 'HEBERGEMENT'}
_PATTERNS_CHERTE = {'CHERTE', 'CHERTEVIE'}
_PATTERNS_DISCIPLINE = {'DISCIPLINE', 'ASSIDUITE'}
_PATTERNS_ANCIENNETE = {'ANCIENNETE', 'ANCIENNET'}
_PATTERNS_RENDEMENT = {'RENDEMENT', 'PERFORMANCE', 'OBJECTIF', 'PRODUCTIVITE'}
_PATTERNS_FONCTION = {'FONCTION', 'RESPONSABILITE', 'RESPONSABILIT'}
_PATTERNS_RISQUE = {'RISQUE', 'DANGER', 'PENIBILITE', 'INSALUBRITE', 'SALISSURE'}
_PATTERNS_REPAS = {'REPAS', 'PANIER', 'NOURRITURE', 'RESTAURATION'}

_MOTS_TRANSPORT = {
    'transport', 'déplacement', 'deplacement', 'indemnité de transport',
    'indemnite de transport', 'allocation transport', 'prime de transport',
    'frais de transport', 'carburant',
}
_MOTS_LOGEMENT = {
    'logement', 'hébergement', 'hebergement', 'allocation logement',
    'prime de logement', 'indemnité de logement', 'indemnite de logement',
}
_MOTS_CHERTE = {
    'cherté', 'cherte', 'vie chère', 'vie chere', 'cherté de vie',
    'cherte de vie', 'indemnité de vie chère',
}
_MOTS_DISCIPLINE = {
    'discipline', 'prime de discipline', 'assiduité', 'assiduite',
    'prime d\'assiduité', 'ponctualité', 'ponctualite',
}
_MOTS_ANCIENNETE = {'ancienneté', 'anciennete', 'prime d\'ancienneté'}
_MOTS_RENDEMENT = {
    'rendement', 'performance', 'objectif', 'productivité',
    'productivite', 'prime de rendement', 'prime de performance',
}
_MOTS_REPAS = {'repas', 'panier', 'nourriture', 'restauration', 'indemnité de repas'}

ETAT_PAIE_HEADERS = [
    'N°', 'Profession', 'Prénom', 'Nom', 'Matricule',
    'Salaire de Base', 'RTS', 'CNSS', 'VF',
    'Prime Discipline', 'Heures Supp', 'Montant HS',
    'Cher. Vie', 'Ind. Transport', 'Ind. Logement',
    'Brut', 'Salaire Net',
    'Jours Mois', 'Jours Travaillés', 'Dim. Travaillés',
    'Paie Dim.', 'Jours Repos', 'Jours Absents',
    'Fériés Payés', 'Net à Payer',
]

ETAT_PAIE_KEYS = [
    'num', 'profession', 'prenom', 'nom', 'matricule',
    'salaire_base', 'rts', 'cnss', 'vf',
    'prime_discipline', 'heures_sup', 'montant_hs',
    'cherte_vie', 'ind_transport', 'ind_logement',
    'brut', 'salaire_net',
    'jours_mois', 'jours_travailles', 'dim_travailles',
    'paie_dim', 'jours_repos', 'jours_absents',
    'feries_payes', 'net_a_payer',
]


def _classer_ligne_bulletin(ligne):
    """Classe une LigneBulletin par catégorie — détection multi-niveaux.

    Niveaux de détection (par priorité) :
    1. Code rubrique exact (PT, PL, PCV)
    2. Pattern dans le code rubrique (TRANSPORT, LOGEMENT, etc.)
    3. categorie_rubrique du modèle RubriquePaie
    4. Mots-clés dans le libellé (personnalisé ou rubrique)
    """
    rub = ligne.rubrique
    code = (rub.code_rubrique or '').upper().replace(' ', '').replace('-', '').replace('_', '')
    libelle = (ligne.libelle_personnalise or rub.libelle_rubrique or '').lower()
    categorie = getattr(rub, 'categorie_rubrique', '') or ''
    type_rub = getattr(rub, 'type_rubrique', '') or ''

    # --- Ignorer les retenues / cotisations (CNSS, RTS sont déjà dans colonnes dédiées) ---
    if type_rub in ('retenue', 'cotisation'):
        return 'retenue'

    # --- Niveau 1 : Codes exacts ---
    if code in ('PT',):
        return 'transport'
    if code in ('PL',):
        return 'logement'
    if code in ('PCV',):
        return 'cherte_vie'

    # --- Niveau 2 : Patterns dans le code ---
    for pat in _PATTERNS_TRANSPORT:
        if pat in code:
            return 'transport'
    for pat in _PATTERNS_LOGEMENT:
        if pat in code:
            return 'logement'
    for pat in _PATTERNS_CHERTE:
        if pat in code:
            return 'cherte_vie'
    for pat in _PATTERNS_DISCIPLINE:
        if pat in code:
            return 'discipline'
    for pat in _PATTERNS_ANCIENNETE:
        if pat in code:
            return 'anciennete'
    for pat in _PATTERNS_RENDEMENT:
        if pat in code:
            return 'rendement'
    for pat in _PATTERNS_FONCTION:
        if pat in code:
            return 'fonction'
    for pat in _PATTERNS_RISQUE:
        if pat in code:
            return 'risque'
    for pat in _PATTERNS_REPAS:
        if pat in code:
            return 'repas'

    # --- Niveau 3 : Mots-clés dans le libellé ---
    if any(m in libelle for m in _MOTS_TRANSPORT):
        return 'transport'
    if any(m in libelle for m in _MOTS_LOGEMENT):
        return 'logement'
    if any(m in libelle for m in _MOTS_CHERTE):
        return 'cherte_vie'
    if any(m in libelle for m in _MOTS_DISCIPLINE):
        return 'discipline'
    if any(m in libelle for m in _MOTS_ANCIENNETE):
        return 'anciennete'
    if any(m in libelle for m in _MOTS_RENDEMENT):
        return 'rendement'
    if any(m in libelle for m in _MOTS_REPAS):
        return 'repas'

    # --- Niveau 4 : categorie_rubrique comme fallback ---
    if categorie == 'salaire_base':
        return 'salaire_base'
    if categorie == 'indemnite':
        return 'indemnite'
    if categorie == 'prime':
        return 'prime'
    if categorie == 'avantage':
        return 'avantage'

    return 'autre_gain' if type_rub == 'gain' else 'autre'


def _get_salaire_base_intelligent(employe, bulletin_base, annee, mois):
    """
    Récupère le salaire de base du travailleur de façon intelligente et automatique.

    Stratégie de fallback:
    1. Salaire_base du bulletin courant (si non-nul)
    2. Dernier avenant de contrat (nouveau_salaire)
    3. Dernière promotion ou changement de salaire
    4. Dernier bulletin validé/payé de l'employé
    5. Moyenne des 3 derniers bulletins valides
    6. Zéro si rien trouvé
    """
    from employes.models import AvenantContrat, Promotion

    # 1. Utiliser le bulletin courant s'il existe et n'est pas zéro
    if bulletin_base and bulletin_base > 0:
        return bulletin_base

    eid = employe.pk

    # 2. Chercher le dernier avenant de contrat avec nouveau_salaire
    try:
        avenant = AvenantContrat.objects.filter(
            contrat__employe_id=eid,
            motif__in=['augmentation', 'promotion', 'autre'],
        ).order_by('-date_avenant').first()
        if avenant and avenant.nouveau_salaire and avenant.nouveau_salaire > 0:
            return avenant.nouveau_salaire
    except Exception:
        pass

    # 3. Chercher la dernière promotion avec nouveau_salaire
    try:
        promotion = Promotion.objects.filter(
            employe_id=eid,
        ).order_by('-date_decision').first()
        if promotion and promotion.nouveau_salaire and promotion.nouveau_salaire > 0:
            return promotion.nouveau_salaire
    except Exception:
        pass

    # 4. Chercher le dernier bulletin validé/payé de l'employé
    try:
        last_bulletin = BulletinPaie.objects.filter(
            employe_id=eid,
            statut_bulletin__in=['valide', 'paye'],
            salaire_base__gt=0,
        ).exclude(
            pk=bulletin_base.pk if hasattr(bulletin_base, 'pk') else None,
        ).order_by('-annee_paie', '-mois_paie').first()
        if last_bulletin and last_bulletin.salaire_base > 0:
            return last_bulletin.salaire_base
    except Exception:
        pass

    # 5. Moyenne des 3 derniers bulletins valides
    try:
        last_three = BulletinPaie.objects.filter(
            employe_id=eid,
            statut_bulletin__in=['valide', 'paye'],
            salaire_base__gt=0,
        ).order_by('-annee_paie', '-mois_paie')[:3].aggregate(
            avg_salaire=Avg('salaire_base')
        )
        if last_three and last_three.get('avg_salaire'):
            return Decimal(str(last_three['avg_salaire']))
    except Exception:
        pass

    # 6. Fallback zéro
    return Decimal('0')


def _fmt_gnf(val):
    """Formate un nombre avec séparateur de milliers."""
    if val is None:
        return '0'
    v = int(val)
    if v < 0:
        return '-' + f'{-v:,}'.replace(',', ' ')
    return f'{v:,}'.replace(',', ' ')


def _construire_donnees_etat_paie(bulletins_qs, annee, mois, entreprise):
    """Construit les données du tableau État de paie (24 colonnes).

    Remplissage intelligent multi-sources :
    - BulletinPaie : salaires, charges, HS stockées sur le bulletin
    - LigneBulletin → RubriquePaie : classification intelligente des composantes
    - Pointage : jours travaillés, dimanches, fériés (avec fallback si pas de pointages)
    - Absence : absences maladie, injustifiées, permissions
    - Conge : congés approuvés du mois
    - HeureSupplementaire : HS dimanche/nuit détaillées
    - JourFerie : fériés du mois
    - PeriodePaie : nombre_jours_travailles comme fallback
    - ArretTravail : arrêts maladie/accident
    """
    from temps_travail.models import Pointage, HeureSupplementaire, JourFerie
    from temps_travail.models import Absence, Conge

    mois_int = int(mois)
    annee_int = int(annee)
    nb_jours_mois = calendar.monthrange(annee_int, mois_int)[1]
    date_debut = date(annee_int, mois_int, 1)
    date_fin = date(annee_int, mois_int, nb_jours_mois)

    # ---- Calendrier du mois ----
    feries = set(
        JourFerie.objects.filter(
            Q(entreprise=entreprise) | Q(entreprise__isnull=True),
            date_jour_ferie__gte=date_debut,
            date_jour_ferie__lte=date_fin,
        ).values_list('date_jour_ferie', flat=True)
    )

    dimanches = set()
    jours_ouvrables = 0
    d = date_debut
    while d <= date_fin:
        if d.weekday() == 6:
            dimanches.add(d)
        else:
            if d not in feries:
                jours_ouvrables += 1
        d += timedelta(days=1)

    # ---- PeriodePaie : fallback pour jours travaillés ----
    from .models import PeriodePaie
    periode_jours = None
    try:
        pp = PeriodePaie.objects.get(entreprise=entreprise, annee=annee_int, mois=mois_int)
        periode_jours = pp.nombre_jours_travailles
    except PeriodePaie.DoesNotExist:
        pass

    # ---- Bulletins avec prefetch intelligent ----
    bulletins = list(bulletins_qs.filter(
        mois_paie=mois_int,
    ).select_related(
        'employe', 'employe__poste', 'employe__service',
    ).prefetch_related(
        Prefetch('lignes', queryset=LigneBulletin.objects.select_related('rubrique'))
    ).order_by('employe__nom', 'employe__prenoms'))

    if not bulletins:
        return [], {k: Decimal('0') for k in ETAT_PAIE_KEYS[5:]}, nb_jours_mois

    employe_ids = [b.employe_id for b in bulletins]

    # ---- BULK : Pointages ----
    pointages_map = defaultdict(lambda: defaultdict(int))
    for p in Pointage.objects.filter(
        employe_id__in=employe_ids,
        date_pointage__gte=date_debut,
        date_pointage__lte=date_fin,
    ):
        pointages_map[p.employe_id][p.statut_pointage] += 1

    # Pointages dimanche (présent un dimanche)
    dim_pointages = defaultdict(int)
    if dimanches:
        for p in Pointage.objects.filter(
            employe_id__in=employe_ids,
            date_pointage__in=dimanches,
            statut_pointage__in=['present', 'retard'],
        ):
            dim_pointages[p.employe_id] += 1

    # ---- BULK : Absences (hors pointage — module Absence dédié) ----
    absences_map = defaultdict(int)
    try:
        for ab in Absence.objects.filter(
            employe_id__in=employe_ids,
            date_absence__gte=date_debut,
            date_absence__lte=date_fin,
        ):
            absences_map[ab.employe_id] += 1
    except Exception:
        pass  # Module absence pas utilisé

    # ---- BULK : Congés approuvés chevauchant le mois ----
    conges_jours_map = defaultdict(int)
    try:
        for cg in Conge.objects.filter(
            employe_id__in=employe_ids,
            statut_demande='approuve',
            date_debut__lte=date_fin,
            date_fin__gte=date_debut,
        ):
            # Compter les jours dans le mois seulement
            d_start = max(cg.date_debut, date_debut)
            d_end = min(cg.date_fin, date_fin)
            conges_jours_map[cg.employe_id] += (d_end - d_start).days + 1
    except Exception:
        pass  # Module congé pas utilisé

    # ---- BULK : HeureSupplementaire détaillées ----
    hs_dim_map = defaultdict(Decimal)
    hs_nuit_map = defaultdict(Decimal)
    hs_detail_heures = defaultdict(Decimal)
    for hs in HeureSupplementaire.objects.filter(
        employe_id__in=employe_ids,
        date_hs__gte=date_debut,
        date_hs__lte=date_fin,
        statut__in=['valide', 'paye'],
    ):
        montant = hs.montant_hs or Decimal('0')
        heures = hs.nombre_heures or Decimal('0')
        hs_detail_heures[hs.employe_id] += heures
        if hs.type_hs in ('dimanche_75', 'dimanche_nuit_100'):
            hs_dim_map[hs.employe_id] += montant
        if hs.type_hs in ('nuit_50', 'dimanche_nuit_100'):
            hs_nuit_map[hs.employe_id] += montant

    # ---- BULK : Fériés travaillés ----
    feries_map = defaultdict(int)
    if feries:
        for p in Pointage.objects.filter(
            employe_id__in=employe_ids,
            date_pointage__in=feries,
            statut_pointage__in=['present', 'retard'],
        ):
            feries_map[p.employe_id] += 1

    # ---- BULK : Prêts en cours (échéances du mois) ----
    prets_map = defaultdict(Decimal)
    try:
        from paie.models_pret import EcheancePret
        for ech in EcheancePret.objects.filter(
            pret__employe_id__in=employe_ids,
            pret__statut='en_cours',
            date_echeance__gte=date_debut,
            date_echeance__lte=date_fin,
            statut__in=['en_attente', 'paye'],
        ).select_related('pret'):
            prets_map[ech.pret.employe_id] += ech.montant_echeance or Decimal('0')
    except Exception:
        pass

    # ---- BULK : Avances salaire en cours ----
    avances_map = defaultdict(Decimal)
    for av in AvanceSalaire.objects.filter(
        employe_id__in=employe_ids,
        statut='en_cours',
    ):
        avances_map[av.employe_id] += av.montant_mensuel or Decimal('0')

    # ---- BULK : Saisies-arrêts actives ----
    saisies_map = defaultdict(Decimal)
    try:
        from .models import SaisieArret
        for sa in SaisieArret.objects.filter(
            employe_id__in=employe_ids,
            statut='active',
            date_debut__lte=date_fin,
        ).filter(Q(date_fin__isnull=True) | Q(date_fin__gte=date_debut)):
            saisies_map[sa.employe_id] += sa.montant_mensuel or Decimal('0')
    except Exception:
        pass

    # ---- Construction des lignes ----
    rows = []
    totaux = {k: Decimal('0') for k in ETAT_PAIE_KEYS[5:]}

    for idx, b in enumerate(bulletins, 1):
        emp = b.employe
        eid = emp.pk

        # -- Classification intelligente des lignes bulletin --
        transport = Decimal('0')
        logement = Decimal('0')
        cherte_vie = Decimal('0')
        discipline = Decimal('0')
        for ligne in b.lignes.all():
            cat = _classer_ligne_bulletin(ligne)
            montant = ligne.montant or Decimal('0')
            if montant <= 0:
                continue
            if cat == 'transport':
                transport += montant
            elif cat == 'logement':
                logement += montant
            elif cat == 'cherte_vie':
                cherte_vie += montant
            elif cat == 'discipline':
                discipline += montant

        # -- Jours travaillés — fallback intelligent --
        pm = pointages_map.get(eid, {})
        has_pointages = bool(pm)
        jours_presents = pm.get('present', 0) + pm.get('retard', 0)
        jours_absents_pointage = pm.get('absent', 0) + pm.get('absence_justifiee', 0)

        # Enrichir avec absences et congés des modules dédiés
        jours_absents_extra = absences_map.get(eid, 0) + conges_jours_map.get(eid, 0)

        if has_pointages:
            jours_travailles = jours_presents
            jours_absents = jours_absents_pointage + jours_absents_extra
        else:
            # Fallback 1 : heures_normales du bulletin / 8h par jour
            if b.heures_normales and b.heures_normales > 0:
                jours_travailles = int(b.heures_normales / 8)
            # Fallback 2 : PeriodePaie.nombre_jours_travailles
            elif periode_jours:
                jours_travailles = periode_jours - jours_absents_extra
            # Fallback 3 : jours ouvrables du mois
            else:
                jours_travailles = jours_ouvrables - jours_absents_extra
            jours_absents = jours_absents_extra

        # -- HS : préférer le bulletin, enrichir avec HeureSupplementaire --
        hs_bulletin = (b.heures_supplementaires_30 or Decimal('0')) + \
                      (b.heures_supplementaires_60 or Decimal('0')) + \
                      (b.heures_nuit or Decimal('0')) + \
                      (b.heures_feries or Decimal('0')) + \
                      (b.heures_feries_nuit or Decimal('0'))
        hs_detail = hs_detail_heures.get(eid, Decimal('0'))
        # Prendre le max pour ne pas sous-estimer
        heures_sup = max(hs_bulletin, hs_detail)

        montant_hs = (b.prime_heures_sup or Decimal('0')) + \
                     (b.prime_nuit or Decimal('0')) + \
                     (b.prime_feries or Decimal('0')) + \
                     (b.prime_feries_nuit or Decimal('0'))

        # -- Fériés payés : montant depuis le bulletin --
        feries_payes_montant = (b.prime_feries or Decimal('0')) + (b.prime_feries_nuit or Decimal('0'))
        feries_travailles_count = feries_map.get(eid, 0)
        # Si pas de pointages mais fériés payés, estimer depuis le bulletin
        if not feries_travailles_count and feries_payes_montant > 0 and b.heures_feries:
            feries_travailles_count = int(b.heures_feries / 8) if b.heures_feries >= 8 else 1

        # -- Paie dimanche : depuis HS ou bulletin --
        paie_dim = hs_dim_map.get(eid, Decimal('0'))

        # -- Salaire de base : intelligent & automatique --
        salaire_base_intelligent = _get_salaire_base_intelligent(emp, b, annee_int, mois_int)

        row = {
            'num': idx,
            'profession': str(emp.poste) if emp.poste else '-',
            'prenom': emp.prenoms,
            'nom': emp.nom,
            'matricule': emp.matricule,
            'salaire_base': salaire_base_intelligent,
            'rts': b.irg or Decimal('0'),
            'cnss': b.cnss_employe or Decimal('0'),
            'vf': b.versement_forfaitaire or Decimal('0'),
            'prime_discipline': discipline,
            'heures_sup': heures_sup,
            'montant_hs': montant_hs,
            'cherte_vie': cherte_vie,
            'ind_transport': transport,
            'ind_logement': logement,
            'brut': b.salaire_brut or Decimal('0'),
            'salaire_net': (b.salaire_brut or Decimal('0')) - (b.total_retenues or Decimal('0')),
            'jours_mois': nb_jours_mois,
            'jours_travailles': jours_travailles,
            'dim_travailles': dim_pointages.get(eid, 0),
            'paie_dim': paie_dim,
            'jours_repos': len(dimanches),
            'jours_absents': jours_absents,
            'feries_payes': feries_travailles_count,
            'net_a_payer': b.net_a_payer or Decimal('0'),
        }
        rows.append(row)

        for k in totaux:
            totaux[k] += Decimal(str(row.get(k, 0)))

    return rows, totaux, nb_jours_mois


@login_required
@entreprise_active_required
@reauth_required
def rapport_etat_paie(request):
    """Rapport État de paie — tableau récapitulatif mensuel 24 colonnes."""
    entreprise = request.user.entreprise
    annee_courante = date.today().year

    annee = int(request.GET.get('annee', annee_courante))
    mois = request.GET.get('mois', str(date.today().month))
    service_id = request.GET.get('service', '')

    services = Service.objects.filter(entreprise=entreprise, actif=True).order_by('nom_service')
    annees_dispo = BulletinPaie.objects.filter(
        employe__entreprise=entreprise
    ).values_list('annee_paie', flat=True).distinct().order_by('-annee_paie')

    bulletins = _get_bulletins_filtrés(
        entreprise, annee, 'mois', mois=mois, service_id=service_id or None,
    )

    rows = []
    totaux = {}
    nb_jours_mois = 30
    if mois:
        rows, totaux, nb_jours_mois = _construire_donnees_etat_paie(bulletins, annee, mois, entreprise)

    context = {
        'annee': annee,
        'mois': mois,
        'mois_label': MOIS_FR[int(mois)] if mois else '',
        'service_id': service_id,
        'services': services,
        'annees_dispo': annees_dispo,
        'headers': ETAT_PAIE_HEADERS,
        'keys': ETAT_PAIE_KEYS,
        'rows': rows,
        'totaux': totaux,
        'nb_employes': len(rows),
        'mois_fr': MOIS_FR,
        'mois_liste': list(range(1, 13)),
    }
    return render(request, 'paie/rapport_etat_paie.html', context)


@login_required
@entreprise_active_required
@reauth_required
def rapport_etat_paie_excel(request):
    """Export Excel de l'État de paie."""
    if not OPENPYXL_AVAILABLE:
        return HttpResponse("openpyxl non disponible", status=500)

    entreprise = request.user.entreprise
    annee = int(request.GET.get('annee', date.today().year))
    mois = request.GET.get('mois', str(date.today().month))
    service_id = request.GET.get('service', '')

    bulletins = _get_bulletins_filtrés(
        entreprise, annee, 'mois', mois=mois, service_id=service_id or None,
    )
    rows, totaux, nb_jours_mois = _construire_donnees_etat_paie(bulletins, annee, mois, entreprise)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "État de Paie"

    header_font = Font(bold=True, color="FFFFFF", size=9)
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    center = Alignment(horizontal='center', wrap_text=True)
    right_align = Alignment(horizontal='right')
    thin = Side(style='thin')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    total_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")

    mois_label = MOIS_FR[int(mois)] if mois else ''
    nb_cols = len(ETAT_PAIE_HEADERS)

    # Titre
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=nb_cols)
    ws.cell(row=1, column=1, value=f"ÉTAT DE PAIE — {mois_label} {annee}").font = Font(bold=True, size=14)
    ws.cell(row=1, column=1).alignment = Alignment(horizontal='center')

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=nb_cols)
    ws.cell(row=2, column=1, value=f"Généré le {date.today().strftime('%d/%m/%Y')} — {len(rows)} employés").alignment = Alignment(horizontal='center')

    # En-têtes
    for col_idx, h in enumerate(ETAT_PAIE_HEADERS, 1):
        c = ws.cell(row=4, column=col_idx, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = center
        c.border = border

    # Données
    row_num = 5
    for row in rows:
        for col_idx, key in enumerate(ETAT_PAIE_KEYS, 1):
            val = row.get(key, '')
            if isinstance(val, Decimal):
                val = float(val)
            c = ws.cell(row=row_num, column=col_idx, value=val)
            c.border = border
            if col_idx >= 6:
                c.number_format = '#,##0'
        row_num += 1

    # Ligne totaux
    ws.cell(row=row_num, column=1, value='TOTAUX').font = Font(bold=True)
    for col_idx, key in enumerate(ETAT_PAIE_KEYS, 1):
        c = ws.cell(row=row_num, column=col_idx)
        if key in totaux:
            c.value = float(totaux[key])
            c.number_format = '#,##0'
        c.font = Font(bold=True)
        c.fill = total_fill
        c.border = border

    # Largeurs colonnes
    widths = [5, 18, 14, 14, 12] + [14] * 20
    for i, w in enumerate(widths[:nb_cols], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="etat_paie_{annee}_{mois}.xlsx"'
    return response


@login_required
@entreprise_active_required
@reauth_required
def rapport_etat_paie_pdf(request):
    """Export PDF de l'État de paie — landscape A4, font 6pt."""
    if not REPORTLAB_AVAILABLE:
        return HttpResponse("ReportLab non disponible", status=500)

    entreprise = request.user.entreprise
    annee = int(request.GET.get('annee', date.today().year))
    mois = request.GET.get('mois', str(date.today().month))
    service_id = request.GET.get('service', '')

    bulletins = _get_bulletins_filtrés(
        entreprise, annee, 'mois', mois=mois, service_id=service_id or None,
    )
    rows, totaux, nb_jours_mois = _construire_donnees_etat_paie(bulletins, annee, mois, entreprise)
    mois_label = MOIS_FR[int(mois)] if mois else ''

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4),
                            leftMargin=0.8*cm, rightMargin=0.8*cm,
                            topMargin=1.5*cm, bottomMargin=1.5*cm)
    styles = getSampleStyleSheet()
    elements = []

    title_style = ParagraphStyle('EPTitle', parent=styles['Title'],
                                  fontSize=14, spaceAfter=4, alignment=TA_CENTER)
    sub_style = ParagraphStyle('EPSub', parent=styles['Normal'],
                                fontSize=8, spaceAfter=8, alignment=TA_CENTER)

    elements.append(Paragraph(f"ÉTAT DE PAIE — {mois_label} {annee}", title_style))
    elements.append(Paragraph(
        f"Généré le {date.today().strftime('%d/%m/%Y')} — {len(rows)} employés", sub_style))

    # Préparer données tableau
    cell_style_tiny = ParagraphStyle('Tiny', parent=styles['Normal'], fontSize=5, leading=6)
    header_ps = ParagraphStyle('HdrTiny', parent=styles['Normal'],
                                fontSize=5, leading=6, fontName='Helvetica-Bold',
                                textColor=colors.white)

    def _p(text, style=cell_style_tiny):
        return Paragraph(str(text), style)

    table_data = [[_p(h, header_ps) for h in ETAT_PAIE_HEADERS]]

    for row in rows:
        table_data.append([
            _p(row['num']),
            _p(row['profession']),
            _p(row['prenom']),
            _p(row['nom']),
            _p(row['matricule']),
            _p(_fmt_gnf(row['salaire_base'])),
            _p(_fmt_gnf(row['rts'])),
            _p(_fmt_gnf(row['cnss'])),
            _p(_fmt_gnf(row['vf'])),
            _p(_fmt_gnf(row['prime_discipline'])),
            _p(row['heures_sup']),
            _p(_fmt_gnf(row['montant_hs'])),
            _p(_fmt_gnf(row['cherte_vie'])),
            _p(_fmt_gnf(row['ind_transport'])),
            _p(_fmt_gnf(row['ind_logement'])),
            _p(_fmt_gnf(row['brut'])),
            _p(_fmt_gnf(row['salaire_net'])),
            _p(row['jours_mois']),
            _p(row['jours_travailles']),
            _p(row['dim_travailles']),
            _p(_fmt_gnf(row['paie_dim'])),
            _p(row['jours_repos']),
            _p(row['jours_absents']),
            _p(row['feries_payes']),
            _p(_fmt_gnf(row['net_a_payer'])),
        ])

    # Ligne totaux
    total_row = [_p('', cell_style_tiny)] * 5
    for key in ETAT_PAIE_KEYS[5:]:
        total_row.append(_p(_fmt_gnf(totaux.get(key, 0)), ParagraphStyle(
            'TotTiny', parent=cell_style_tiny, fontName='Helvetica-Bold')))
    table_data.append(total_row)

    col_widths = [0.8*cm, 2.5*cm, 2*cm, 2*cm, 1.8*cm] + [1.1*cm] * 20
    t = Table(table_data, colWidths=col_widths, repeatRows=1)
    t.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1F4E79')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('GRID', (0, 0), (-1, -1), 0.3, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, colors.HexColor('#F5F8FF')]),
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#D9E2F3')),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('ALIGN', (5, 0), (-1, -1), 'RIGHT'),
        ('ALIGN', (0, 0), (4, -1), 'LEFT'),
    ]))
    elements.append(t)

    doc.build(elements)
    buffer.seek(0)
    response = HttpResponse(buffer.read(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="etat_paie_{annee}_{mois}.pdf"'
    return response


# ============================================================================
# FEUILLE DE PRÉSENCE — Grille journalière mensuelle
# ============================================================================

_CODES_POINTAGE = {
    'present': 'P',
    'present_am': 'AM',
    'present_pm': 'PM',
    'present_am_pm': 'AM+PM',
    'absent': 'A',
    'retard': 'R',
    'absence_justifiee': 'AJ',
    'malade': 'M',
    'p': 'P',
    'a': 'A',
}


def _construire_donnees_feuille_presence(entreprise, annee, mois, service_id=None):
    """Construit les données de la feuille de présence mensuelle."""
    from temps_travail.models import Pointage, HeureSupplementaire, JourFerie
    from employes.models import Employe as EmployeModel

    mois_int = int(mois)
    annee_int = int(annee)
    nb_jours = calendar.monthrange(annee_int, mois_int)[1]
    date_debut = date(annee_int, mois_int, 1)
    date_fin = date(annee_int, mois_int, nb_jours)

    # Employés actifs
    emp_qs = EmployeModel.objects.filter(
        entreprise=entreprise,
        statut_employe='actif',
    ).select_related('poste', 'service').order_by('nom', 'prenoms')
    if service_id:
        emp_qs = emp_qs.filter(service_id=service_id)

    employes = list(emp_qs)
    emp_ids = [e.pk for e in employes]

    # Jours fériés
    feries = set(
        JourFerie.objects.filter(
            Q(entreprise=entreprise) | Q(entreprise__isnull=True),
            date_jour_ferie__gte=date_debut,
            date_jour_ferie__lte=date_fin,
        ).values_list('date_jour_ferie', flat=True)
    )

    # Dimanches
    dimanches = set()
    d = date_debut
    while d <= date_fin:
        if d.weekday() == 6:
            dimanches.add(d)
        d += timedelta(days=1)

    # Pointages bulk
    pointages_bulk = defaultdict(dict)
    for p in Pointage.objects.filter(
        employe_id__in=emp_ids,
        date_pointage__gte=date_debut,
        date_pointage__lte=date_fin,
    ):
        pointages_bulk[p.employe_id][p.date_pointage.day] = p.statut_pointage

    # Bulletins du mois
    bulletins_map = {}
    for b in BulletinPaie.objects.filter(
        employe_id__in=emp_ids,
        annee_paie=annee_int,
        mois_paie=mois_int,
        statut_bulletin__in=['calcule', 'valide', 'paye'],
    ):
        bulletins_map[b.employe_id] = b

    # HS dimanches
    hs_dim_map = defaultdict(Decimal)
    for hs in HeureSupplementaire.objects.filter(
        employe_id__in=emp_ids,
        date_hs__gte=date_debut,
        date_hs__lte=date_fin,
        type_hs__in=['dimanche_75', 'dimanche_nuit_100'],
        statut__in=['valide', 'paye'],
    ):
        hs_dim_map[hs.employe_id] += hs.montant_hs or Decimal('0')

    # Avances en cours
    avances_map = defaultdict(Decimal)
    for av in AvanceSalaire.objects.filter(
        employe_id__in=emp_ids,
        statut='en_cours',
    ):
        avances_map[av.employe_id] += av.montant_mensuel or Decimal('0')

    # Construire les lignes
    rows = []
    jours_info = []
    for j in range(1, nb_jours + 1):
        d = date(annee_int, mois_int, j)
        jour_nom = JOURS_FR[d.weekday()]
        est_dim = d in dimanches
        est_ferie = d in feries
        jours_info.append({'jour': j, 'nom': jour_nom, 'dimanche': est_dim, 'ferie': est_ferie})

    for emp in employes:
        eid = emp.pk
        grille = []
        # Separate counters for each statut type
        count_p = 0      # Présent
        count_am = 0     # Présent AM
        count_pm = 0     # Présent PM
        count_am_pm = 0  # Présent AM et PM
        count_a = 0      # Absent
        count_r = 0      # Retard
        count_aj = 0     # Absence justifiée
        count_m = 0      # Malade
        nb_dim_travailles = 0

        for j in range(1, nb_jours + 1):
            statut = pointages_bulk.get(eid, {}).get(j, '')
            code = _CODES_POINTAGE.get(statut, '')
            grille.append(code)

            # Count by specific statut
            if statut == 'present':
                count_p += 1
                d_j = date(annee_int, mois_int, j)
                if d_j in dimanches:
                    nb_dim_travailles += 1
            elif statut == 'present_am':
                count_am += 1
                d_j = date(annee_int, mois_int, j)
                if d_j in dimanches:
                    nb_dim_travailles += 1
            elif statut == 'present_pm':
                count_pm += 1
                d_j = date(annee_int, mois_int, j)
                if d_j in dimanches:
                    nb_dim_travailles += 1
            elif statut == 'present_am_pm':
                count_am_pm += 1
                d_j = date(annee_int, mois_int, j)
                if d_j in dimanches:
                    nb_dim_travailles += 1
            elif statut == 'absent':
                count_a += 1
            elif statut == 'retard':
                count_r += 1
                d_j = date(annee_int, mois_int, j)
                if d_j in dimanches:
                    nb_dim_travailles += 1
            elif statut == 'absence_justifiee':
                count_aj += 1
            elif statut == 'malade':
                count_m += 1
            elif statut in ('p', 'a'):
                # Handle abbreviations: 'p' counts as present, 'a' counts as absent
                if statut == 'p':
                    count_p += 1
                    d_j = date(annee_int, mois_int, j)
                    if d_j in dimanches:
                        nb_dim_travailles += 1
                else:  # statut == 'a'
                    count_a += 1

        bull = bulletins_map.get(eid)
        feries_travailles = 0
        if feries:
            for f_date in feries:
                statut_f = pointages_bulk.get(eid, {}).get(f_date.day, '')
                if statut_f in ('present', 'present_am', 'present_pm', 'present_am_pm', 'retard', 'p'):
                    feries_travailles += 1

        # Calculate totals for display
        nb_present = count_p + count_am + count_pm + count_am_pm
        nb_absent = count_a + count_aj

        # Salaire de base : intelligent & automatique
        salaire_base_bull = bull.salaire_base if bull else None
        salaire_base_intelligent = _get_salaire_base_intelligent(emp, salaire_base_bull, annee_int, mois_int)

        row = {
            'employe': emp,
            'nom': emp.nom,
            'prenom': emp.prenoms,
            'matricule': emp.matricule,
            'profession': str(emp.poste) if emp.poste else '-',
            'grille': grille,
            'count_p': count_p,
            'count_am': count_am,
            'count_pm': count_pm,
            'count_am_pm': count_am_pm,
            'count_a': count_a,
            'count_r': count_r,
            'count_aj': count_aj,
            'count_m': count_m,
            'nb_present': nb_present,
            'nb_absent': nb_absent,
            'nb_dim_travailles': nb_dim_travailles,
            'feries_travailles': feries_travailles,
            'salaire_base': salaire_base_intelligent,
            'brut': bull.salaire_brut if bull else Decimal('0'),
            'net': bull.net_a_payer if bull else Decimal('0'),
            'paie_dim': hs_dim_map.get(eid, Decimal('0')),
            'feries_payes': bull.prime_feries if bull else Decimal('0'),
            'avance': avances_map.get(eid, Decimal('0')),
        }
        rows.append(row)

    return rows, jours_info, nb_jours


@login_required
@entreprise_active_required
@reauth_required
def rapport_feuille_presence(request):
    """Rapport Feuille de présence — grille journalière mensuelle."""
    entreprise = request.user.entreprise
    annee_courante = date.today().year

    annee = int(request.GET.get('annee', annee_courante))
    mois = request.GET.get('mois', str(date.today().month))
    service_id = request.GET.get('service', '')

    services = Service.objects.filter(entreprise=entreprise, actif=True).order_by('nom_service')
    annees_dispo = BulletinPaie.objects.filter(
        employe__entreprise=entreprise
    ).values_list('annee_paie', flat=True).distinct().order_by('-annee_paie')

    rows = []
    jours_info = []
    nb_jours = 30
    if mois:
        rows, jours_info, nb_jours = _construire_donnees_feuille_presence(
            entreprise, annee, mois, service_id or None)

    context = {
        'annee': annee,
        'mois': mois,
        'mois_label': MOIS_FR[int(mois)] if mois else '',
        'service_id': service_id,
        'services': services,
        'annees_dispo': annees_dispo,
        'rows': rows,
        'jours_info': jours_info,
        'nb_jours': nb_jours,
        'nb_employes': len(rows),
        'mois_fr': MOIS_FR,
        'mois_liste': list(range(1, 13)),
    }
    return render(request, 'paie/rapport_feuille_presence.html', context)


@login_required
@entreprise_active_required
@reauth_required
def rapport_feuille_presence_excel(request):
    """Export Excel de la feuille de présence."""
    if not OPENPYXL_AVAILABLE:
        return HttpResponse("openpyxl non disponible", status=500)

    entreprise = request.user.entreprise
    annee = int(request.GET.get('annee', date.today().year))
    mois = request.GET.get('mois', str(date.today().month))
    service_id = request.GET.get('service', '')

    rows, jours_info, nb_jours = _construire_donnees_feuille_presence(
        entreprise, annee, mois, service_id or None)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Feuille de Présence"
    mois_label = MOIS_FR[int(mois)] if mois else ''

    header_font = Font(bold=True, color="FFFFFF", size=9)
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    dim_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    ferie_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    present_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    absent_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    center = Alignment(horizontal='center')
    thin = Side(style='thin')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    nb_cols = 4 + nb_jours + 13  # nom,prenom,matricule,profession + jours + résumé (P, AM, PM, AM+PM, A, R, AJ, M, Dim.T, Fér.T, Sal.Base, Net, Avance)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=nb_cols)
    ws.cell(row=1, column=1, value=f"FEUILLE DE PRÉSENCE — {mois_label} {annee}").font = Font(bold=True, size=14)
    ws.cell(row=1, column=1).alignment = Alignment(horizontal='center')

    # En-têtes
    hdr_row = 3
    fixed_headers = ['Nom', 'Prénom', 'Matricule', 'Profession']
    for col, h in enumerate(fixed_headers, 1):
        c = ws.cell(row=hdr_row, column=col, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = center
        c.border = border

    for idx, ji in enumerate(jours_info):
        col = 5 + idx
        c = ws.cell(row=hdr_row, column=col, value=f"{ji['jour']}\n{ji['nom']}")
        c.font = header_font
        c.alignment = Alignment(horizontal='center', wrap_text=True)
        c.border = border
        if ji['dimanche']:
            c.fill = dim_fill
        elif ji['ferie']:
            c.fill = ferie_fill
        else:
            c.fill = header_fill

    resume_headers = ['P', 'AM', 'PM', 'AM+PM', 'A', 'R', 'AJ', 'M', 'Dim. Trav.', 'Fériés Trav.', 'Salaire Base', 'Net', 'Avance']
    for idx, h in enumerate(resume_headers):
        col = 5 + nb_jours + idx
        c = ws.cell(row=hdr_row, column=col, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = center
        c.border = border

    # Données
    data_row = 4
    for row in rows:
        ws.cell(row=data_row, column=1, value=row['nom']).border = border
        ws.cell(row=data_row, column=2, value=row['prenom']).border = border
        ws.cell(row=data_row, column=3, value=row['matricule']).border = border
        ws.cell(row=data_row, column=4, value=row['profession']).border = border

        for idx, code in enumerate(row['grille']):
            col = 5 + idx
            c = ws.cell(row=data_row, column=col, value=code)
            c.alignment = center
            c.border = border
            if code in ('P', 'AM', 'PM', 'AM+PM'):
                c.fill = present_fill
            elif code in ('A', 'AJ'):
                c.fill = absent_fill

        base_col = 5 + nb_jours
        ws.cell(row=data_row, column=base_col, value=row['count_p']).border = border
        ws.cell(row=data_row, column=base_col + 1, value=row['count_am']).border = border
        ws.cell(row=data_row, column=base_col + 2, value=row['count_pm']).border = border
        ws.cell(row=data_row, column=base_col + 3, value=row['count_am_pm']).border = border
        ws.cell(row=data_row, column=base_col + 4, value=row['count_a']).border = border
        ws.cell(row=data_row, column=base_col + 5, value=row['count_r']).border = border
        ws.cell(row=data_row, column=base_col + 6, value=row['count_aj']).border = border
        ws.cell(row=data_row, column=base_col + 7, value=row['count_m']).border = border
        ws.cell(row=data_row, column=base_col + 8, value=row['nb_dim_travailles']).border = border
        ws.cell(row=data_row, column=base_col + 9, value=row['feries_travailles']).border = border
        ws.cell(row=data_row, column=base_col + 10, value=float(row['salaire_base'])).border = border
        ws.cell(row=data_row, column=base_col + 10).number_format = '#,##0'
        ws.cell(row=data_row, column=base_col + 11, value=float(row['net'])).border = border
        ws.cell(row=data_row, column=base_col + 11).number_format = '#,##0'
        ws.cell(row=data_row, column=base_col + 12, value=float(row['avance'])).border = border
        ws.cell(row=data_row, column=base_col + 12).number_format = '#,##0'
        data_row += 1

    # Largeurs
    for col in range(1, 5):
        ws.column_dimensions[get_column_letter(col)].width = 14
    for col in range(5, 5 + nb_jours):
        ws.column_dimensions[get_column_letter(col)].width = 5
    for col in range(5 + nb_jours, nb_cols + 1):
        ws.column_dimensions[get_column_letter(col)].width = 14

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="feuille_presence_{annee}_{mois}.xlsx"'
    return response


@login_required
@entreprise_active_required
@reauth_required
def rapport_feuille_presence_pdf(request):
    """Export PDF de la feuille de présence — landscape large."""
    if not REPORTLAB_AVAILABLE:
        return HttpResponse("ReportLab non disponible", status=500)

    entreprise = request.user.entreprise
    annee = int(request.GET.get('annee', date.today().year))
    mois = request.GET.get('mois', str(date.today().month))
    service_id = request.GET.get('service', '')

    rows, jours_info, nb_jours = _construire_donnees_feuille_presence(
        entreprise, annee, mois, service_id or None)
    mois_label = MOIS_FR[int(mois)] if mois else ''

    # A3 landscape (42 x 29.7 cm)
    page_w = 42 * cm
    page_h = 29.7 * cm

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=(page_w, page_h),
                            leftMargin=0.5*cm, rightMargin=0.5*cm,
                            topMargin=1*cm, bottomMargin=1*cm)
    styles = getSampleStyleSheet()
    elements = []

    title_style = ParagraphStyle('FPTitle', parent=styles['Title'],
                                  fontSize=12, spaceAfter=4, alignment=TA_CENTER)
    sub_style = ParagraphStyle('FPSub', parent=styles['Normal'],
                                fontSize=7, spaceAfter=6, alignment=TA_CENTER)

    elements.append(Paragraph(f"FEUILLE DE PRÉSENCE — {mois_label} {annee}", title_style))
    elements.append(Paragraph(
        f"Généré le {date.today().strftime('%d/%m/%Y')} — {len(rows)} employés", sub_style))

    cell_s = ParagraphStyle('CellS', parent=styles['Normal'], fontSize=5, leading=6)
    hdr_s = ParagraphStyle('HdrS', parent=styles['Normal'],
                            fontSize=5, leading=6, fontName='Helvetica-Bold',
                            textColor=colors.white)

    def _p(text, style=cell_s):
        return Paragraph(str(text), style)

    # En-tête
    header_row = [_p('Nom', hdr_s), _p('Prénom', hdr_s), _p('Mat.', hdr_s)]
    for ji in jours_info:
        header_row.append(_p(f"{ji['jour']}", hdr_s))
    header_row += [_p('P', hdr_s), _p('AM', hdr_s), _p('PM', hdr_s), _p('AM+PM', hdr_s),
                   _p('A', hdr_s), _p('R', hdr_s), _p('AJ', hdr_s), _p('M', hdr_s),
                   _p('Dim.T', hdr_s), _p('Fér.T', hdr_s), _p('Base', hdr_s), _p('Net', hdr_s), _p('Av.', hdr_s)]

    table_data = [header_row]
    for row in rows:
        r = [_p(row['nom']), _p(row['prenom']), _p(row['matricule'])]
        for code in row['grille']:
            r.append(_p(code))
        r += [
            _p(row['count_p']),
            _p(row['count_am']),
            _p(row['count_pm']),
            _p(row['count_am_pm']),
            _p(row['count_a']),
            _p(row['count_r']),
            _p(row['count_aj']),
            _p(row['count_m']),
            _p(row['nb_dim_travailles']),
            _p(row['feries_travailles']),
            _p(_fmt_gnf(row['salaire_base'])),
            _p(_fmt_gnf(row['net'])),
            _p(_fmt_gnf(row['avance'])),
        ]
        table_data.append(r)

    # Largeurs colonnes
    fixed_w = [2.2*cm, 2*cm, 1.5*cm]
    jour_w = [0.7*cm] * nb_jours
    resume_w = [0.8*cm] * 8 + [0.9*cm, 0.9*cm, 1.5*cm, 1.5*cm, 1.2*cm]
    col_widths = fixed_w + jour_w + resume_w

    t = Table(table_data, colWidths=col_widths, repeatRows=1)

    style_cmds = [
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1F4E79')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('GRID', (0, 0), (-1, -1), 0.2, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F8FAFF')]),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('ALIGN', (3, 0), (3 + nb_jours - 1, -1), 'CENTER'),
    ]

    # Colorer les dimanches/fériés dans l'en-tête
    for idx, ji in enumerate(jours_info):
        col = 3 + idx
        if ji['dimanche']:
            style_cmds.append(('BACKGROUND', (col, 0), (col, 0), colors.HexColor('#5B9BD5')))
        elif ji['ferie']:
            style_cmds.append(('BACKGROUND', (col, 0), (col, 0), colors.HexColor('#FFD966')))

    t.setStyle(TableStyle(style_cmds))
    elements.append(t)

    doc.build(elements)
    buffer.seek(0)
    response = HttpResponse(buffer.read(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="feuille_presence_{annee}_{mois}.pdf"'
    return response
