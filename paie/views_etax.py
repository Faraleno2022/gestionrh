"""
Vues eTax Guinée: préparation, exports et suivi des déclarations fiscales.
"""
import io
from decimal import Decimal
from datetime import date, datetime

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db.models import Q, Sum
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from django.utils import timezone
from django.views.decorators.http import require_POST

from core.decorators import entreprise_active_required, reauth_required
from .models import BulletinPaie, DeclarationEtax, PeriodePaie

try:
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from reportlab.lib.enums import TA_CENTER
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False


MOIS_LABELS = {
    1: 'Janvier', 2: 'Février', 3: 'Mars', 4: 'Avril',
    5: 'Mai', 6: 'Juin', 7: 'Juillet', 8: 'Août',
    9: 'Septembre', 10: 'Octobre', 11: 'Novembre', 12: 'Décembre',
}


def _to_decimal(value):
    return value if value is not None else Decimal('0')


def _parse_date(value):
    if not value:
        return None
    try:
        return datetime.strptime(value, '%Y-%m-%d').date()
    except ValueError:
        return None


def _bulletins_etax(entreprise, annee, mois):
    periodes = PeriodePaie.objects.filter(entreprise=entreprise, annee=annee, mois=mois)
    filtre_periode = Q(periode__in=periodes) | Q(annee_paie=annee, mois_paie=mois)
    return BulletinPaie.objects.filter(
        filtre_periode,
        statut_bulletin__in=['calcule', 'valide', 'paye'],
        employe__entreprise=entreprise,
    ).select_related('employe', 'periode').order_by('employe__matricule', 'employe__nom')


def get_etax_data(entreprise, annee, mois):
    bulletins = _bulletins_etax(entreprise, annee, mois)
    totaux = bulletins.aggregate(
        total_brut=Sum('salaire_brut'),
        total_net=Sum('net_a_payer'),
        total_base_rts=Sum('base_rts'),
        total_rts=Sum('irg'),
        total_vf=Sum('versement_forfaitaire'),
        total_ta=Sum('taxe_apprentissage'),
        total_onfpp=Sum('contribution_onfpp'),
        total_cnss_employe=Sum('cnss_employe'),
        total_cnss_employeur=Sum('cnss_employeur'),
    )

    data = {key: _to_decimal(value) for key, value in totaux.items()}
    data['effectif'] = bulletins.values('employe').distinct().count()
    data['total_fiscal'] = (
        data['total_rts'] + data['total_vf'] + data['total_ta'] + data['total_onfpp']
    )
    data['total_cnss'] = data['total_cnss_employe'] + data['total_cnss_employeur']
    data['total_general'] = data['total_fiscal'] + data['total_cnss']
    data['mois_label'] = MOIS_LABELS.get(mois, str(mois))
    data['detail_employes'] = []

    for bulletin in bulletins:
        emp = bulletin.employe
        data['detail_employes'].append({
            'matricule': emp.matricule or '',
            'num_cnss': emp.num_cnss_individuel or '',
            'nom_complet': f"{emp.nom} {emp.prenoms}".strip(),
            'brut': bulletin.salaire_brut,
            'base_rts': getattr(bulletin, 'base_rts', Decimal('0')) or Decimal('0'),
            'rts': bulletin.irg,
            'cnss_employe': bulletin.cnss_employe,
            'cnss_employeur': bulletin.cnss_employeur,
            'vf': getattr(bulletin, 'versement_forfaitaire', Decimal('0')) or Decimal('0'),
            'ta': getattr(bulletin, 'taxe_apprentissage', Decimal('0')) or Decimal('0'),
            'onfpp': getattr(bulletin, 'contribution_onfpp', Decimal('0')) or Decimal('0'),
            'net': bulletin.net_a_payer,
        })

    return data


def _reference_etax_interne(entreprise, annee, mois):
    return f"ETAX-{entreprise.id}-{annee}{mois:02d}"


def creer_ou_actualiser_declaration_etax(entreprise, annee, mois, utilisateur=None):
    data = get_etax_data(entreprise, annee, mois)
    periode = PeriodePaie.objects.filter(entreprise=entreprise, annee=annee, mois=mois).first()
    declaration, _ = DeclarationEtax.objects.update_or_create(
        entreprise=entreprise,
        annee=annee,
        mois=mois,
        defaults={
            'periode': periode,
            'reference': _reference_etax_interne(entreprise, annee, mois),
            'effectif_declare': data['effectif'],
            'total_brut': data['total_brut'],
            'total_net': data['total_net'],
            'total_base_rts': data['total_base_rts'],
            'total_rts': data['total_rts'],
            'total_vf': data['total_vf'],
            'total_ta': data['total_ta'],
            'total_onfpp': data['total_onfpp'],
            'total_fiscal': data['total_fiscal'],
            'total_cnss_employe': data['total_cnss_employe'],
            'total_cnss_employeur': data['total_cnss_employeur'],
            'total_cnss': data['total_cnss'],
            'total_general': data['total_general'],
            'statut': 'generee',
            'date_generation': timezone.now(),
            'genere_par': utilisateur,
        }
    )
    return declaration


@login_required
@entreprise_active_required
@reauth_required
def declaration_etax(request):
    annee = int(request.GET.get('annee', timezone.now().year))
    mois = int(request.GET.get('mois', timezone.now().month))
    entreprise = request.user.entreprise

    data = get_etax_data(entreprise, annee, mois)
    declaration = DeclarationEtax.objects.filter(
        entreprise=entreprise, annee=annee, mois=mois
    ).first()
    historique = DeclarationEtax.objects.filter(entreprise=entreprise).order_by('-annee', '-mois')[:12]
    annees = PeriodePaie.objects.filter(
        entreprise=entreprise
    ).values_list('annee', flat=True).distinct().order_by('-annee')

    return render(request, 'paie/declarations_etax.html', {
        'annee': annee,
        'mois': mois,
        'mois_labels': MOIS_LABELS,
        'annees': annees,
        'data': data,
        'declaration': declaration,
        'historique': historique,
        'statuts': DeclarationEtax.STATUTS,
    })


@login_required
@entreprise_active_required
@reauth_required
@require_POST
def generer_declaration_etax(request):
    annee = int(request.POST.get('annee', timezone.now().year))
    mois = int(request.POST.get('mois', timezone.now().month))
    declaration = creer_ou_actualiser_declaration_etax(
        request.user.entreprise, annee, mois, request.user
    )
    messages.success(request, f"Déclaration eTax {declaration.reference} générée.")
    return redirect(f"{reverse('paie:declaration_etax')}?annee={annee}&mois={mois}")


@login_required
@entreprise_active_required
@reauth_required
@require_POST
def mettre_a_jour_declaration_etax(request, pk):
    declaration = get_object_or_404(DeclarationEtax, pk=pk, entreprise=request.user.entreprise)
    declaration.statut = request.POST.get('statut') or declaration.statut
    declaration.reference_etax = request.POST.get('reference_etax') or None
    declaration.date_declaration = _parse_date(request.POST.get('date_declaration'))
    declaration.date_paiement = _parse_date(request.POST.get('date_paiement'))
    declaration.observations = request.POST.get('observations') or None
    if request.FILES.get('fichier_recu'):
        declaration.fichier_recu = request.FILES['fichier_recu']
    declaration.save()
    messages.success(request, "Suivi eTax mis à jour.")
    return redirect(f"{reverse('paie:declaration_etax')}?annee={declaration.annee}&mois={declaration.mois}")


def _filename(entreprise, annee, mois, ext):
    nom = ''.join(c for c in entreprise.nom_entreprise if c.isalnum() or c in (' ', '_', '-')).strip()
    nom = nom.replace(' ', '_') or 'entreprise'
    return f"eTax_{nom}_{annee}_{mois:02d}.{ext}"


@login_required
@entreprise_active_required
def declaration_etax_excel(request):
    if not OPENPYXL_AVAILABLE:
        return HttpResponse("Module openpyxl non disponible", status=500)

    annee = int(request.GET.get('annee', timezone.now().year))
    mois = int(request.GET.get('mois', timezone.now().month))
    entreprise = request.user.entreprise
    data = get_etax_data(entreprise, annee, mois)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Declaration eTax"
    bold = Font(bold=True)
    title = Font(bold=True, size=14)
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    ws.merge_cells('A1:J1')
    ws['A1'] = "DÉCLARATION eTax - DONNÉES PRÉPARÉES PAR GUINÉERH"
    ws['A1'].font = title
    ws['A1'].alignment = Alignment(horizontal='center')
    ws.merge_cells('A2:J2')
    ws['A2'] = f"Période: {data['mois_label']} {annee}"
    ws['A2'].alignment = Alignment(horizontal='center')

    rows = [
        ("Entreprise", entreprise.nom_entreprise),
        ("NIF", getattr(entreprise, 'nif', '') or ''),
        ("N° CNSS", getattr(entreprise, 'num_cnss', '') or ''),
        ("Effectif déclaré", data['effectif']),
    ]
    row = 4
    for label, value in rows:
        ws.cell(row=row, column=1, value=label).font = bold
        ws.cell(row=row, column=2, value=value)
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="RÉCAPITULATIF").font = title
    row += 1
    recap = [
        ("Masse salariale brute", data['total_brut']),
        ("Base imposable RTS", data['total_base_rts']),
        ("RTS", data['total_rts']),
        ("VF", data['total_vf']),
        ("TA", data['total_ta']),
        ("ONFPP", data['total_onfpp']),
        ("Total fiscal eTax", data['total_fiscal']),
        ("CNSS employé", data['total_cnss_employe']),
        ("CNSS employeur", data['total_cnss_employeur']),
        ("Total CNSS", data['total_cnss']),
        ("Total général", data['total_general']),
    ]
    for label, amount in recap:
        ws.cell(row=row, column=1, value=label).border = border
        cell = ws.cell(row=row, column=2, value=float(amount))
        cell.border = border
        cell.number_format = '#,##0'
        row += 1

    row += 2
    headers = ["Matricule", "N° CNSS", "Nom et prénoms", "Brut", "Base RTS", "RTS", "CNSS employé", "CNSS employeur", "VF", "TA", "ONFPP", "Net"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal='center')
    row += 1

    for emp in data['detail_employes']:
        values = [
            emp['matricule'], emp['num_cnss'], emp['nom_complet'], emp['brut'], emp['base_rts'],
            emp['rts'], emp['cnss_employe'], emp['cnss_employeur'], emp['vf'], emp['ta'], emp['onfpp'], emp['net']
        ]
        for col, value in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=float(value) if isinstance(value, Decimal) else value)
            cell.border = border
            if col >= 4:
                cell.number_format = '#,##0'
        row += 1

    for col in range(1, 13):
        ws.column_dimensions[get_column_letter(col)].width = 16 if col != 3 else 28

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    response = HttpResponse(output.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{_filename(entreprise, annee, mois, "xlsx")}"'
    return response


@login_required
@entreprise_active_required
def declaration_etax_pdf(request):
    if not REPORTLAB_AVAILABLE:
        return HttpResponse("Module reportlab non disponible", status=500)

    annee = int(request.GET.get('annee', timezone.now().year))
    mois = int(request.GET.get('mois', timezone.now().month))
    entreprise = request.user.entreprise
    data = get_etax_data(entreprise, annee, mois)

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), topMargin=0.8*cm, bottomMargin=0.8*cm)
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('EtaxTitle', parent=styles['Heading1'], alignment=TA_CENTER, fontSize=15)
    elements = [
        Paragraph("DÉCLARATION eTax - RÉCAPITULATIF MENSUEL", title_style),
        Paragraph(f"{entreprise.nom_entreprise} - {data['mois_label']} {annee}", styles['Normal']),
        Spacer(1, 10),
    ]

    recap = [
        ["Indicateur", "Montant"],
        ["Masse salariale brute", f"{data['total_brut']:,.0f} GNF"],
        ["Base imposable RTS", f"{data['total_base_rts']:,.0f} GNF"],
        ["RTS", f"{data['total_rts']:,.0f} GNF"],
        ["VF", f"{data['total_vf']:,.0f} GNF"],
        ["TA", f"{data['total_ta']:,.0f} GNF"],
        ["ONFPP", f"{data['total_onfpp']:,.0f} GNF"],
        ["Total fiscal eTax", f"{data['total_fiscal']:,.0f} GNF"],
        ["Total CNSS", f"{data['total_cnss']:,.0f} GNF"],
        ["Total général", f"{data['total_general']:,.0f} GNF"],
    ]
    recap_table = Table(recap, colWidths=[7*cm, 5*cm])
    recap_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1F4E79')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 0.4, colors.grey),
        ('ALIGN', (1, 1), (1, -1), 'RIGHT'),
        ('FONTNAME', (0, -2), (-1, -1), 'Helvetica-Bold'),
    ]))
    elements.append(recap_table)
    elements.append(Spacer(1, 12))

    table_data = [["Matricule", "Nom", "Brut", "Base RTS", "RTS", "CNSS emp.", "CNSS pat.", "VF", "TA", "ONFPP", "Net"]]
    for emp in data['detail_employes']:
        table_data.append([
            emp['matricule'],
            emp['nom_complet'][:24],
            f"{emp['brut']:,.0f}",
            f"{emp['base_rts']:,.0f}",
            f"{emp['rts']:,.0f}",
            f"{emp['cnss_employe']:,.0f}",
            f"{emp['cnss_employeur']:,.0f}",
            f"{emp['vf']:,.0f}",
            f"{emp['ta']:,.0f}",
            f"{emp['onfpp']:,.0f}",
            f"{emp['net']:,.0f}",
        ])
    detail_table = Table(table_data, repeatRows=1)
    detail_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#E9ECEF')),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 0.25, colors.grey),
        ('FONTSIZE', (0, 0), (-1, -1), 7),
        ('ALIGN', (2, 1), (-1, -1), 'RIGHT'),
    ]))
    elements.append(detail_table)
    elements.append(Spacer(1, 8))
    elements.append(Paragraph(
        "Document préparé pour saisie ou import manuel sur le portail eTax Guinée. "
        "Aucune transmission automatique à la DGI n'est effectuée.",
        styles['Italic']
    ))

    doc.build(elements)
    buffer.seek(0)
    response = HttpResponse(buffer.read(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{_filename(entreprise, annee, mois, "pdf")}"'
    return response
