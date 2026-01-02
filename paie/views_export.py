"""
Vues pour l'export des déclarations sociales et fiscales.
- Export CNSS (Bordereau de cotisations)
- Export DMU (Déclaration Mensuelle Unique)
"""
from django.shortcuts import render, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.db.models import Sum
from django.utils import timezone
from decimal import Decimal
from datetime import date
import io

from .models import PeriodePaie, BulletinPaie, Constante
from employes.models import Employe
from core.decorators import entreprise_active_required

# Imports pour génération Excel
try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

# Imports pour génération PDF
try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm, mm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False


def get_declarations_data(entreprise, annee, mois=None):
    """Récupère les données pour les déclarations sociales"""
    from django.db.models import Q
    
    # Filtrer les périodes de l'entreprise
    periodes = PeriodePaie.objects.filter(entreprise=entreprise, annee=annee)
    if mois:
        periodes = periodes.filter(mois=mois)
    
    # Filtrer les bulletins (inclure calcule, valide et paye)
    # Accepter les bulletins liés à une période OU directement par annee_paie/mois_paie
    bulletins_filter = Q(periode__in=periodes) | Q(annee_paie=annee)
    if mois:
        bulletins_filter = Q(periode__in=periodes) | Q(annee_paie=annee, mois_paie=mois)
    
    bulletins = BulletinPaie.objects.filter(
        bulletins_filter,
        statut_bulletin__in=['calcule', 'valide', 'paye'],
    ).filter(
        Q(employe__entreprise=entreprise) | Q(employe__entreprise__isnull=True)
    ).select_related('employe', 'periode')
    
    # Récupérer les constantes
    def get_constante(code, default):
        const = Constante.objects.filter(code=code, actif=True).first()
        return const.valeur if const else Decimal(str(default))
    
    plancher_cnss = get_constante('PLANCHER_CNSS', 550000)
    plafond_cnss = get_constante('PLAFOND_CNSS', 2500000)
    taux_cnss_employe = get_constante('TAUX_CNSS_EMPLOYE', 5)
    taux_cnss_employeur = get_constante('TAUX_CNSS_EMPLOYEUR', 18)
    taux_vf = get_constante('TAUX_VF', 6)
    taux_ta = get_constante('TAUX_TA', 2.0)
    
    # Calculer les totaux
    totaux = bulletins.aggregate(
        masse_salariale=Sum('salaire_brut'),
        total_cnss_employe=Sum('cnss_employe'),
        total_cnss_employeur=Sum('cnss_employeur'),
        total_rts=Sum('irg'),
    )
    
    masse_salariale = totaux['masse_salariale'] or Decimal('0')
    total_cnss_employe = totaux['total_cnss_employe'] or Decimal('0')
    total_cnss_employeur = totaux['total_cnss_employeur'] or Decimal('0')
    total_rts = totaux['total_rts'] or Decimal('0')
    
    # Calculer VF et TA
    total_vf = (masse_salariale * taux_vf / Decimal('100')).quantize(Decimal('1'))
    total_ta = (masse_salariale * taux_ta / Decimal('100')).quantize(Decimal('1'))
    
    # Détail par employé
    detail_employes = []
    for bulletin in bulletins:
        emp = bulletin.employe
        detail_employes.append({
            'matricule': emp.matricule,
            'num_cnss': emp.num_cnss_individuel or '',
            'nom': emp.nom,
            'prenoms': emp.prenoms,
            'nom_complet': f"{emp.nom} {emp.prenoms}",
            'salaire_brut': bulletin.salaire_brut,
            'cnss_employe': bulletin.cnss_employe,
            'cnss_employeur': bulletin.cnss_employeur,
            'total_cnss': bulletin.cnss_employe + bulletin.cnss_employeur,
            'rts': bulletin.irg,
            'net_a_payer': bulletin.net_a_payer,
            'periode': str(bulletin.periode),
        })
    
    return {
        'entreprise': entreprise,
        'annee': annee,
        'mois': mois,
        'nb_salaries': bulletins.values('employe').distinct().count(),
        'masse_salariale': masse_salariale,
        'plancher_cnss': plancher_cnss,
        'plafond_cnss': plafond_cnss,
        'taux_cnss_employe': taux_cnss_employe,
        'taux_cnss_employeur': taux_cnss_employeur,
        'total_cnss_employe': total_cnss_employe,
        'total_cnss_employeur': total_cnss_employeur,
        'total_cnss': total_cnss_employe + total_cnss_employeur,
        'taux_vf': taux_vf,
        'taux_ta': taux_ta,
        'total_rts': total_rts,
        'total_vf': total_vf,
        'total_ta': total_ta,
        'detail_employes': detail_employes,
        'date_generation': timezone.now(),
    }


# ============================================================================
# EXPORT CNSS
# ============================================================================

@login_required
@entreprise_active_required
def export_cnss_excel(request):
    """Exporte le bordereau CNSS au format Excel"""
    if not OPENPYXL_AVAILABLE:
        return HttpResponse("Module openpyxl non disponible", status=500)
    
    annee = request.GET.get('annee', date.today().year)
    mois = request.GET.get('mois')
    
    data = get_declarations_data(request.user.entreprise, int(annee), int(mois) if mois else None)
    
    # Créer le workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Bordereau CNSS"
    
    # Styles
    header_font = Font(bold=True, size=14)
    title_font = Font(bold=True, size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(bold=True, color="FFFFFF")
    
    # En-tête du document
    ws.merge_cells('A1:H1')
    ws['A1'] = "BORDEREAU DE COTISATIONS CNSS"
    ws['A1'].font = header_font
    ws['A1'].alignment = Alignment(horizontal='center')
    
    ws.merge_cells('A2:H2')
    ws['A2'] = f"Période: {data['mois']:02d}/{data['annee']}" if data['mois'] else f"Année: {data['annee']}"
    ws['A2'].alignment = Alignment(horizontal='center')
    
    # Informations entreprise
    row = 4
    ws[f'A{row}'] = "Entreprise:"
    ws[f'B{row}'] = data['entreprise'].nom_entreprise
    ws[f'B{row}'].font = title_font
    
    row += 1
    ws[f'A{row}'] = "N° CNSS Employeur:"
    ws[f'B{row}'] = getattr(data['entreprise'], 'num_cnss', 'N/A')
    
    row += 1
    ws[f'A{row}'] = "Date de génération:"
    ws[f'B{row}'] = data['date_generation'].strftime('%d/%m/%Y %H:%M')
    
    # Récapitulatif
    row += 2
    ws[f'A{row}'] = "RÉCAPITULATIF DES COTISATIONS"
    ws[f'A{row}'].font = title_font
    
    row += 1
    recap_data = [
        ["Nombre de salariés", data['nb_salaries']],
        ["Masse salariale brute", f"{data['masse_salariale']:,.0f} GNF"],
        ["Plancher CNSS", f"{data['plancher_cnss']:,.0f} GNF"],
        ["Plafond CNSS", f"{data['plafond_cnss']:,.0f} GNF"],
        ["", ""],
        ["Part salariale (5%)", f"{data['total_cnss_employe']:,.0f} GNF"],
        ["Part patronale (18%)", f"{data['total_cnss_employeur']:,.0f} GNF"],
        ["TOTAL À VERSER (23%)", f"{data['total_cnss']:,.0f} GNF"],
    ]
    
    for item in recap_data:
        ws[f'A{row}'] = item[0]
        ws[f'C{row}'] = item[1]
        if "TOTAL" in str(item[0]):
            ws[f'A{row}'].font = title_font
            ws[f'C{row}'].font = title_font
        row += 1
    
    # Liste nominative
    row += 2
    ws[f'A{row}'] = "LISTE NOMINATIVE DES ASSURÉS"
    ws[f'A{row}'].font = title_font
    
    row += 1
    headers = ["N°", "Matricule", "N° CNSS", "Nom", "Prénoms", "Salaire Brut", "CNSS Employé", "CNSS Employeur", "Total CNSS"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center')
    
    row += 1
    for idx, emp in enumerate(data['detail_employes'], 1):
        ws.cell(row=row, column=1, value=idx).border = border
        ws.cell(row=row, column=2, value=emp['matricule']).border = border
        ws.cell(row=row, column=3, value=emp['num_cnss']).border = border
        ws.cell(row=row, column=4, value=emp['nom']).border = border
        ws.cell(row=row, column=5, value=emp['prenoms']).border = border
        ws.cell(row=row, column=6, value=float(emp['salaire_brut'])).border = border
        ws.cell(row=row, column=7, value=float(emp['cnss_employe'])).border = border
        ws.cell(row=row, column=8, value=float(emp['cnss_employeur'])).border = border
        ws.cell(row=row, column=9, value=float(emp['total_cnss'])).border = border
        row += 1
    
    # Totaux
    ws.cell(row=row, column=5, value="TOTAUX").font = title_font
    ws.cell(row=row, column=6, value=float(data['masse_salariale'])).font = title_font
    ws.cell(row=row, column=7, value=float(data['total_cnss_employe'])).font = title_font
    ws.cell(row=row, column=8, value=float(data['total_cnss_employeur'])).font = title_font
    ws.cell(row=row, column=9, value=float(data['total_cnss'])).font = title_font
    
    # Ajuster les largeurs de colonnes
    for col in range(1, 10):
        ws.column_dimensions[get_column_letter(col)].width = 15
    
    # Générer le fichier
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"CNSS_{data['entreprise'].nom_entreprise}_{data['annee']}"
    if data['mois']:
        filename += f"_{data['mois']:02d}"
    filename += ".xlsx"
    
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


@login_required
@entreprise_active_required
def export_cnss_pdf(request):
    """Exporte le bordereau CNSS au format PDF"""
    if not REPORTLAB_AVAILABLE:
        return HttpResponse("Module reportlab non disponible", status=500)
    
    annee = request.GET.get('annee', date.today().year)
    mois = request.GET.get('mois')
    
    data = get_declarations_data(request.user.entreprise, int(annee), int(mois) if mois else None)
    
    # Créer le PDF
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=1*cm, bottomMargin=1*cm)
    elements = []
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('Title', parent=styles['Heading1'], alignment=TA_CENTER, fontSize=16)
    subtitle_style = ParagraphStyle('Subtitle', parent=styles['Normal'], alignment=TA_CENTER, fontSize=12)
    section_style = ParagraphStyle('Section', parent=styles['Heading2'], fontSize=12, spaceAfter=10)
    
    # Titre
    elements.append(Paragraph("BORDEREAU DE COTISATIONS CNSS", title_style))
    periode_text = f"Période: {data['mois']:02d}/{data['annee']}" if data['mois'] else f"Année: {data['annee']}"
    elements.append(Paragraph(periode_text, subtitle_style))
    elements.append(Spacer(1, 20))
    
    # Informations entreprise
    elements.append(Paragraph("INFORMATIONS EMPLOYEUR", section_style))
    info_data = [
        ["Entreprise:", data['entreprise'].nom_entreprise],
        ["N° CNSS Employeur:", getattr(data['entreprise'], 'num_cnss', 'N/A')],
        ["Date de génération:", data['date_generation'].strftime('%d/%m/%Y %H:%M')],
    ]
    info_table = Table(info_data, colWidths=[5*cm, 10*cm])
    info_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
    ]))
    elements.append(info_table)
    elements.append(Spacer(1, 20))
    
    # Récapitulatif
    elements.append(Paragraph("RÉCAPITULATIF DES COTISATIONS", section_style))
    recap_data = [
        ["Nombre de salariés", str(data['nb_salaries'])],
        ["Masse salariale brute", f"{data['masse_salariale']:,.0f} GNF"],
        ["Plancher CNSS", f"{data['plancher_cnss']:,.0f} GNF"],
        ["Plafond CNSS", f"{data['plafond_cnss']:,.0f} GNF"],
        ["", ""],
        ["Part salariale (5%)", f"{data['total_cnss_employe']:,.0f} GNF"],
        ["Part patronale (18%)", f"{data['total_cnss_employeur']:,.0f} GNF"],
        ["TOTAL À VERSER (23%)", f"{data['total_cnss']:,.0f} GNF"],
    ]
    recap_table = Table(recap_data, colWidths=[7*cm, 5*cm])
    recap_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('BACKGROUND', (0, -1), (-1, -1), colors.lightgrey),
        ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
    ]))
    elements.append(recap_table)
    elements.append(Spacer(1, 20))
    
    # Liste nominative
    elements.append(Paragraph("LISTE NOMINATIVE DES ASSURÉS", section_style))
    
    # En-têtes du tableau
    table_data = [["N°", "Matricule", "Nom Complet", "Salaire Brut", "CNSS Emp.", "CNSS Pat.", "Total"]]
    
    for idx, emp in enumerate(data['detail_employes'], 1):
        table_data.append([
            str(idx),
            emp['matricule'],
            emp['nom_complet'][:25],
            f"{emp['salaire_brut']:,.0f}",
            f"{emp['cnss_employe']:,.0f}",
            f"{emp['cnss_employeur']:,.0f}",
            f"{emp['total_cnss']:,.0f}",
        ])
    
    # Totaux
    table_data.append([
        "", "", "TOTAUX",
        f"{data['masse_salariale']:,.0f}",
        f"{data['total_cnss_employe']:,.0f}",
        f"{data['total_cnss_employeur']:,.0f}",
        f"{data['total_cnss']:,.0f}",
    ])
    
    emp_table = Table(table_data, colWidths=[1*cm, 2*cm, 5*cm, 2.5*cm, 2.5*cm, 2.5*cm, 2.5*cm])
    emp_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('BACKGROUND', (0, -1), (-1, -1), colors.lightgrey),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('ALIGN', (2, 1), (2, -1), 'LEFT'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
    ]))
    elements.append(emp_table)
    
    # Générer le PDF
    doc.build(elements)
    buffer.seek(0)
    
    filename = f"CNSS_{data['entreprise'].nom_entreprise}_{data['annee']}"
    if data['mois']:
        filename += f"_{data['mois']:02d}"
    filename += ".pdf"
    
    response = HttpResponse(buffer.read(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


# ============================================================================
# EXPORT DMU (Déclaration Mensuelle Unique)
# ============================================================================

@login_required
@entreprise_active_required
def export_dmu_excel(request):
    """Exporte la DMU au format Excel"""
    if not OPENPYXL_AVAILABLE:
        return HttpResponse("Module openpyxl non disponible", status=500)
    
    annee = request.GET.get('annee', date.today().year)
    mois = request.GET.get('mois')
    
    if not mois:
        return HttpResponse("Le mois est requis pour la DMU", status=400)
    
    data = get_declarations_data(request.user.entreprise, int(annee), int(mois))
    
    # Créer le workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "DMU"
    
    # Styles
    header_font = Font(bold=True, size=14)
    title_font = Font(bold=True, size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    header_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
    header_font_white = Font(bold=True, color="FFFFFF")
    
    # En-tête du document
    ws.merge_cells('A1:J1')
    ws['A1'] = "DÉCLARATION MENSUELLE UNIQUE (DMU)"
    ws['A1'].font = header_font
    ws['A1'].alignment = Alignment(horizontal='center')
    
    ws.merge_cells('A2:J2')
    ws['A2'] = f"Période: {data['mois']:02d}/{data['annee']}"
    ws['A2'].alignment = Alignment(horizontal='center')
    
    # Informations entreprise
    row = 4
    ws[f'A{row}'] = "IDENTIFICATION EMPLOYEUR"
    ws[f'A{row}'].font = title_font
    
    row += 1
    ws[f'A{row}'] = "Raison sociale:"
    ws[f'C{row}'] = data['entreprise'].nom_entreprise
    
    row += 1
    ws[f'A{row}'] = "NIF:"
    ws[f'C{row}'] = getattr(data['entreprise'], 'nif', 'N/A')
    
    row += 1
    ws[f'A{row}'] = "RCCM:"
    ws[f'C{row}'] = getattr(data['entreprise'], 'rccm', 'N/A')
    
    row += 1
    ws[f'A{row}'] = "N° CNSS:"
    ws[f'C{row}'] = getattr(data['entreprise'], 'num_cnss', 'N/A')
    
    row += 1
    ws[f'A{row}'] = "Adresse:"
    ws[f'C{row}'] = getattr(data['entreprise'], 'adresse', 'N/A')
    
    # Liste des salariés
    row += 2
    ws[f'A{row}'] = "LISTE DES SALARIÉS"
    ws[f'A{row}'].font = title_font
    
    row += 1
    headers = ["N°", "Matricule", "N° CNSS", "Nom", "Prénoms", "Salaire Brut", "CNSS (5%)", "RTS", "Net à Payer"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center')
    
    row += 1
    for idx, emp in enumerate(data['detail_employes'], 1):
        ws.cell(row=row, column=1, value=idx).border = border
        ws.cell(row=row, column=2, value=emp['matricule']).border = border
        ws.cell(row=row, column=3, value=emp['num_cnss']).border = border
        ws.cell(row=row, column=4, value=emp['nom']).border = border
        ws.cell(row=row, column=5, value=emp['prenoms']).border = border
        ws.cell(row=row, column=6, value=float(emp['salaire_brut'])).border = border
        ws.cell(row=row, column=7, value=float(emp['cnss_employe'])).border = border
        ws.cell(row=row, column=8, value=float(emp['rts'])).border = border
        ws.cell(row=row, column=9, value=float(emp['net_a_payer'])).border = border
        row += 1
    
    # Totaux salariés
    ws.cell(row=row, column=5, value="TOTAUX").font = title_font
    ws.cell(row=row, column=6, value=float(data['masse_salariale'])).font = title_font
    ws.cell(row=row, column=7, value=float(data['total_cnss_employe'])).font = title_font
    ws.cell(row=row, column=8, value=float(data['total_rts'])).font = title_font
    
    # Récapitulatif des impôts et taxes
    row += 3
    ws[f'A{row}'] = "RÉCAPITULATIF DES IMPÔTS ET TAXES"
    ws[f'A{row}'].font = title_font
    
    row += 1
    recap_headers = ["Désignation", "Assiette", "Taux", "Montant"]
    for col, header in enumerate(recap_headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = border
    
    row += 1
    recap_data = [
        ["RTS (Retenue sur Traitements et Salaires)", float(data['masse_salariale']), "Barème", float(data['total_rts'])],
        ["VF (Versement Forfaitaire)", float(data['masse_salariale']), f"{data['taux_vf']}%", float(data['total_vf'])],
        ["TA (Taxe d'Apprentissage)", float(data['masse_salariale']), f"{data['taux_ta']}%", float(data['total_ta'])],
    ]
    
    for item in recap_data:
        for col, val in enumerate(item, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.border = border
        row += 1
    
    # Total à verser DNI
    total_dni = data['total_rts'] + data['total_vf'] + data['total_ta']
    ws.cell(row=row, column=1, value="TOTAL À VERSER À LA DNI").font = title_font
    ws.cell(row=row, column=4, value=float(total_dni)).font = title_font
    for col in range(1, 5):
        ws.cell(row=row, column=col).border = border
    
    # CNSS
    row += 2
    ws[f'A{row}'] = "COTISATIONS CNSS"
    ws[f'A{row}'].font = title_font
    
    row += 1
    cnss_data = [
        ["Part salariale (5%)", float(data['total_cnss_employe'])],
        ["Part patronale (18%)", float(data['total_cnss_employeur'])],
        ["TOTAL À VERSER À LA CNSS", float(data['total_cnss'])],
    ]
    
    for item in cnss_data:
        ws.cell(row=row, column=1, value=item[0]).border = border
        ws.cell(row=row, column=2, value=item[1]).border = border
        if "TOTAL" in item[0]:
            ws.cell(row=row, column=1).font = title_font
            ws.cell(row=row, column=2).font = title_font
        row += 1
    
    # Ajuster les largeurs
    col_widths = [5, 12, 15, 15, 15, 15, 12, 12, 15]
    for col, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width
    
    # Générer le fichier
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"DMU_{data['entreprise'].nom_entreprise}_{data['annee']}_{data['mois']:02d}.xlsx"
    
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


@login_required
@entreprise_active_required
def export_dmu_pdf(request):
    """Exporte la DMU au format PDF"""
    if not REPORTLAB_AVAILABLE:
        return HttpResponse("Module reportlab non disponible", status=500)
    
    annee = request.GET.get('annee', date.today().year)
    mois = request.GET.get('mois')
    
    if not mois:
        return HttpResponse("Le mois est requis pour la DMU", status=400)
    
    data = get_declarations_data(request.user.entreprise, int(annee), int(mois))
    
    # Créer le PDF
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=1*cm, bottomMargin=1*cm)
    elements = []
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('Title', parent=styles['Heading1'], alignment=TA_CENTER, fontSize=16)
    subtitle_style = ParagraphStyle('Subtitle', parent=styles['Normal'], alignment=TA_CENTER, fontSize=12)
    section_style = ParagraphStyle('Section', parent=styles['Heading2'], fontSize=12, spaceAfter=10)
    
    # Titre
    elements.append(Paragraph("DÉCLARATION MENSUELLE UNIQUE (DMU)", title_style))
    elements.append(Paragraph(f"Période: {data['mois']:02d}/{data['annee']}", subtitle_style))
    elements.append(Spacer(1, 15))
    
    # Informations employeur
    elements.append(Paragraph("IDENTIFICATION EMPLOYEUR", section_style))
    info_data = [
        ["Raison sociale:", data['entreprise'].nom_entreprise],
        ["NIF:", getattr(data['entreprise'], 'nif', 'N/A')],
        ["RCCM:", getattr(data['entreprise'], 'rccm', 'N/A')],
        ["N° CNSS:", getattr(data['entreprise'], 'num_cnss', 'N/A')],
    ]
    info_table = Table(info_data, colWidths=[4*cm, 10*cm])
    info_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
    ]))
    elements.append(info_table)
    elements.append(Spacer(1, 15))
    
    # Liste des salariés
    elements.append(Paragraph("LISTE DES SALARIÉS", section_style))
    
    table_data = [["N°", "Matricule", "Nom Complet", "Brut", "CNSS", "RTS", "Net"]]
    for idx, emp in enumerate(data['detail_employes'], 1):
        table_data.append([
            str(idx),
            emp['matricule'],
            emp['nom_complet'][:20],
            f"{emp['salaire_brut']:,.0f}",
            f"{emp['cnss_employe']:,.0f}",
            f"{emp['rts']:,.0f}",
            f"{emp['net_a_payer']:,.0f}",
        ])
    
    table_data.append([
        "", "", "TOTAUX",
        f"{data['masse_salariale']:,.0f}",
        f"{data['total_cnss_employe']:,.0f}",
        f"{data['total_rts']:,.0f}",
        "",
    ])
    
    emp_table = Table(table_data, colWidths=[0.8*cm, 2*cm, 4*cm, 2.5*cm, 2*cm, 2*cm, 2.5*cm])
    emp_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2E7D32')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('BACKGROUND', (0, -1), (-1, -1), colors.lightgrey),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('ALIGN', (2, 1), (2, -1), 'LEFT'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
    ]))
    elements.append(emp_table)
    elements.append(Spacer(1, 15))
    
    # Récapitulatif impôts
    elements.append(Paragraph("RÉCAPITULATIF DES IMPÔTS ET TAXES", section_style))
    
    total_dni = data['total_rts'] + data['total_vf'] + data['total_ta']
    
    recap_data = [
        ["Désignation", "Taux", "Montant"],
        ["RTS (Retenue sur Traitements et Salaires)", "Barème", f"{data['total_rts']:,.0f} GNF"],
        ["VF (Versement Forfaitaire)", f"{data['taux_vf']}%", f"{data['total_vf']:,.0f} GNF"],
        ["TA (Taxe d'Apprentissage)", f"{data['taux_ta']}%", f"{data['total_ta']:,.0f} GNF"],
        ["TOTAL À VERSER À LA DNI", "", f"{total_dni:,.0f} GNF"],
    ]
    
    recap_table = Table(recap_data, colWidths=[8*cm, 3*cm, 4*cm])
    recap_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1565C0')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#BBDEFB')),
        ('ALIGN', (1, 0), (-1, -1), 'RIGHT'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
    ]))
    elements.append(recap_table)
    elements.append(Spacer(1, 15))
    
    # CNSS
    elements.append(Paragraph("COTISATIONS CNSS", section_style))
    
    cnss_data = [
        ["Désignation", "Montant"],
        ["Part salariale (5%)", f"{data['total_cnss_employe']:,.0f} GNF"],
        ["Part patronale (18%)", f"{data['total_cnss_employeur']:,.0f} GNF"],
        ["TOTAL À VERSER À LA CNSS", f"{data['total_cnss']:,.0f} GNF"],
    ]
    
    cnss_table = Table(cnss_data, colWidths=[8*cm, 5*cm])
    cnss_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#E65100')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#FFE0B2')),
        ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
    ]))
    elements.append(cnss_table)
    
    # Générer le PDF
    doc.build(elements)
    buffer.seek(0)
    
    filename = f"DMU_{data['entreprise'].nom_entreprise}_{data['annee']}_{data['mois']:02d}.pdf"
    
    response = HttpResponse(buffer.read(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response
