"""
Générateur Excel de vérification automatique de paie - Guinée
Conforme CGI 2022 + Code du Travail

Usage: python docs/Verification_Paie_Guinee.py
Génère: docs/Verification_Paie_Guinee.xlsx
"""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, NamedStyle
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import FormulaRule


def creer_excel_verification():
    wb = Workbook()
    
    # Styles
    titre_style = Font(bold=True, size=14, color="FFFFFF")
    header_style = Font(bold=True, size=11)
    money_format = '#,##0" GNF"'
    percent_format = '0.00%'
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    ok_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    error_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    param_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    
    # ========== FEUILLE 1: PARAMÈTRES ==========
    ws1 = wb.active
    ws1.title = "Paramètres"
    
    # Titre
    ws1.merge_cells('A1:D1')
    ws1['A1'] = "PARAMÈTRES LÉGAUX - GUINÉE (CGI 2022)"
    ws1['A1'].font = titre_style
    ws1['A1'].fill = header_fill
    ws1['A1'].alignment = Alignment(horizontal='center')
    
    # CNSS
    ws1['A3'] = "COTISATIONS CNSS"
    ws1['A3'].font = header_style
    
    params_cnss = [
        ('A4', 'Plancher CNSS:', 'B4', 550000),
        ('A5', 'Plafond CNSS:', 'B5', 2500000),
        ('A6', 'Taux salarié:', 'B6', 0.05),
        ('A7', 'Taux employeur:', 'B7', 0.18),
    ]
    
    for label_cell, label, value_cell, value in params_cnss:
        ws1[label_cell] = label
        ws1[value_cell] = value
        ws1[value_cell].fill = param_fill
        ws1[value_cell].border = thin_border
        if isinstance(value, float) and value < 1:
            ws1[value_cell].number_format = percent_format
        else:
            ws1[value_cell].number_format = money_format
    
    # Charges patronales
    ws1['A9'] = "CHARGES PATRONALES"
    ws1['A9'].font = header_style
    
    params_charges = [
        ('A10', 'Versement Forfaitaire (VF):', 'B10', 0.06),
        ('A11', 'Taxe Apprentissage (TA):', 'B11', 0.015),
    ]
    
    for label_cell, label, value_cell, value in params_charges:
        ws1[label_cell] = label
        ws1[value_cell] = value
        ws1[value_cell].fill = param_fill
        ws1[value_cell].border = thin_border
        ws1[value_cell].number_format = percent_format
    
    # Barème RTS
    ws1['A13'] = "BARÈME RTS (IRG) - CGI 2022"
    ws1['A13'].font = header_style
    
    ws1['A14'] = "Tranche"
    ws1['B14'] = "De"
    ws1['C14'] = "À"
    ws1['D14'] = "Taux"
    for cell in ['A14', 'B14', 'C14', 'D14']:
        ws1[cell].font = header_style
        ws1[cell].border = thin_border
    
    tranches_rts = [
        (1, 0, 1000000, 0),
        (2, 1000001, 3000000, 0.05),
        (3, 3000001, 5000000, 0.08),
        (4, 5000001, 10000000, 0.10),
        (5, 10000001, 20000000, 0.15),
        (6, 20000001, None, 0.20),
    ]
    
    for i, (num, de, a, taux) in enumerate(tranches_rts):
        row = 15 + i
        ws1[f'A{row}'] = f"Tranche {num}"
        ws1[f'B{row}'] = de
        ws1[f'C{row}'] = a if a else "Illimité"
        ws1[f'D{row}'] = taux
        ws1[f'B{row}'].number_format = money_format
        if a:
            ws1[f'C{row}'].number_format = money_format
        ws1[f'D{row}'].number_format = percent_format
        for col in ['A', 'B', 'C', 'D']:
            ws1[f'{col}{row}'].border = thin_border
    
    # Ajuster largeurs
    ws1.column_dimensions['A'].width = 30
    ws1.column_dimensions['B'].width = 18
    ws1.column_dimensions['C'].width = 18
    ws1.column_dimensions['D'].width = 12
    
    # ========== FEUILLE 2: SAISIE SALARIÉ ==========
    ws2 = wb.create_sheet("Saisie Salarié")
    
    ws2.merge_cells('A1:D1')
    ws2['A1'] = "SAISIE DES DONNÉES SALARIÉ"
    ws2['A1'].font = titre_style
    ws2['A1'].fill = header_fill
    ws2['A1'].alignment = Alignment(horizontal='center')
    
    # Identification
    ws2['A3'] = "IDENTIFICATION"
    ws2['A3'].font = header_style
    
    saisie_id = [
        ('A4', 'Nom:', 'B4', 'DIALLO'),
        ('A5', 'Prénom:', 'B5', 'Mamadou'),
        ('A6', 'Matricule:', 'B6', 'EMP001'),
        ('A7', 'Poste:', 'B7', 'Comptable Principal'),
        ('A8', 'Mois/Année:', 'B8', 'Janvier 2026'),
    ]
    
    for label_cell, label, value_cell, value in saisie_id:
        ws2[label_cell] = label
        ws2[value_cell] = value
        ws2[value_cell].fill = param_fill
        ws2[value_cell].border = thin_border
    
    # Rémunération
    ws2['A10'] = "RÉMUNÉRATION"
    ws2['A10'].font = header_style
    
    saisie_remun = [
        ('A11', 'Salaire de base:', 'B11', 2800000),
        ('A12', 'Prime de transport:', 'B12', 800000),
        ('A13', 'Prime de logement:', 'B13', 0),
        ('A14', 'Autres primes:', 'B14', 0),
    ]
    
    for label_cell, label, value_cell, value in saisie_remun:
        ws2[label_cell] = label
        ws2[value_cell] = value
        ws2[value_cell].fill = param_fill
        ws2[value_cell].border = thin_border
        ws2[value_cell].number_format = money_format
    
    # Retenues
    ws2['A16'] = "RETENUES OPTIONNELLES"
    ws2['A16'].font = header_style
    
    saisie_retenues = [
        ('A17', 'Avance sur salaire:', 'B17', 160000),
        ('A18', 'Prêt (mensualité):', 'B18', 0),
        ('A19', 'Autres retenues:', 'B19', 0),
    ]
    
    for label_cell, label, value_cell, value in saisie_retenues:
        ws2[label_cell] = label
        ws2[value_cell] = value
        ws2[value_cell].fill = param_fill
        ws2[value_cell].border = thin_border
        ws2[value_cell].number_format = money_format
    
    # Ajuster largeurs
    ws2.column_dimensions['A'].width = 25
    ws2.column_dimensions['B'].width = 20
    
    # ========== FEUILLE 3: CALCUL AUTOMATIQUE ==========
    ws3 = wb.create_sheet("Calcul Automatique")
    
    ws3.merge_cells('A1:D1')
    ws3['A1'] = "CALCUL AUTOMATIQUE DE PAIE"
    ws3['A1'].font = titre_style
    ws3['A1'].fill = header_fill
    ws3['A1'].alignment = Alignment(horizontal='center')
    
    # Formules de calcul
    ws3['A3'] = "SALAIRE BRUT"
    ws3['A3'].font = header_style
    
    ws3['A4'] = "Salaire de base"
    ws3['B4'] = "='Saisie Salarié'!B11"
    ws3['A5'] = "Prime transport"
    ws3['B5'] = "='Saisie Salarié'!B12"
    ws3['A6'] = "Prime logement"
    ws3['B6'] = "='Saisie Salarié'!B13"
    ws3['A7'] = "Autres primes"
    ws3['B7'] = "='Saisie Salarié'!B14"
    ws3['A8'] = "TOTAL BRUT"
    ws3['A8'].font = header_style
    ws3['B8'] = "=SUM(B4:B7)"
    ws3['B8'].font = header_style
    
    for row in range(4, 9):
        ws3[f'B{row}'].number_format = money_format
        ws3[f'B{row}'].border = thin_border
    
    # CNSS
    ws3['A10'] = "COTISATION CNSS"
    ws3['A10'].font = header_style
    
    ws3['A11'] = "Assiette CNSS"
    ws3['B11'] = "=MAX(MIN(B8,Paramètres!B5),Paramètres!B4)"
    ws3['C11'] = "(plafonnée)"
    
    ws3['A12'] = "CNSS Salarié (5%)"
    ws3['B12'] = "=ROUND(B11*Paramètres!B6,0)"
    
    ws3['A13'] = "CNSS Employeur (18%)"
    ws3['B13'] = "=ROUND(B11*Paramètres!B7,0)"
    
    for row in range(11, 14):
        ws3[f'B{row}'].number_format = money_format
        ws3[f'B{row}'].border = thin_border
    
    # IRG/RTS
    ws3['A15'] = "IMPÔT SUR LE REVENU (IRG/RTS)"
    ws3['A15'].font = header_style
    
    ws3['A16'] = "Base imposable"
    ws3['B16'] = "=B8-B12"
    ws3['C16'] = "(Brut - CNSS)"
    
    # Calcul IRG par tranches
    ws3['A17'] = "Tranche 1 (0%)"
    ws3['B17'] = "=IF(B16>1000000,0,0)"
    ws3['A18'] = "Tranche 2 (5%)"
    ws3['B18'] = "=IF(B16>1000000,ROUND(MIN(B16,3000000)-1000000,0)*0.05,0)"
    ws3['A19'] = "Tranche 3 (8%)"
    ws3['B19'] = "=IF(B16>3000000,ROUND(MIN(B16,5000000)-3000000,0)*0.08,0)"
    ws3['A20'] = "Tranche 4 (10%)"
    ws3['B20'] = "=IF(B16>5000000,ROUND(MIN(B16,10000000)-5000000,0)*0.1,0)"
    ws3['A21'] = "Tranche 5 (15%)"
    ws3['B21'] = "=IF(B16>10000000,ROUND(MIN(B16,20000000)-10000000,0)*0.15,0)"
    ws3['A22'] = "Tranche 6 (20%)"
    ws3['B22'] = "=IF(B16>20000000,ROUND(B16-20000000,0)*0.2,0)"
    
    ws3['A23'] = "TOTAL IRG"
    ws3['A23'].font = header_style
    ws3['B23'] = "=SUM(B17:B22)"
    ws3['B23'].font = header_style
    
    for row in range(16, 24):
        ws3[f'B{row}'].number_format = money_format
        ws3[f'B{row}'].border = thin_border
    
    # Retenues
    ws3['A25'] = "TOTAL RETENUES"
    ws3['A25'].font = header_style
    
    ws3['A26'] = "CNSS Salarié"
    ws3['B26'] = "=B12"
    ws3['A27'] = "IRG"
    ws3['B27'] = "=B23"
    ws3['A28'] = "Avance sur salaire"
    ws3['B28'] = "='Saisie Salarié'!B17"
    ws3['A29'] = "Prêt"
    ws3['B29'] = "='Saisie Salarié'!B18"
    ws3['A30'] = "Autres retenues"
    ws3['B30'] = "='Saisie Salarié'!B19"
    ws3['A31'] = "TOTAL RETENUES"
    ws3['A31'].font = header_style
    ws3['B31'] = "=SUM(B26:B30)"
    ws3['B31'].font = header_style
    
    for row in range(26, 32):
        ws3[f'B{row}'].number_format = money_format
        ws3[f'B{row}'].border = thin_border
    
    # Net à payer
    ws3['A33'] = "NET À PAYER"
    ws3['A33'].font = Font(bold=True, size=14, color="006400")
    ws3['B33'] = "=B8-B31"
    ws3['B33'].font = Font(bold=True, size=14, color="006400")
    ws3['B33'].number_format = money_format
    ws3['B33'].border = thin_border
    ws3['B33'].fill = ok_fill
    
    # Charges patronales
    ws3['A35'] = "CHARGES PATRONALES"
    ws3['A35'].font = header_style
    
    ws3['A36'] = "CNSS Employeur (18%)"
    ws3['B36'] = "=B13"
    ws3['A37'] = "VF (6%)"
    ws3['B37'] = "=ROUND(B8*Paramètres!B10,0)"
    ws3['A38'] = "TA (1,5%)"
    ws3['B38'] = "=ROUND(B8*Paramètres!B11,0)"
    ws3['A39'] = "TOTAL CHARGES PATRONALES"
    ws3['A39'].font = header_style
    ws3['B39'] = "=SUM(B36:B38)"
    ws3['B39'].font = header_style
    
    for row in range(36, 40):
        ws3[f'B{row}'].number_format = money_format
        ws3[f'B{row}'].border = thin_border
    
    # Coût total
    ws3['A41'] = "COÛT TOTAL EMPLOYEUR"
    ws3['A41'].font = Font(bold=True, size=12)
    ws3['B41'] = "=B8+B39"
    ws3['B41'].font = Font(bold=True, size=12)
    ws3['B41'].number_format = money_format
    ws3['B41'].border = thin_border
    
    # Ajuster largeurs
    ws3.column_dimensions['A'].width = 28
    ws3.column_dimensions['B'].width = 20
    ws3.column_dimensions['C'].width = 15
    
    # ========== FEUILLE 4: COMPARAISON ==========
    ws4 = wb.create_sheet("Comparaison")
    
    ws4.merge_cells('A1:E1')
    ws4['A1'] = "COMPARAISON EXCEL vs APPLICATION"
    ws4['A1'].font = titre_style
    ws4['A1'].fill = header_fill
    ws4['A1'].alignment = Alignment(horizontal='center')
    
    # En-têtes
    headers = ['Élément', 'Excel', 'Application', 'Écart', 'Statut']
    for col, header in enumerate(headers, 1):
        cell = ws4.cell(row=3, column=col, value=header)
        cell.font = header_style
        cell.border = thin_border
        cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    
    # Lignes de comparaison
    comparaisons = [
        ('Salaire brut', "='Calcul Automatique'!B8"),
        ('CNSS Salarié', "='Calcul Automatique'!B12"),
        ('IRG', "='Calcul Automatique'!B23"),
        ('Avance', "='Calcul Automatique'!B28"),
        ('Total retenues', "='Calcul Automatique'!B31"),
        ('Net à payer', "='Calcul Automatique'!B33"),
        ('CNSS Employeur', "='Calcul Automatique'!B36"),
        ('VF', "='Calcul Automatique'!B37"),
        ('TA', "='Calcul Automatique'!B38"),
        ('Total charges patronales', "='Calcul Automatique'!B39"),
    ]
    
    for i, (elem, formule) in enumerate(comparaisons):
        row = 4 + i
        ws4[f'A{row}'] = elem
        ws4[f'B{row}'] = formule
        ws4[f'C{row}'] = 0  # À remplir manuellement
        ws4[f'C{row}'].fill = param_fill
        ws4[f'D{row}'] = f"=B{row}-C{row}"
        ws4[f'E{row}'] = f'=IF(ABS(D{row})<10,"✅ OK","❌ ÉCART")'
        
        for col in ['A', 'B', 'C', 'D', 'E']:
            ws4[f'{col}{row}'].border = thin_border
        ws4[f'B{row}'].number_format = money_format
        ws4[f'C{row}'].number_format = money_format
        ws4[f'D{row}'].number_format = money_format
    
    # Résumé
    ws4['A16'] = "RÉSULTAT GLOBAL"
    ws4['A16'].font = header_style
    ws4['B16'] = '=IF(COUNTIF(E4:E13,"❌ ÉCART")=0,"✅ TOUS LES CALCULS SONT CORRECTS","❌ DES ÉCARTS ONT ÉTÉ DÉTECTÉS")'
    ws4['B16'].font = Font(bold=True, size=12)
    
    # Instructions
    ws4['A18'] = "INSTRUCTIONS:"
    ws4['A18'].font = header_style
    ws4['A19'] = "1. Saisir les données du salarié dans l'onglet 'Saisie Salarié'"
    ws4['A20'] = "2. Reporter les valeurs de l'application dans la colonne 'Application' (C)"
    ws4['A21'] = "3. Vérifier que tous les statuts affichent '✅ OK'"
    ws4['A22'] = "4. En cas d'écart > 10 GNF, investiguer la différence"
    
    # Ajuster largeurs
    ws4.column_dimensions['A'].width = 25
    ws4.column_dimensions['B'].width = 18
    ws4.column_dimensions['C'].width = 18
    ws4.column_dimensions['D'].width = 15
    ws4.column_dimensions['E'].width = 12
    
    # Sauvegarder
    output_path = os.path.join(os.path.dirname(__file__), 'Verification_Paie_Guinee.xlsx')
    wb.save(output_path)
    print(f"✅ Fichier Excel créé: {output_path}")
    return output_path


if __name__ == "__main__":
    creer_excel_verification()
