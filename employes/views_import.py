"""
Module d'importation des employés et données RH
Supporte les formats Excel (.xlsx) et CSV (.csv)
"""
import os
import io
import csv
import traceback
from datetime import datetime, date

from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from gestionrh.decorators import reauth_required
from django.contrib import messages
from django.db import transaction
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from .models import Employe
from core.models import Etablissement, Service, Poste, Entreprise
from core.views import log_activity


# ============================================================
#  COLONNES DU TEMPLATE D'IMPORTATION
# ============================================================
IMPORT_COLUMNS = [
    # (nom_colonne, champ_model, obligatoire, description, exemple)
    # --- RENSEIGNEMENTS ---
    ('Matricule', 'matricule', False, 'Laisser vide pour génération automatique', 'EMP-001'),
    ('Civilité', 'civilite', False, 'M. / Mme / Mlle', 'M.'),
    ('Nom*', 'nom', True, 'Nom de famille (obligatoire)', 'DIALLO'),
    ('Prénoms*', 'prenoms', True, 'Prénoms (obligatoire)', 'Mamadou Alpha'),
    ('Sexe', 'sexe', False, 'M ou F', 'M'),
    ('N° CNSS', 'num_cnss_individuel', False, 'Numéro CNSS individuel', 'CNSS-2024-001'),
    ('N° Pièce identité', 'numero_piece_identite', False, 'Numéro du document', 'CNI-123456'),
    ('Statut', 'statut_employe', False, 'actif / suspendu / demissionnaire / licencie / retraite', 'actif'),
    # --- CONDUCTEUR / TRANSPORT ---
    ('ID Conducteur', 'id_conducteur', False, 'Identifiant conducteur', 'COND-001'),
    ('Tracteur', 'tracteur', False, 'Immatriculation ou nom tracteur', 'RC-1234-AB'),
    ('Citerne', 'citerne', False, 'Immatriculation ou nom citerne', 'CT-5678-CD'),
    ('N° Permis', 'numero_permis', False, 'Numéro du permis de conduire', 'PG-2024-12345'),
    ('Date obtention permis', 'date_obtention_permis', False, 'Format JJ/MM/AAAA', '10/05/2015'),
    ('Date validité permis', 'date_validite_permis', False, 'Format JJ/MM/AAAA', '10/05/2030'),
    ('Groupe sanguin', 'groupe_sanguin', False, 'A+ / A- / B+ / B- / AB+ / AB- / O+ / O-', 'O+'),
    ('Chauffeur basé à', 'base_chauffeur', False, 'Lieu d\'affectation du chauffeur', 'Conakry'),
    ('Camion/Voiture assigné(e)', 'vehicule_assigne', False, 'Véhicule assigné (immat. ou description)', 'RC-9999-AB / Toyota Hilux'),
    # --- IMMATRICULATIONS / ÉTAT CIVIL ---
    ('Date de naissance', 'date_naissance', False, 'Format JJ/MM/AAAA', '15/03/1990'),
    ('Lieu de naissance', 'lieu_naissance', False, 'Ville de naissance', 'Conakry'),
    ('Nationalité', 'nationalite', False, 'Par défaut : Guinéenne', 'Guinéenne'),
    ('Situation matrimoniale', 'situation_matrimoniale', False, 'celibataire / marie / divorce / veuf', 'celibataire'),
    ('Téléphone', 'telephone_principal', False, 'Numéro principal', '+224 620 00 00 00'),
    ('Email professionnel', 'email_professionnel', False, 'Adresse email pro', 'mdiallo@entreprise.gn'),
    ('Adresse', 'adresse_actuelle', False, 'Adresse de résidence', 'Kaloum, Conakry'),
    ('Date embauche', 'date_embauche', False, 'Format JJ/MM/AAAA', '01/01/2024'),
    # --- PROFESSIONNEL ---
    ('Établissement', 'etablissement', False, 'Nom de l\'établissement (doit exister)', 'Siège Conakry'),
    ('Service', 'service', False, 'Nom du service (doit exister)', 'Ressources Humaines'),
    ('Poste', 'poste', False, 'Intitulé du poste (doit exister)', 'Responsable RH'),
    ('Type contrat', 'type_contrat', False, 'CDI / CDD / Stage / etc.', 'CDI'),
    # --- FORMATION TRANSPORT / SÉCURITÉ ---
    ('Date formation APTH', 'date_formation_apth', False, 'Format JJ/MM/AAAA', '15/06/2023'),
    ('Ancienneté transport HCL (ans)', 'anciennete_transport_hcl', False, 'Nombre d\'années', '5'),
    ('Date dernier recyclage', 'date_dernier_recyclage', False, 'Format JJ/MM/AAAA', '20/01/2025'),
    ('Formation extincteur', 'formation_extincteur', False, 'OUI / NON', 'OUI'),
    # --- VISITE MÉDICALE ---
    ('Date visite médicale', 'date_derniere_visite_medicale', False, 'Format JJ/MM/AAAA', '10/02/2025'),
    ('Service médical accrédité', 'service_medical_accredite', False, 'Nom du centre médical', 'CMC Kaloum'),
    ('Prochaine visite médicale', 'date_prochaine_visite_medicale', False, 'Format JJ/MM/AAAA', '10/02/2026'),
    # --- FILIATION ---
    ('Nombre de femmes', 'nombre_femmes', False, 'Nombre entier', '1'),
    ('Nombre enfants', 'nombre_enfants', False, 'Nombre entier', '2'),
    ('Nom du père', 'nom_pere', False, 'Nom complet du père', 'DIALLO Ibrahima'),
    ('Nom de la mère', 'nom_mere', False, 'Nom complet de la mère', 'BAH Fatoumata'),
    # --- CONTACT D\'URGENCE ---
    ('Contact urgence - Nom', 'contact_urgence_nom', False, 'Nom et prénoms', 'CAMARA Moussa'),
    ('Contact urgence - Lien', 'contact_urgence_lien', False, 'Lien de parenté', 'Frère'),
    ('Contact urgence - Téléphone', 'contact_urgence_telephone', False, 'Numéro de téléphone', '+224 621 11 11 11'),
    # --- BANCAIRE ---
    ('Mode paiement', 'mode_paiement', False, 'virement / cheque / especes / mobile_money', 'virement'),
    ('Nom banque', 'nom_banque', False, 'Nom de la banque', 'BCRG'),
    ('N° Compte', 'numero_compte', False, 'Numéro de compte bancaire', '001-234567-89'),
    ('RIB', 'rib', False, 'Relevé d\'identité bancaire', 'GN12345678901234'),
]


def _generate_matricule(entreprise):
    """Génère un matricule automatique unique"""
    prefix = 'EMP'
    last = Employe.objects.filter(
        entreprise=entreprise,
        matricule__startswith=prefix
    ).order_by('-matricule').first()

    if last:
        try:
            num = int(last.matricule.replace(f'{prefix}-', '').replace(prefix, ''))
            return f'{prefix}-{num + 1:04d}'
        except (ValueError, AttributeError):
            pass

    count = Employe.objects.filter(entreprise=entreprise).count()
    return f'{prefix}-{count + 1:04d}'


def _parse_date(value):
    """Parse une date depuis différents formats"""
    if value is None or value == '':
        return None
    if isinstance(value, (date, datetime)):
        if isinstance(value, datetime):
            return value.date()
        return value

    value = str(value).strip()
    formats = ['%d/%m/%Y', '%Y-%m-%d', '%d-%m-%Y', '%d.%m.%Y', '%Y/%m/%d']
    for fmt in formats:
        try:
            return datetime.strptime(value, fmt).date()
        except ValueError:
            continue
    return None


def _parse_int(value, default=0):
    """Parse un entier"""
    if value is None or value == '':
        return default
    try:
        return int(float(str(value).strip()))
    except (ValueError, TypeError):
        return default


def _clean_str(value):
    """Nettoie une chaîne"""
    if value is None:
        return ''
    return str(value).strip()


# ============================================================
#  TÉLÉCHARGER LE TEMPLATE D'IMPORTATION
# ============================================================
@reauth_required
@login_required
def telecharger_template_import(request):
    """Génère et télécharge le template Excel d'importation des employés"""
    wb = Workbook()

    # ── Feuille 1 : Template principal ──
    ws = wb.active
    ws.title = "Employés à importer"

    # Styles
    header_font = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
    header_fill_required = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    header_fill_optional = PatternFill(start_color='2E75B6', end_color='2E75B6', fill_type='solid')
    desc_font = Font(name='Calibri', italic=True, color='808080', size=9)
    example_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Ligne 1 : En-têtes
    for col_idx, (col_name, _, required, _, _) in enumerate(IMPORT_COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill_required if required else header_fill_optional
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border

    # Ligne 2 : Descriptions
    for col_idx, (_, _, _, desc, _) in enumerate(IMPORT_COLUMNS, 1):
        cell = ws.cell(row=2, column=col_idx, value=desc)
        cell.font = desc_font
        cell.alignment = Alignment(wrap_text=True)
        cell.border = border

    # Ligne 3 : Exemple
    for col_idx, (_, _, _, _, example) in enumerate(IMPORT_COLUMNS, 1):
        cell = ws.cell(row=3, column=col_idx, value=example)
        cell.fill = example_fill
        cell.border = border

    # Largeur colonnes (auto pour toutes les colonnes)
    for col_idx, (col_name, _, _, desc, _) in enumerate(IMPORT_COLUMNS, 1):
        # Largeur basée sur le plus long entre nom de colonne et description
        w = max(len(col_name), len(desc) // 2, 14)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(w, 30)

    # Figer la première ligne
    ws.freeze_panes = 'A4'

    # ── Validations de données (dynamique par nom de colonne) ──
    col_index = {col[1]: i + 1 for i, col in enumerate(IMPORT_COLUMNS)}

    def _add_dv(field_name, formula, error_msg='Valeur invalide'):
        """Ajoute une validation dropdown sur la colonne du champ"""
        if field_name in col_index:
            letter = get_column_letter(col_index[field_name])
            dv = DataValidation(type='list', formula1=formula, allow_blank=True)
            dv.error = error_msg
            ws.add_data_validation(dv)
            dv.add(f'{letter}4:{letter}1000')

    _add_dv('civilite', '"M.,Mme,Mlle"', 'Choisir M., Mme ou Mlle')
    _add_dv('sexe', '"M,F"', 'Choisir M ou F')
    _add_dv('statut_employe', '"actif,suspendu,demissionnaire,licencie,retraite"', 'Statut invalide')
    _add_dv('situation_matrimoniale', '"celibataire,marie,divorce,veuf"', 'Situation invalide')
    _add_dv('type_contrat', '"CDI,CDD,CDImp,CTI,stage,apprentissage,temporaire"', 'Type contrat invalide')
    _add_dv('mode_paiement', '"virement,cheque,especes,mobile_money"', 'Mode paiement invalide')
    _add_dv('groupe_sanguin', '"A+,A-,B+,B-,AB+,AB-,O+,O-"', 'Groupe sanguin invalide')
    _add_dv('formation_extincteur', '"OUI,NON"', 'Choisir OUI ou NON')

    # ── Feuille 2 : Données de référence ──
    ws_ref = wb.create_sheet("Données de référence")
    ws_ref.sheet_properties.tabColor = "FFC000"

    ref_header_font = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
    ref_header_fill = PatternFill(start_color='E06C00', end_color='E06C00', fill_type='solid')

    # Charger les données de référence de l'entreprise
    entreprise = request.user.entreprise

    # Établissements
    ws_ref.cell(row=1, column=1, value="Établissements disponibles").font = ref_header_font
    ws_ref['A1'].fill = ref_header_fill
    etablissements = Etablissement.objects.filter(
        societe__entreprise=entreprise, actif=True
    ).values_list('nom_etablissement', flat=True)
    for i, nom in enumerate(etablissements, 2):
        ws_ref.cell(row=i, column=1, value=nom)

    # Services
    ws_ref.cell(row=1, column=3, value="Services disponibles").font = ref_header_font
    ws_ref['C1'].fill = ref_header_fill
    services = Service.objects.filter(
        entreprise=entreprise, actif=True
    ).values_list('nom_service', flat=True)
    for i, nom in enumerate(services, 2):
        ws_ref.cell(row=i, column=3, value=nom)

    # Postes
    ws_ref.cell(row=1, column=5, value="Postes disponibles").font = ref_header_font
    ws_ref['E1'].fill = ref_header_fill
    postes = Poste.objects.filter(
        entreprise=entreprise, actif=True
    ).values_list('intitule_poste', flat=True)
    for i, nom in enumerate(postes, 2):
        ws_ref.cell(row=i, column=5, value=nom)

    # Largeurs colonnes référence
    ws_ref.column_dimensions['A'].width = 30
    ws_ref.column_dimensions['C'].width = 30
    ws_ref.column_dimensions['E'].width = 30

    # ── Feuille 3 : Instructions ──
    ws_instr = wb.create_sheet("Instructions")
    ws_instr.sheet_properties.tabColor = "00B050"

    instructions = [
        ("INSTRUCTIONS D'IMPORTATION - GuineeRH", ""),
        ("", ""),
        ("RÈGLES GÉNÉRALES", ""),
        ("1.", "Les colonnes marquées avec * sont obligatoires (Nom, Prénoms)"),
        ("2.", "Commencez à saisir vos données à partir de la ligne 4 (après l'en-tête, la description et l'exemple)"),
        ("3.", "La ligne 3 (grisée) est un exemple - supprimez-la avant l'importation"),
        ("4.", "Le matricule est généré automatiquement si laissé vide"),
        ("5.", "Les dates doivent être au format JJ/MM/AAAA (ex: 15/03/1990)"),
        ("", ""),
        ("DONNÉES DE RÉFÉRENCE", ""),
        ("6.", "Les Établissements, Services et Postes doivent exister dans le système"),
        ("7.", "Consultez l'onglet 'Données de référence' pour les valeurs acceptées"),
        ("8.", "Si un établissement/service/poste n'existe pas, il sera ignoré (l'employé sera créé sans)"),
        ("", ""),
        ("VALEURS ACCEPTÉES", ""),
        ("Civilité:", "M. / Mme / Mlle"),
        ("Sexe:", "M / F"),
        ("Situation matrimoniale:", "celibataire / marie / divorce / veuf"),
        ("Type contrat:", "CDI / CDD / CDImp / CTI / stage / apprentissage / temporaire"),
        ("Mode paiement:", "virement / cheque / especes / mobile_money"),
        ("", ""),
        ("DOUBLONS", ""),
        ("9.", "Les matricules et N° CNSS doivent être uniques"),
        ("10.", "Si un matricule existe déjà, la ligne sera ignorée avec un avertissement"),
        ("", ""),
        ("LIMITES", ""),
        ("11.", "Maximum 500 employés par importation"),
        ("12.", "Les photos ne peuvent pas être importées via Excel (ajoutez-les manuellement)"),
    ]

    for row_idx, (col_a, col_b) in enumerate(instructions, 1):
        cell_a = ws_instr.cell(row=row_idx, column=1, value=col_a)
        ws_instr.cell(row=row_idx, column=2, value=col_b)
        if row_idx == 1:
            cell_a.font = Font(name='Calibri', bold=True, size=14, color='1F4E79')
        elif col_a in ("RÈGLES GÉNÉRALES", "DONNÉES DE RÉFÉRENCE", "VALEURS ACCEPTÉES", "DOUBLONS", "LIMITES"):
            cell_a.font = Font(name='Calibri', bold=True, size=12, color='E06C00')

    ws_instr.column_dimensions['A'].width = 35
    ws_instr.column_dimensions['B'].width = 70

    # Réponse HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=Template_Import_Employes_GuineeRH.xlsx'
    wb.save(response)

    log_activity(request, "Téléchargement template import employés", 'employes')
    return response


# ============================================================
#  PAGE D'IMPORTATION
# ============================================================
@reauth_required
@login_required
def import_employes_page(request):
    """Page d'importation des employés"""
    return render(request, 'employes/import.html')


# ============================================================
#  PRÉVISUALISATION AVANT IMPORT
# ============================================================
@reauth_required
@login_required
def import_employes_preview(request):
    """Analyse le fichier uploadé et affiche un aperçu avant importation"""
    if request.method != 'POST' or 'fichier' not in request.FILES:
        messages.error(request, 'Veuillez sélectionner un fichier Excel (.xlsx) ou CSV (.csv)')
        return redirect('employes:import_page')

    fichier = request.FILES['fichier']
    nom_fichier = fichier.name.lower()

    if not (nom_fichier.endswith('.xlsx') or nom_fichier.endswith('.csv')):
        messages.error(request, 'Format non supporté. Utilisez .xlsx ou .csv')
        return redirect('employes:import_page')

    # Lire le fichier
    try:
        if nom_fichier.endswith('.xlsx'):
            rows = _read_xlsx(fichier)
        else:
            rows = _read_csv(fichier)
    except Exception as e:
        messages.error(request, f'Erreur de lecture du fichier : {str(e)}')
        return redirect('employes:import_page')

    if len(rows) == 0:
        messages.warning(request, 'Le fichier est vide ou ne contient aucune donnée.')
        return redirect('employes:import_page')

    if len(rows) > 500:
        messages.error(request, f'Le fichier contient {len(rows)} lignes. Maximum autorisé : 500.')
        return redirect('employes:import_page')

    entreprise = request.user.entreprise

    # Préparer les caches de référence
    etabs_cache = {e.nom_etablissement.lower(): e for e in Etablissement.objects.filter(
        societe__entreprise=entreprise, actif=True
    )}
    services_cache = {s.nom_service.lower(): s for s in Service.objects.filter(
        entreprise=entreprise, actif=True
    )}
    postes_cache = {p.intitule_poste.lower(): p for p in Poste.objects.filter(
        entreprise=entreprise, actif=True
    )}

    # Matricules et CNSS existants
    matricules_existants = set(Employe.objects.filter(
        entreprise=entreprise
    ).values_list('matricule', flat=True))
    cnss_existants = set(Employe.objects.filter(
        entreprise=entreprise, num_cnss_individuel__isnull=False
    ).exclude(num_cnss_individuel='').values_list('num_cnss_individuel', flat=True))

    # Analyser chaque ligne
    preview_data = []
    total_ok = 0
    total_warn = 0
    total_err = 0
    matricules_fichier = set()
    cnss_fichier = set()

    for idx, row in enumerate(rows, 1):
        line = {'numero': idx, 'donnees': row, 'erreurs': [], 'avertissements': [], 'statut': 'ok'}

        nom = _clean_str(row.get('nom', row.get('Nom*', row.get('Nom', ''))))
        prenoms = _clean_str(row.get('prenoms', row.get('Prénoms*', row.get('Prénoms', ''))))
        matricule = _clean_str(row.get('matricule', row.get('Matricule', '')))
        cnss = _clean_str(row.get('num_cnss_individuel', row.get('N° CNSS', '')))

        # Validations obligatoires
        if not nom:
            line['erreurs'].append('Nom obligatoire')
        if not prenoms:
            line['erreurs'].append('Prénoms obligatoire')

        # Doublons matricule
        if matricule:
            if matricule in matricules_existants:
                line['erreurs'].append(f'Matricule "{matricule}" existe déjà')
            elif matricule in matricules_fichier:
                line['erreurs'].append(f'Matricule "{matricule}" en doublon dans le fichier')
            else:
                matricules_fichier.add(matricule)

        # Doublons CNSS
        if cnss:
            if cnss in cnss_existants:
                line['erreurs'].append(f'N° CNSS "{cnss}" existe déjà')
            elif cnss in cnss_fichier:
                line['erreurs'].append(f'N° CNSS "{cnss}" en doublon dans le fichier')
            else:
                cnss_fichier.add(cnss)

        # Validation dates
        date_naissance_str = row.get('date_naissance', row.get('Date de naissance', ''))
        if date_naissance_str and not _parse_date(date_naissance_str):
            line['avertissements'].append(f'Date naissance invalide : "{date_naissance_str}"')

        date_embauche_str = row.get('date_embauche', row.get('Date embauche', ''))
        if date_embauche_str and not _parse_date(date_embauche_str):
            line['avertissements'].append(f'Date embauche invalide : "{date_embauche_str}"')

        # Validations dates supplémentaires (conducteur, formation, visite)
        for field, label in [
            ('date_obtention_permis', 'Date obtention permis'),
            ('date_validite_permis', 'Date validité permis'),
            ('date_formation_apth', 'Date formation APTH'),
            ('date_dernier_recyclage', 'Date dernier recyclage'),
            ('date_derniere_visite_medicale', 'Date visite médicale'),
            ('date_prochaine_visite_medicale', 'Prochaine visite médicale'),
        ]:
            val = row.get(field, row.get(label, ''))
            if val and not _parse_date(val):
                line['avertissements'].append(f'{label} invalide : "{val}"')

        # Validation groupe sanguin
        gs = _clean_str(row.get('groupe_sanguin', row.get('Groupe sanguin', '')))
        if gs and gs not in ('A+', 'A-', 'B+', 'B-', 'AB+', 'AB-', 'O+', 'O-'):
            line['avertissements'].append(f'Groupe sanguin invalide : "{gs}"')

        # Validation références
        etab = _clean_str(row.get('etablissement', row.get('Établissement', '')))
        if etab and etab.lower() not in etabs_cache:
            line['avertissements'].append(f'Établissement "{etab}" introuvable')

        serv = _clean_str(row.get('service', row.get('Service', '')))
        if serv and serv.lower() not in services_cache:
            line['avertissements'].append(f'Service "{serv}" introuvable')

        poste = _clean_str(row.get('poste', row.get('Poste', '')))
        if poste and poste.lower() not in postes_cache:
            line['avertissements'].append(f'Poste "{poste}" introuvable')

        # Validation valeurs autorisées
        sexe = _clean_str(row.get('sexe', row.get('Sexe', '')))
        if sexe and sexe not in ('M', 'F'):
            line['avertissements'].append(f'Sexe invalide : "{sexe}" (attendu M ou F)')

        type_contrat = _clean_str(row.get('type_contrat', row.get('Type contrat', '')))
        types_valides = ('CDI', 'CDD', 'CDImp', 'CTI', 'stage', 'apprentissage', 'temporaire')
        if type_contrat and type_contrat not in types_valides:
            line['avertissements'].append(f'Type contrat invalide : "{type_contrat}"')

        # Statut final
        if line['erreurs']:
            line['statut'] = 'erreur'
            total_err += 1
        elif line['avertissements']:
            line['statut'] = 'avertissement'
            total_warn += 1
        else:
            total_ok += 1

        preview_data.append(line)

    # Sauvegarder en session pour l'import effectif
    request.session['import_data'] = rows
    request.session['import_filename'] = fichier.name

    context = {
        'preview_data': preview_data,
        'total_lignes': len(rows),
        'total_ok': total_ok,
        'total_warn': total_warn,
        'total_err': total_err,
        'nom_fichier': fichier.name,
        'colonnes': [c[0] for c in IMPORT_COLUMNS],
    }

    return render(request, 'employes/import_preview.html', context)


def _read_xlsx(fichier):
    """Lit un fichier Excel et retourne une liste de dictionnaires"""
    from openpyxl import load_workbook

    wb = load_workbook(fichier, read_only=True, data_only=True)
    ws = wb.active

    rows_iter = ws.iter_rows(values_only=True)
    headers_raw = next(rows_iter, None)
    if not headers_raw:
        return []

    # Nettoyer les en-têtes
    headers = [_clean_str(h) if h else f'col_{i}' for i, h in enumerate(headers_raw)]

    # Mapper les en-têtes du template vers les noms de champs
    header_map = {}
    for col_name, field_name, _, _, _ in IMPORT_COLUMNS:
        header_map[col_name.lower().rstrip('*')] = field_name
        header_map[field_name.lower()] = field_name

    # Ligne 2 = descriptions, Ligne 3 = exemple → à ignorer
    next(rows_iter, None)  # skip ligne 2 (descriptions)
    next(rows_iter, None)  # skip ligne 3 (exemple)

    rows = []
    for row_values in rows_iter:
        if not any(v for v in row_values):
            continue  # ligne vide

        row_dict = {}
        for i, val in enumerate(row_values):
            if i < len(headers):
                header_lower = headers[i].lower().rstrip('*')
                field = header_map.get(header_lower, headers[i])
                row_dict[field] = val
                # Garder aussi le header original pour le preview
                row_dict[headers[i]] = val

        rows.append(row_dict)

    wb.close()
    return rows


def _read_csv(fichier):
    """Lit un fichier CSV et retourne une liste de dictionnaires"""
    content = fichier.read().decode('utf-8-sig')
    reader = csv.DictReader(io.StringIO(content), delimiter=';')

    header_map = {}
    for col_name, field_name, _, _, _ in IMPORT_COLUMNS:
        header_map[col_name.lower().rstrip('*')] = field_name
        header_map[field_name.lower()] = field_name

    rows = []
    row_iter = iter(reader)

    # Si le CSV vient du template, les 2 premières lignes de données
    # sont les descriptions et l'exemple → les ignorer
    first_row = next(row_iter, None)
    if first_row:
        # Détecter si c'est la ligne de description du template
        first_vals = [str(v).strip().lower() for v in first_row.values() if v]
        is_template_desc = any(
            kw in ' '.join(first_vals)
            for kw in ['obligatoire', 'laisser vide', 'format jj/mm', 'numéro']
        )
        if is_template_desc:
            # C'est la ligne description → skip aussi la ligne exemple
            next(row_iter, None)
        else:
            # C'est une vraie donnée → la traiter
            row_dict = {}
            for key, val in first_row.items():
                if key:
                    key_lower = key.strip().lower().rstrip('*')
                    field = header_map.get(key_lower, key.strip())
                    row_dict[field] = val
                    row_dict[key.strip()] = val
            rows.append(row_dict)

    for row in row_iter:
        row_dict = {}
        for key, val in row.items():
            if key:
                key_lower = key.strip().lower().rstrip('*')
                field = header_map.get(key_lower, key.strip())
                row_dict[field] = val
                row_dict[key.strip()] = val
        rows.append(row_dict)

    return rows


# ============================================================
#  IMPORTATION EFFECTIVE
# ============================================================
@reauth_required
@login_required
def import_employes_execute(request):
    """Exécute l'importation après prévisualisation"""
    if request.method != 'POST':
        return redirect('employes:import_page')

    rows = request.session.get('import_data')
    if not rows:
        messages.error(request, 'Aucune donnée à importer. Veuillez recommencer.')
        return redirect('employes:import_page')

    entreprise = request.user.entreprise

    # Caches de référence
    etabs_cache = {e.nom_etablissement.lower(): e for e in Etablissement.objects.filter(
        societe__entreprise=entreprise, actif=True
    )}
    services_cache = {s.nom_service.lower(): s for s in Service.objects.filter(
        entreprise=entreprise, actif=True
    )}
    postes_cache = {p.intitule_poste.lower(): p for p in Poste.objects.filter(
        entreprise=entreprise, actif=True
    )}

    matricules_existants = set(Employe.objects.filter(
        entreprise=entreprise
    ).values_list('matricule', flat=True))

    cnss_existants = set(Employe.objects.filter(
        entreprise=entreprise, num_cnss_individuel__isnull=False
    ).exclude(num_cnss_individuel='').values_list('num_cnss_individuel', flat=True))

    created = 0
    errors = []
    warnings = []

    for idx, row in enumerate(rows, 1):
        nom = _clean_str(row.get('nom', row.get('Nom*', row.get('Nom', ''))))
        prenoms = _clean_str(row.get('prenoms', row.get('Prénoms*', row.get('Prénoms', ''))))

        if not nom or not prenoms:
            errors.append(f'Ligne {idx}: Nom ou Prénoms manquant - ignorée')
            continue

        matricule = _clean_str(row.get('matricule', row.get('Matricule', '')))
        cnss = _clean_str(row.get('num_cnss_individuel', row.get('N° CNSS', '')))

        # Vérifier doublons
        if matricule and matricule in matricules_existants:
            errors.append(f'Ligne {idx}: Matricule "{matricule}" existe déjà - ignorée')
            continue

        if cnss and cnss in cnss_existants:
            errors.append(f'Ligne {idx}: N° CNSS "{cnss}" existe déjà - ignorée')
            continue

        # Générer matricule si vide
        if not matricule:
            matricule = _generate_matricule(entreprise)
            # S'assurer de l'unicité
            while matricule in matricules_existants:
                try:
                    num = int(matricule.split('-')[-1])
                    matricule = f'EMP-{num + 1:04d}'
                except ValueError:
                    matricule = f'EMP-{created + 1:04d}'

        # Résoudre les références
        etab_nom = _clean_str(row.get('etablissement', row.get('Établissement', '')))
        etab = etabs_cache.get(etab_nom.lower()) if etab_nom else None

        serv_nom = _clean_str(row.get('service', row.get('Service', '')))
        serv = services_cache.get(serv_nom.lower()) if serv_nom else None

        poste_nom = _clean_str(row.get('poste', row.get('Poste', '')))
        poste = postes_cache.get(poste_nom.lower()) if poste_nom else None

        if etab_nom and not etab:
            warnings.append(f'Ligne {idx}: Établissement "{etab_nom}" introuvable')
        if serv_nom and not serv:
            warnings.append(f'Ligne {idx}: Service "{serv_nom}" introuvable')
        if poste_nom and not poste:
            warnings.append(f'Ligne {idx}: Poste "{poste_nom}" introuvable')

        try:
            with transaction.atomic():
                # Résoudre le booléen extincteur
                ext_val = _clean_str(row.get('formation_extincteur', row.get('Formation extincteur', '')))
                formation_ext = ext_val.upper() in ('OUI', 'YES', '1', 'TRUE', 'O') if ext_val else False

                # Résoudre le statut
                statut_val = _clean_str(row.get('statut_employe', row.get('Statut', '')))
                statut_valides = ['actif', 'suspendu', 'demissionnaire', 'licencie', 'retraite']
                statut_final = statut_val if statut_val in statut_valides else 'actif'

                employe = Employe(
                    entreprise=entreprise,
                    matricule=matricule,
                    nom=nom,
                    prenoms=prenoms,
                    civilite=_clean_str(row.get('civilite', row.get('Civilité', ''))) or None,
                    sexe=_clean_str(row.get('sexe', row.get('Sexe', ''))) or None,
                    date_naissance=_parse_date(row.get('date_naissance', row.get('Date de naissance', ''))),
                    lieu_naissance=_clean_str(row.get('lieu_naissance', row.get('Lieu de naissance', ''))) or None,
                    nationalite=_clean_str(row.get('nationalite', row.get('Nationalité', ''))) or 'Guinéenne',
                    situation_matrimoniale=_clean_str(row.get('situation_matrimoniale', row.get('Situation matrimoniale', ''))) or None,
                    nombre_enfants=_parse_int(row.get('nombre_enfants', row.get('Nombre enfants', 0))),
                    numero_piece_identite=_clean_str(row.get('numero_piece_identite', row.get('N° Pièce identité', ''))) or None,
                    num_cnss_individuel=cnss or None,
                    telephone_principal=_clean_str(row.get('telephone_principal', row.get('Téléphone', ''))) or None,
                    email_professionnel=_clean_str(row.get('email_professionnel', row.get('Email professionnel', ''))) or None,
                    adresse_actuelle=_clean_str(row.get('adresse_actuelle', row.get('Adresse', ''))) or None,
                    etablissement=etab,
                    service=serv,
                    poste=poste,
                    date_embauche=_parse_date(row.get('date_embauche', row.get('Date embauche', ''))),
                    type_contrat=_clean_str(row.get('type_contrat', row.get('Type contrat', ''))) or None,
                    statut_employe=statut_final,
                    # Conducteur / Transport
                    id_conducteur=_clean_str(row.get('id_conducteur', row.get('ID Conducteur', ''))) or None,
                    tracteur=_clean_str(row.get('tracteur', row.get('Tracteur', ''))) or None,
                    citerne=_clean_str(row.get('citerne', row.get('Citerne', ''))) or None,
                    numero_permis=_clean_str(row.get('numero_permis', row.get('N° Permis', ''))) or None,
                    date_obtention_permis=_parse_date(row.get('date_obtention_permis', row.get('Date obtention permis', ''))),
                    date_validite_permis=_parse_date(row.get('date_validite_permis', row.get('Date validité permis', ''))),
                    groupe_sanguin=_clean_str(row.get('groupe_sanguin', row.get('Groupe sanguin', ''))) or None,
                    base_chauffeur=_clean_str(row.get('base_chauffeur', row.get('Chauffeur basé à', ''))) or None,
                    vehicule_assigne=_clean_str(row.get('vehicule_assigne', row.get('Camion/Voiture assigné(e)', ''))) or None,
                    # Formation transport / sécurité
                    date_formation_apth=_parse_date(row.get('date_formation_apth', row.get('Date formation APTH', ''))),
                    anciennete_transport_hcl=_parse_int(row.get('anciennete_transport_hcl', row.get('Ancienneté transport HCL (ans)', ''))) or None,
                    date_dernier_recyclage=_parse_date(row.get('date_dernier_recyclage', row.get('Date dernier recyclage', ''))),
                    formation_extincteur=formation_ext,
                    # Visite médicale
                    date_derniere_visite_medicale=_parse_date(row.get('date_derniere_visite_medicale', row.get('Date visite médicale', ''))),
                    service_medical_accredite=_clean_str(row.get('service_medical_accredite', row.get('Service médical accrédité', ''))) or None,
                    date_prochaine_visite_medicale=_parse_date(row.get('date_prochaine_visite_medicale', row.get('Prochaine visite médicale', ''))),
                    # Filiation
                    nombre_femmes=_parse_int(row.get('nombre_femmes', row.get('Nombre de femmes', 0))),
                    nom_pere=_clean_str(row.get('nom_pere', row.get('Nom du père', ''))) or None,
                    nom_mere=_clean_str(row.get('nom_mere', row.get('Nom de la mère', ''))) or None,
                    # Contact d'urgence
                    contact_urgence_nom=_clean_str(row.get('contact_urgence_nom', row.get('Contact urgence - Nom', ''))) or None,
                    contact_urgence_lien=_clean_str(row.get('contact_urgence_lien', row.get('Contact urgence - Lien', ''))) or None,
                    contact_urgence_telephone=_clean_str(row.get('contact_urgence_telephone', row.get('Contact urgence - Téléphone', ''))) or None,
                    # Bancaire
                    mode_paiement=_clean_str(row.get('mode_paiement', row.get('Mode paiement', ''))) or 'virement',
                    nom_banque=_clean_str(row.get('nom_banque', row.get('Nom banque', ''))) or None,
                    numero_compte=_clean_str(row.get('numero_compte', row.get('N° Compte', ''))) or None,
                    rib=_clean_str(row.get('rib', row.get('RIB', ''))) or None,
                    utilisateur_creation=request.user,
                    utilisateur_modification=request.user,
                )
                employe.save()
                matricules_existants.add(matricule)
                if cnss:
                    cnss_existants.add(cnss)
                created += 1

        except Exception as e:
            errors.append(f'Ligne {idx} ({nom} {prenoms}): {str(e)}')

    # Nettoyer la session
    if 'import_data' in request.session:
        del request.session['import_data']
    if 'import_filename' in request.session:
        del request.session['import_filename']

    # Log
    log_activity(
        request,
        f"Import employés : {created} créés, {len(errors)} erreurs",
        'employes'
    )

    # Message de résultat
    if created > 0:
        messages.success(request, f'{created} employé(s) importé(s) avec succès !')

    for w in warnings[:10]:
        messages.warning(request, w)
    if len(warnings) > 10:
        messages.warning(request, f'... et {len(warnings) - 10} autres avertissements')

    for e in errors[:10]:
        messages.error(request, e)
    if len(errors) > 10:
        messages.error(request, f'... et {len(errors) - 10} autres erreurs')

    if created == 0 and errors:
        messages.error(request, "Aucun employé n'a été importé. Vérifiez votre fichier.")
        return redirect('employes:import_page')

    return redirect('employes:list')


# ============================================================
#  TEMPLATE D'IMPORTATION DES ÉLÉMENTS DE SALAIRE
# ============================================================
@reauth_required
@login_required
def telecharger_template_import_salaires(request):
    """Template pour l'import des éléments de salaire"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Salaires à importer"

    header_font = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    desc_font = Font(name='Calibri', italic=True, color='808080', size=9)
    example_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    salaire_cols = [
        ('Matricule*', 'Matricule de l\'employé (obligatoire)', 'EMP-001'),
        ('Salaire de base*', 'Montant du salaire de base', '2850000'),
        ('Prime transport', 'Prime de transport mensuelle', '300000'),
        ('Prime logement', 'Prime de logement mensuelle', '350000'),
        ('Autres primes', 'Autres primes (imposables)', '300000'),
        ('Indemnité non imposable', 'Indemnités exonérées', '0'),
    ]

    for col_idx, (name, desc, example) in enumerate(salaire_cols, 1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = border

        cell_desc = ws.cell(row=2, column=col_idx, value=desc)
        cell_desc.font = desc_font
        cell_desc.border = border

        cell_ex = ws.cell(row=3, column=col_idx, value=example)
        cell_ex.fill = example_fill
        cell_ex.border = border

    ws.freeze_panes = 'A4'
    for i in range(1, 7):
        ws.column_dimensions[get_column_letter(i)].width = 22

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=Template_Import_Salaires_GuineeRH.xlsx'
    wb.save(response)

    log_activity(request, "Téléchargement template import salaires", 'paie')
    return response
